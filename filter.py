import csv
import ast
import json
import os
import random
from collections.abc import Callable
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
import re
import string
import sys
import time
from pathlib import Path
from urllib.parse import parse_qsl, quote_plus, unquote, urlencode, urljoin, urlparse, urlunparse

import requests

from runtime_paths import app_dir

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)


BASE_DIR = app_dir()
ACTOR_ID = "aqPbs3KeH9aD8b22w"
# Collabs outside discovery: fallback khi actor Similarweb mặc định không trả traffic (tắt bằng COLLABS_OUTSIDE_SIMILARWEB_FALLBACK_ACTOR="").
DEFAULT_OUTSIDE_SIMILARWEB_FALLBACK_ACTOR = "radeance~similarweb-scraper"
MIN_VISITS = int(os.getenv("MIN_VISITS", "9000") or "9000")
TOP_KEYWORDS_COUNT = int(os.getenv("TOP_KEYWORDS_COUNT", "5") or "5")

# Mặc định cố định (không cấu hình qua UI/.env cho luồng chính)
DEFAULT_APIFY_MAX_DOMAINS_PER_RUN = 50
DEFAULT_UPPROMOTE_PAGE_DELAY_MS = 250
DEFAULT_GOAFFPRO_PAGE_DELAY_MS = 250
DEFAULT_REFERSION_PAGE_DELAY_MS = 250
DEFAULT_COLLABS_PAGE_DELAY_MS = 50
DEFAULT_OFFERS_PER_PAGE = 50
MAX_OFFERS_PER_PAGE = 50
MIN_OFFERS_PER_PAGE = 10
OFFERS_PER_PAGE_STEP = 10
DEFAULT_COLLABS_LIMIT = 12


def clamp_offers_per_page(raw) -> int:
    """Số offer/trang (Up) hoặc limit request (Go): bội số của 10, trong [10, 50]. Rỗng/sai → mặc định."""
    if raw is None:
        return DEFAULT_OFFERS_PER_PAGE
    s = str(raw).strip()
    if not s:
        return DEFAULT_OFFERS_PER_PAGE
    try:
        n = int(float(s))
    except (TypeError, ValueError):
        return DEFAULT_OFFERS_PER_PAGE
    n = max(MIN_OFFERS_PER_PAGE, min(MAX_OFFERS_PER_PAGE, n))
    n = (n // OFFERS_PER_PAGE_STEP) * OFFERS_PER_PAGE_STEP
    if n < MIN_OFFERS_PER_PAGE:
        n = MIN_OFFERS_PER_PAGE
    return n


def clamp_collabs_limit(raw) -> int:
    """Collabs cố định 12 brand/request (giữ hàm để tương thích call-site cũ)."""
    return DEFAULT_COLLABS_LIMIT


def enforce_fixed_fetch_defaults() -> None:
    """Apify 50 domain/lần; Uppromote & Goaffpro trễ 250ms/trang; không giới hạn số trang; offer/trang (Up & Go) theo .env/UI đã chuẩn hóa."""
    os.environ["APIFY_MAX_DOMAINS_PER_RUN"] = str(DEFAULT_APIFY_MAX_DOMAINS_PER_RUN)
    os.environ["UPPROMOTE_PAGE_DELAY_MS"] = str(DEFAULT_UPPROMOTE_PAGE_DELAY_MS)
    os.environ.pop("UPPROMOTE_MAX_PAGES", None)
    os.environ["GOAFFPRO_PAGE_DELAY_MS"] = str(DEFAULT_GOAFFPRO_PAGE_DELAY_MS)
    os.environ.pop("GOAFFPRO_MAX_PAGES", None)
    os.environ["REFERSION_PAGE_DELAY_MS"] = str(DEFAULT_REFERSION_PAGE_DELAY_MS)
    os.environ.pop("REFERSION_MAX_PAGES", None)
    os.environ["COLLABS_PAGE_DELAY_MS"] = str(DEFAULT_COLLABS_PAGE_DELAY_MS)
    os.environ.pop("COLLABS_MAX_PAGES", None)
    os.environ["UPPROMOTE_PER_PAGE"] = str(clamp_offers_per_page(os.getenv("UPPROMOTE_PER_PAGE")))
    os.environ["GOAFFPRO_LIMIT"] = str(clamp_offers_per_page(os.getenv("GOAFFPRO_LIMIT")))
    os.environ["COLLABS_LIMIT"] = str(clamp_collabs_limit(os.getenv("COLLABS_LIMIT")))


def uppromote_max_pages_cap() -> int | None:
    """None = không giới hạn trang. Chỉ dùng khi UPPROMOTE_MAX_PAGES được set thủ công (CLI / thử nghiệm)."""
    raw = (os.getenv("UPPROMOTE_MAX_PAGES") or "").strip().lower()
    if not raw or raw in ("0", "unlimited", "none", "no", "inf"):
        return None
    try:
        n = int(raw)
    except ValueError:
        return None
    return n if n > 0 else None


def goaffpro_max_pages_cap() -> int | None:
    """None = không giới hạn trang Goaffpro."""
    raw = (os.getenv("GOAFFPRO_MAX_PAGES") or "").strip().lower()
    if not raw or raw in ("0", "unlimited", "none", "no", "inf"):
        return None
    try:
        n = int(raw)
    except ValueError:
        return None
    return n if n > 0 else None


def refersion_max_pages_cap() -> int | None:
    """None = không giới hạn trang Refersion."""
    raw = (os.getenv("REFERSION_MAX_PAGES") or "").strip().lower()
    if not raw or raw in ("0", "unlimited", "none", "no", "inf"):
        return None
    try:
        n = int(raw)
    except ValueError:
        return None
    return n if n > 0 else None


def collabs_max_pages_cap() -> int | None:
    """None = không giới hạn trang Shopify Collabs."""
    raw = (os.getenv("COLLABS_MAX_PAGES") or "").strip().lower()
    if not raw or raw in ("0", "unlimited", "none", "no", "inf"):
        return None
    try:
        n = int(raw)
    except ValueError:
        return None
    return n if n > 0 else None


def _unescape_dotenv_double_quoted(inner: str) -> str:
    out: list[str] = []
    i = 0
    while i < len(inner):
        if inner[i] == "\\" and i + 1 < len(inner):
            n = inner[i + 1]
            if n == "n":
                out.append("\n")
            elif n == "r":
                out.append("\r")
            elif n == "t":
                out.append("\t")
            elif n in ('"', "\\"):
                out.append(n)
            else:
                out.append(inner[i])
                out.append(n)
            i += 2
            continue
        out.append(inner[i])
        i += 1
    return "".join(out)


def parse_env_file(path: Path) -> dict:
    """Đọc .env → dict. utf-8-sig (bỏ BOM); giá trị trong \"…\" được unescape chuẩn dotenv."""
    out: dict[str, str] = {}
    if not path.exists():
        return out
    text = path.read_text(encoding="utf-8-sig")
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, val = line.split("=", 1)
        k = key.strip()
        v = val.strip()
        if len(v) >= 2 and v[0] == '"' and v[-1] == '"':
            v = _unescape_dotenv_double_quoted(v[1:-1])
        elif len(v) >= 2 and v[0] == "'" and v[-1] == "'":
            v = v[1:-1]
        out[k] = v
    return out


def load_env_file(path: Path):
    for k, v in parse_env_file(path).items():
        if k and k not in os.environ:
            os.environ[k] = v


load_env_file(BASE_DIR / ".env")


def host_key(raw: str) -> str:
    if not raw:
        return ""
    value = str(raw).strip().lower()
    value = re.sub(r"^https?://", "", value, flags=re.I)
    value = value.split("/")[0]
    if value.startswith("www."):
        value = value[4:]
    return value


def parse_visits_value(raw) -> float:
    """Chuyển Visits từ Apify (số, chuỗi có dấu phẩy, đôi khi dạng 1.2M) sang float."""
    if raw is None:
        return 0.0
    if isinstance(raw, bool):
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        s = raw.replace(",", "").replace("\u00a0", " ").strip()
        if not s:
            return 0.0
        s_lower = s.lower().strip()
        for suf, mul in (("k", 1e3), ("m", 1e6), ("b", 1e9)):
            if s_lower.endswith(suf) and len(s_lower) > 1:
                try:
                    return float(s_lower[:-1].strip()) * mul
                except ValueError:
                    break
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def apify_site_field(item: dict) -> str:
    """Tên site/domain trong item Apify (nhiều actor đặt tên khác nhau)."""
    if not item:
        return ""
    # Chuẩn hoá map key lowercase để hỗ trợ actor trả key khác kiểu chữ.
    low = {str(k).strip().lower(): v for k, v in item.items()}
    return (
        item.get("SiteName")
        or item.get("siteName")
        or item.get("Domain")
        or item.get("domain")
        or item.get("Website")
        or item.get("website")
        or item.get("Url")
        or item.get("url")
        or item.get("Site")
        or item.get("site")
        or low.get("sitename")
        or low.get("domain")
        or low.get("website")
        or low.get("url")
        or low.get("site")
        or ""
    )


def engagement_from_item(item: dict) -> dict:
    """Khối engagement / hoặc Visits nằm ngang item."""
    if not item:
        return {}
    low = {str(k).strip().lower(): v for k, v in item.items()}
    eng = (
        item.get("Engagments")
        or item.get("Engagements")
        or item.get("engagement")
        or low.get("engagments")
        or low.get("engagements")
        or low.get("engagement")
    )
    if isinstance(eng, dict) and eng:
        return eng
    # Một số actor để traffic nằm ngang item với key khác casing.
    if (
        item.get("Visits") is not None
        or item.get("VisitsFormatted") is not None
        or low.get("visits") is not None
        or low.get("visitsformatted") is not None
        or low.get("estimatedmonthlyvisits") is not None
        or low.get("monthlyvisits") is not None
    ):
        return item
    return {}


def parse_visits_from_engagement(eng: dict) -> float:
    if not eng:
        return 0.0
    low = {str(k).strip().lower(): v for k, v in eng.items()} if isinstance(eng, dict) else {}
    raw = eng.get("Visits") if isinstance(eng, dict) else None
    if raw is None:
        raw = (
            (eng.get("EstimatedVisits") if isinstance(eng, dict) else None)
            or (eng.get("Traffic") if isinstance(eng, dict) else None)
            or (eng.get("MonthlyVisits") if isinstance(eng, dict) else None)
            or low.get("visits")
            or low.get("estimatedvisits")
            or low.get("traffic")
            or low.get("monthlyvisits")
            or low.get("estimatedmonthlyvisits")
        )
    return parse_visits_value(raw)


def estimated_monthly_visits_formatted(item: dict, eng: dict | None = None) -> str:
    """Chuỗi traffic theo tháng từ Apify, ưu tiên field gốc trên item."""
    src_eng = eng if isinstance(eng, dict) else {}
    item_low = {str(k).strip().lower(): v for k, v in (item or {}).items()}
    eng_low = {str(k).strip().lower(): v for k, v in src_eng.items()}
    candidates = (
        (item or {}).get("EstimatedMonthlyVisitsFormatted"),
        src_eng.get("EstimatedMonthlyVisitsFormatted"),
        (item or {}).get("EstimatedMonthlyVisits"),
        src_eng.get("EstimatedMonthlyVisits"),
        item_low.get("estimatedmonthlyvisitsformatted"),
        eng_low.get("estimatedmonthlyvisitsformatted"),
        item_low.get("estimatedmonthlyvisits"),
        eng_low.get("estimatedmonthlyvisits"),
    )
    for v in candidates:
        if v is not None and str(v).strip():
            return format_estimated_monthly_visits(v)
    return ""


def visits_formatted_from_engagement(eng: dict | None) -> str:
    if not isinstance(eng, dict) or not eng:
        return ""
    low = {str(k).strip().lower(): v for k, v in eng.items()}
    raw = eng.get("VisitsFormatted") or low.get("visitsformatted")
    if raw is not None and str(raw).strip():
        return str(raw).strip()
    visits = parse_visits_from_engagement(eng)
    if visits <= 0:
        return ""
    if float(visits).is_integer():
        return str(int(visits))
    return f"{visits:.2f}".rstrip("0").rstrip(".")


def format_estimated_monthly_visits(raw) -> str:
    """
    Chuẩn hóa traffic theo tháng:
    {'2026-01-01': '577', '2026-02-01': '226'} -> T1(577), T2(226)
    """
    if raw is None:
        return ""
    obj = raw
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return ""
        obj = text
        if text.startswith("{") and text.endswith("}"):
            try:
                obj = json.loads(text)
            except Exception:
                try:
                    obj = ast.literal_eval(text)
                except Exception:
                    obj = text
    if isinstance(obj, dict):
        parts = []
        for k in sorted(obj.keys(), key=lambda x: str(x)):
            key_text = str(k)
            month_match = re.search(r"-(\d{2})-", key_text)
            month_label = f"T{int(month_match.group(1))}" if month_match else key_text
            v = obj.get(k)
            value_text = "" if v is None else str(v).strip()
            parts.append(f"{month_label}({value_text})")
        return ", ".join(parts)
    return str(obj).strip()


def _highlight_max_monthly_traffic(cell, raw_text: str) -> None:
    """
    Tô đỏ phần tháng có traffic lớn nhất trong chuỗi:
    T1(577), T2(226), T3(314)
    """
    text = (raw_text or "").strip()
    if not text:
        return
    matches = list(re.finditer(r"(T\d+\()([^)]+)(\))", text))
    if not matches:
        return

    parsed = []
    for m in matches:
        num_txt = (m.group(2) or "").replace(",", "").strip()
        try:
            val = float(num_txt)
        except ValueError:
            continue
        parsed.append((m.span(), val))
    if not parsed:
        return
    max_val = max(v for _, v in parsed)
    max_spans = [span for span, v in parsed if v == max_val]
    if not max_spans:
        return

    from openpyxl.styles import Font
    # Dùng cách an toàn cho mọi bản openpyxl: tô đỏ toàn bộ ô khi có max.
    # (Tránh rich text gây lỗi save trên một số môi trường.)
    if max_spans:
        cell.font = Font(color="FF0000")


def lookup_apify_item(url: str, by_host: dict) -> dict:
    """Ghép offer URL với bản ghi Apify (khớp host + quét fallback)."""
    if not by_host:
        return {}
    k = host_key(url)
    if not k:
        return {}
    if k in by_host:
        return by_host[k]
    for item in by_host.values():
        sk = host_key(apify_site_field(item))
        if sk and sk == k:
            return item
    return {}


def normalize_bearer(raw: str) -> str:
    if raw is None:
        return ""
    token = str(raw).strip().replace("\r", "")
    if (token.startswith('"') and token.endswith('"')) or (token.startswith("'") and token.endswith("'")):
        token = token[1:-1].strip()
    token = re.sub(r"^Bearer\s+", "", token, flags=re.I)
    return token


def build_uppromote_headers() -> dict:
    token = normalize_bearer(os.getenv("UPPROMOTE_BEARER_TOKEN", ""))
    if not token:
        raise RuntimeError("Thiếu UPPROMOTE_BEARER_TOKEN trong .env")
    return {
        "accept": "application/json",
        "authorization": f"Bearer {token}",
        "user-agent": os.getenv(
            "UPPROMOTE_USER_AGENT",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
        ),
    }


def build_goaffpro_headers() -> dict:
    token = normalize_bearer(os.getenv("GOAFFPRO_BEARER_TOKEN", ""))
    if not token:
        raise RuntimeError("Thiếu GOAFFPRO_BEARER_TOKEN trong .env")
    return {
        "accept": "application/json",
        "authorization": f"Bearer {token}",
        "user-agent": os.getenv(
            "GOAFFPRO_USER_AGENT",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
        ),
    }


def build_refersion_headers() -> dict:
    token = normalize_bearer(os.getenv("REFERSION_TOKEN", ""))
    if not token:
        raise RuntimeError("Thiếu REFERSION_TOKEN trong .env")
    return {
        "accept": "application/json",
        "origin": os.getenv("REFERSION_ORIGIN", "https://marketplace.refersion.com"),
        "referer": os.getenv("REFERSION_REFERER", "https://marketplace.refersion.com/"),
        "refersion-token": token,
        "user-agent": os.getenv(
            "REFERSION_USER_AGENT",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
        ),
    }


def build_collabs_headers() -> dict:
    cookie = (os.getenv("COLLABS_COOKIE") or "").strip()
    csrf = (os.getenv("COLLABS_CSRF_TOKEN") or "").strip()
    if not cookie:
        raise RuntimeError("Thiếu COLLABS_COOKIE trong .env")
    if not csrf:
        raise RuntimeError("Thiếu COLLABS_CSRF_TOKEN trong .env")
    return {
        "accept": "*/*",
        "content-type": "application/json",
        "cookie": cookie,
        "origin": os.getenv("COLLABS_ORIGIN", "https://collabs.shopify.com"),
        "referer": os.getenv("COLLABS_REFERER", "https://collabs.shopify.com/"),
        "x-client-type": os.getenv("COLLABS_CLIENT_TYPE", "web"),
        "x-csrf-token": csrf,
        "user-agent": os.getenv(
            "COLLABS_USER_AGENT",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
        ),
    }


COLLABS_BRANDS_QUERY = (
    "query BrandsQuery($first: Int, $last: Int, $after: String, $before: String, "
    "$brandValues: [BrandValue!], $categories: [ProductCategory!], "
    "$productCategories: [CreatorProductCategory!], $saved: Boolean, $searchQuery: String) {"
    " socialAccounts { id __typename }"
    " brandsNetworkSearch("
    "   first: $first"
    "   last: $last"
    "   after: $after"
    "   before: $before"
    "   brandValues: $brandValues"
    "   categories: $categories"
    "   productCategories: $productCategories"
    "   saved: $saved"
    "   searchQuery: $searchQuery"
    " ) {"
    "   totalCount"
    "   pageInfo { hasNextPage hasPreviousPage endCursor startCursor __typename }"
    "   nodes {"
    "     id"
    "     ... on BrandInterface {"
    "       name"
    "       logoUrl"
    "       images"
    "       backgroundColor"
    "       shopifyStore { id shopifyStoreId __typename }"
    "       saved"
    "       partnershipStatus"
    "       productCategory"
    "       networkCommissionRange"
    "       partnershipState"
    "       targetCountries"
    "       previouslyPurchased"
    "       __typename"
    "     }"
    "     __typename"
    "   }"
    "   __typename"
    " }"
    "}"
)

COLLABS_BRAND_PROFILE_QUERY = (
    "query DiscoverBrandProfileQuery($shopifyStoreId: GID!) {"
    " socialAccounts { id __typename }"
    " brand(shopifyStoreId: $shopifyStoreId) {"
    "   id"
    "   ... on BrandInterface {"
    "     name"
    "     holdingPeriod"
    "     socialLinks { platform url __typename }"
    "     shopifyStore { id shopifyStoreId __typename }"
    "     __typename"
    "   }"
    "   __typename"
    " }"
    " creator { submittedApplications submittedApplicationsLimit __typename }"
    "}"
)

COLLABS_CATEGORY_LABELS = {
    "CLOTHING_AND_ACCESSORIES": "Clothing and accessories",
    "WOMENS_CLOTHING": "Women's clothing and accessories",
    "MENS_CLOTHING": "Men's clothing and accessories",
    "BEAUTY": "Beauty",
    "HEALTH_AND_WELLNESS": "Health and wellness",
    "FOOD_AND_DRINK": "Food and drink",
    "HOME_GOODS_AND_DECOR": "Home goods and decor",
    "BABY_AND_TODDLER": "Baby and toddler",
    "ELECTRONICS": "Electronics",
    "SPORTS_GOODS": "Sports goods",
    "ARTS_AND_CRAFTS": "Arts and crafts",
    "TECH": "Tech",
    "PHOTOGRAPHY": "Photography",
    "PET_SUPPLIES_AND_ACCESSORIES": "Pet supplies and accessories",
    "DIY": "DIY",
    "AUTOMOTIVE": "Automotive",
    "GARDENING": "Gardening",
    "TOBACCO_AND_VAPE": "Tobacco and vape",
    "MATURE": "Mature",
    "MUSICAL_INSTRUMENTS_AND_ACCESSORIES": "Musical instruments and accessories",
}


def collabs_category_label(raw_code: str) -> str:
    code = str(raw_code or "").strip().upper()
    if not code:
        return ""
    return COLLABS_CATEGORY_LABELS.get(code, code)


def format_collabs_holding_period(raw) -> str:
    """Chuẩn hóa holding period: số -> thêm ' day'."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    low = s.lower()
    if "day" in low:
        return s
    try:
        n = float(s.replace(",", ""))
    except ValueError:
        return s
    if abs(n - round(n)) < 1e-9:
        return f"{int(round(n))} day"
    return f"{n:g} day"


def with_page(url: str, page: int) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["page"] = str(page)
    # Always force per_page from env for consistent paging.
    query["per_page"] = os.getenv("UPPROMOTE_PER_PAGE", str(DEFAULT_OFFERS_PER_PAGE))
    new_query = urlencode(query, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def offer_url_from_uppromote(offer: dict) -> str:
    website = offer.get("website")
    if isinstance(website, str) and website.strip():
        return website.strip()
    domain = offer.get("myshopify_domain")
    if isinstance(domain, str) and domain.strip():
        return f"https://{domain.strip()}"
    apply_url = offer.get("apply_url")
    if isinstance(apply_url, str) and apply_url.strip():
        return apply_url.strip()
    return ""


def fetch_uppromote_offer_detail(shop_id) -> dict:
    base = (os.getenv("UPPROMOTE_DETAIL_BASE_URL") or "https://mkp-api.uppromote.com").strip().rstrip("/")
    url = f"{base}/api/v1/marketplace-offer/offer-detail/{shop_id}?mobile=false"
    res = requests.get(url, headers=build_uppromote_headers(), timeout=60)
    text = res.text
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError(f"Uppromote detail parse JSON lỗi (HTTP {res.status_code}): {text[:180]}") from exc
    if not res.ok:
        raise RuntimeError(f"Uppromote detail HTTP {res.status_code}: {text[:300]}")
    if body.get("status") not in (200, "200"):
        raise RuntimeError(f"Uppromote detail API lỗi: {text[:300]}")
    return body.get("data") or {}


def with_goaffpro_paging(url: str, offset: int, limit: int) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["limit"] = str(limit)
    query["offset"] = str(offset)
    new_query = urlencode(query, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def _portal_string(store: dict) -> str:
    raw = (
        store.get("affiliatePortal")
        or store.get("affiliate_portal")
        or store.get("AffiliatePortal")
        or ""
    )
    if raw is None:
        return ""
    if not isinstance(raw, str):
        raw = str(raw)
    return raw.strip()


def goaffpro_apply_url(store: dict) -> str:
    """Base URL https://{token}.goaffpro.com từ chuỗi affiliatePortal (mọi biến thể API)."""
    portal = _portal_string(store)
    if not portal:
        return ""
    # Đã là URL đầy đủ
    m = re.search(r"https?://([a-z0-9_-]+)\.goaffpro\.com/?", portal, re.I)
    if m:
        return f"https://{m.group(1).lower()}.goaffpro.com"
    # {goaffpro_public_token:xxx,...}.goaffpro.com
    m = re.search(r"goaffpro_public_token:\s*([^,}\s]+)", portal, re.I)
    if m:
        token = m.group(1).strip()
        if token:
            return f"https://{token}.goaffpro.com"
    # Token chỉ có chữ/số trước .goaffpro.com (một số response rút gọn)
    m = re.search(r"\b([a-z0-9_-]{2,64})\.goaffpro\.com\b", portal, re.I)
    if m:
        return f"https://{m.group(1).lower()}.goaffpro.com"
    return ""


def affiliate_portal_raw_from_offer(offer: dict) -> str:
    return (
        (offer.get("goaff_affiliate_portal") or "").strip()
        or (offer.get("affiliatePortal") or "").strip()
        or (offer.get("affiliate_portal") or "").strip()
    )


def goaff_create_account_url(offer: dict) -> str:
    """Chỉ từ affiliatePortal (API): https://sub.goaffpro.com/create-account — không dùng website."""
    portal_raw = affiliate_portal_raw_from_offer(offer)
    if not portal_raw:
        return ""
    base = goaffpro_apply_url({"affiliatePortal": portal_raw}).strip().rstrip("/")
    if not base:
        return ""
    if base.endswith("/create-account"):
        return base
    return f"{base}/create-account"


def format_goaff_commission_amount_display(offer: dict) -> str:
    """Số tiền HH kèm đơn vị: % hoặc tiền tệ ($, …) theo loại hoa hồng."""
    raw = offer.get("goaff_commission_amount")
    if raw is None or raw == "":
        return ""
    typ = (offer.get("goaff_commission_type") or "").strip().lower()
    cur = (offer.get("currency") or "").strip().upper()
    try:
        num = float(raw)
    except (TypeError, ValueError):
        return str(raw)
    if typ == "percentage" or "percent" in typ:
        return f"{num:g}%"
    if typ in ("fixed", "flat", "amount") or "fixed" in typ:
        if cur == "USD":
            return f"${num:g}"
        if cur == "EUR":
            return f"€{num:g}"
        if cur == "GBP":
            return f"£{num:g}"
        if cur == "VND":
            return f"{num:,.0f} ₫"
        if cur:
            return f"{num:g} {cur}"
        return f"${num:g}"
    if cur == "USD":
        return f"${num:g}"
    if cur:
        return f"{num:g} {cur}"
    return f"{num:g}"


def fmt_yes_no_01(val) -> str:
    if val in (1, "1", True):
        return "Có"
    if val in (0, "0", False):
        return "Không"
    return ""


def fmt_yes_no_bool_like(val) -> str:
    """Chuẩn hóa bool/0/1/"true"/"false" -> Có/Không cho cột trạng thái."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Có" if val else "Không"
    if isinstance(val, (int, float)):
        return "Có" if int(val) != 0 else "Không"
    s = str(val).strip().lower()
    if not s:
        return ""
    if s in ("1", "true", "yes", "y", "on"):
        return "Có"
    if s in ("0", "false", "no", "n", "off"):
        return "Không"
    return str(val)


def fmt_refersion_scalar(value) -> str:
    """Giữ dữ liệu Refersion dạng chuỗi gọn cho Excel (list/dict -> text)."""
    if value is None:
        return ""
    if isinstance(value, list):
        return join_list(value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def refersion_active_from_denied(denied) -> str:
    """Hoạt động = phủ định của denied."""
    if denied is None:
        return ""
    if isinstance(denied, bool):
        return "Có" if not denied else "Không"
    if isinstance(denied, (int, float)):
        return "Có" if int(denied) == 0 else "Không"
    s = str(denied).strip().lower()
    if not s:
        return ""
    if s in ("1", "true", "yes", "y", "on"):
        return "Không"
    if s in ("0", "false", "no", "n", "off"):
        return "Có"
    return ""


def cookie_days_from_goaffpro(raw) -> str | int | float:
    if raw is None:
        return ""
    try:
        v = float(raw)
    except Exception:
        return ""
    if v > 86400 * 2:
        days = v / 86400.0
        return int(days) if days == int(days) else round(days, 2)
    return int(v) if v == int(v) else v


def commission_str_goaffpro(comm: dict | None) -> str:
    if not isinstance(comm, dict):
        return ""
    typ = (comm.get("type") or "").lower()
    amount = comm.get("amount")
    on = (comm.get("on") or "").strip()
    if typ == "percentage" and amount is not None:
        base = f"{amount}%"
        return f"{base} ({on})" if on else base
    if amount is not None:
        extra = f" {typ}" if typ else ""
        on_part = f" on {on}" if on else ""
        return f"{amount}{extra}{on_part}".strip()
    return ""


def commission_type_goaffpro(comm: dict | None) -> str:
    if not isinstance(comm, dict):
        return ""
    parts = [str(comm.get("type") or "").strip(), str(comm.get("on") or "").strip()]
    return " ".join(p for p in parts if p)


def map_goaffpro_store(store: dict) -> dict:
    website = store.get("website")
    url = website.strip() if isinstance(website, str) and website.strip() else ""
    name_raw = store.get("name")
    brand = ""
    if isinstance(name_raw, str) and name_raw.strip() and not name_raw.strip().lower().startswith("http"):
        brand = name_raw.strip()
    elif url:
        hk = host_key(url)
        brand = hk.replace(".", " ").title() if hk else url
    comm = store.get("commission") if isinstance(store.get("commission"), dict) else {}
    app_auto = store.get("isApprovedAutomatically")
    if app_auto == 1:
        application_review = "auto"
    elif app_auto == 0:
        application_review = "manual"
    else:
        application_review = ""
    return {
        "brand": brand,
        "url": url,
        "offer": commission_str_goaffpro(comm),
        "cookieDays": cookie_days_from_goaffpro(store.get("cookieDuration")),
        "client_url": goaffpro_apply_url(store) or "",
        "offer_id": store.get("id") or "",
        "shop_id": store.get("id") or "",
        "program_id": "",
        "mkp_listing_id": store.get("id") or "",
        "commission_type": commission_type_goaffpro(comm),
        "category": "",
        "epc": "",
        "payments": "",
        "currency": store.get("currency") or "",
        "payout_rate": "",
        "approval_rate": "",
        "offer_score": "",
        "recommend_score": "",
        "application_review": application_review,
        "promotion_details": [],
        "target_audience_customer_channels": [],
        "target_audience_locations": [],
        "target_audience_ages": [],
        "target_audience_genders": [],
        "can_apply_offer": None,
        "is_applied_offer": None,
        # Raw Goaff API fields (CSV / snapshot)
        "goaff_id": store.get("id", ""),
        "goaff_name": (name_raw.strip() if isinstance(name_raw, str) else "") or "",
        "goaff_logo": store.get("logo") or "",
        "goaff_affiliate_portal": _portal_string(store),
        "goaff_cookie_duration_sec": store.get("cookieDuration", ""),
        "goaff_are_registrations_open": store.get("areRegistrationsOpen", ""),
        "goaff_is_approved_automatically": store.get("isApprovedAutomatically", ""),
        "goaff_commission_type": comm.get("type", "") if comm else "",
        "goaff_commission_amount": comm.get("amount", "") if comm else "",
        "goaff_commission_on": comm.get("on", "") if comm else "",
    }


def map_refersion_offer(offer: dict) -> dict:
    category_raw = offer.get("category")
    if category_raw in (None, ""):
        category_raw = offer.get("categories")
    denied_raw = offer.get("denied")
    pending_raw = offer.get("pending")
    payments_raw = offer.get("payments")
    return {
        "brand": offer.get("brand") or "",
        "url": offer.get("url") or "",
        "offer": offer.get("offer") or "",
        "cookieDays": offer.get("cookieDays") or "",
        "client_url": offer.get("client_url") or "",
        "offer_id": offer.get("offer_id") or "",
        "shop_id": offer.get("client_id") or "",
        "program_id": offer.get("id") or "",
        "mkp_listing_id": offer.get("id") or "",
        "commission_type": offer.get("commission_type") or "",
        "category": fmt_refersion_scalar(category_raw),
        "epc": offer.get("epc") or "",
        "payments": fmt_refersion_scalar(payments_raw),
        "currency": offer.get("client_currency_symbol") or "",
        "payout_rate": "",
        "approval_rate": "",
        "offer_score": "",
        "recommend_score": "",
        "application_review": "",
        "promotion_details": [],
        "target_audience_customer_channels": [],
        "target_audience_locations": [],
        "target_audience_ages": [],
        "target_audience_genders": [],
        "can_apply_offer": offer.get("applied"),
        "is_applied_offer": offer.get("applied"),
        "refersion_id": offer.get("id") or "",
        "refersion_client_id": offer.get("client_id") or "",
        "refersion_offer_id": offer.get("offer_id") or "",
        "refersion_visible_url": offer.get("visible_url") or "",
        "refersion_payments": fmt_refersion_scalar(payments_raw),
        "refersion_category": fmt_refersion_scalar(category_raw),
        "refersion_denied": denied_raw,
        "refersion_pending": pending_raw,
    }


def _collabs_storefront_url(social_links) -> str:
    if not isinstance(social_links, list):
        return ""
    for item in social_links:
        if not isinstance(item, dict):
            continue
        platform = str(item.get("platform") or "").strip().upper()
        if platform != "STOREFRONT":
            continue
        url = str(item.get("url") or "").strip()
        if url.startswith("http://") or url.startswith("https://"):
            return url
    return ""


def _collabs_brand_url(node: dict, detail_brand: dict | None = None) -> str:
    # Ưu tiên domain storefront từ API detail, vì list query không trả website đầy đủ.
    if isinstance(detail_brand, dict):
        storefront_url = _collabs_storefront_url(detail_brand.get("socialLinks"))
        if storefront_url:
            return storefront_url
    # Fallback cũ: map sang myshopify khi chỉ có store id.
    store = node.get("shopifyStore") if isinstance(node.get("shopifyStore"), dict) else {}
    sid = str(store.get("shopifyStoreId") or "").strip()
    if sid and sid.isdigit():
        return f"https://{sid}.myshopify.com"
    return ""


def map_collabs_brand(node: dict, detail_brand: dict | None = None) -> dict:
    name = str(node.get("name") or "").strip()
    commission = node.get("networkCommissionRange")
    if commission in (None, ""):
        offer = ""
    else:
        offer = f"{commission}%"
    category_code = str(node.get("productCategory") or "").strip().upper()
    return {
        "brand": name,
        "url": _collabs_brand_url(node, detail_brand),
        "offer": offer,
        "cookieDays": "",
        "client_url": "",
        "offer_id": node.get("id") or "",
        "shop_id": (node.get("shopifyStore") or {}).get("id") if isinstance(node.get("shopifyStore"), dict) else "",
        "program_id": "",
        "mkp_listing_id": node.get("id") or "",
        "commission_type": "network_commission_range",
        "category": collabs_category_label(category_code),
        "collabs_product_category_code": category_code,
        "epc": "",
        "payments": "",
        "currency": "",
        "payout_rate": "",
        "approval_rate": "",
        "offer_score": "",
        "recommend_score": "",
        "application_review": "",
        "promotion_details": [],
        "target_audience_customer_channels": [],
        "target_audience_locations": [],
        "target_audience_ages": [],
        "target_audience_genders": [],
        "can_apply_offer": None,
        "is_applied_offer": None,
        "collabs_partnership_status": node.get("partnershipStatus") or "",
        "collabs_partnership_state": node.get("partnershipState") or "",
        "collabs_saved": node.get("saved"),
        "collabs_previously_purchased": node.get("previouslyPurchased"),
        "collabs_target_countries": join_list(node.get("targetCountries") or []),
        "collabs_logo_url": node.get("logoUrl") or "",
        "collabs_images": join_list(node.get("images") or []),
        "collabs_shopify_store_id": ((node.get("shopifyStore") or {}).get("shopifyStoreId") if isinstance(node.get("shopifyStore"), dict) else ""),
        "collabs_holding_period": (detail_brand.get("holdingPeriod") if isinstance(detail_brand, dict) else ""),
    }


def collabs_shopify_store_gid(node: dict) -> str:
    store = node.get("shopifyStore") if isinstance(node.get("shopifyStore"), dict) else {}
    raw_id = str(store.get("id") or "").strip()
    if not raw_id:
        return ""
    if raw_id.startswith("gid://"):
        return raw_id
    return f"gid://dovetale-api/ShopifyStore/{raw_id}"


def resolve_redirected_url(url: str, timeout_sec: int = 20) -> str:
    """
    Theo dõi redirect để lấy URL đích cuối (brand gốc).
    Trả về URL ban đầu nếu request lỗi/timeout.
    """
    raw = str(url or "").strip()
    if not raw:
        return ""
    if not (raw.startswith("http://") or raw.startswith("https://")):
        return raw
    try:
        res = requests.get(raw, allow_redirects=True, timeout=timeout_sec, stream=True)
        final_url = str(res.url or "").strip()
        try:
            res.close()
        except Exception:
            pass
        return final_url or raw
    except Exception:
        return raw


def _collabs_html_has_outside_cta_phrases(low_html: str) -> bool:
    """
    CTA trên trang (low_html = HTML đã .lower()).
    Logic **HOẶC**: chỉ cần xuất hiện **một** trong các nhóm dưới → True.
    - apply now hoặc apply-now
    - chứa chuỗi con join (joining, rejoin, joint, …)
    - sign up
    - partner with us
    - get started
    - join community
    """
    if not low_html:
        return False
    return any(
        (
            ("apply now" in low_html or "apply-now" in low_html),
            ("join" in low_html),
            ("sign up" in low_html),
            ("partner with us" in low_html),
            ("get started" in low_html),
            ("join community" in low_html),
        )
    )


def _collabs_page_looks_like_signup(html: str, url: str = "") -> bool:
    text = (html or "").lower()
    u = (url or "").lower()
    title_match = re.search(r"<title[^>]*>(.*?)</title>", html or "", flags=re.I | re.S)
    title_text = (title_match.group(1).strip().lower() if title_match else "")
    if "class=\"collabs-page__main\"" in text or "class='collabs-page__main'" in text:
        return True
    if "community" in title_text:
        return True
    signup_path_hints = (
        "/pages/collab",
        "/pages/collabs",
        "/pages/collabs-signup",
        "/pages/partnerships",
        "/pages/ambassador",
        "/pages/affiliate",
        "/pages/affiliates",
        "/pages/collaborators",
        "/pages/affiliate-program",
        "/pages/affiliate-programs",
        "/pages/ambassadors",
        "/pages/ambassador-program",
        "/pages/ambassador-programs",
        "/pages/collaboration",
        "/pages/partners",
        "/pages/partner-program",
        "/pages/partner-programs",
        "/pages/collaborations",
        "/pages/curious-community",
        "/pages/shopify-collabs",
        "/ambassadors",
        "/ambassador",
        "/community",
        "/affiliates",
        "/affiliate",
        "/affiliate-program",
        "/affiliate-programs",
        "/collab",
        "/collabs",
        "/collaborators",
        "/partnerships",
        "/partners",
        "/partner",
        "/partner-program",
        "/partner-programs",
        "/collaborations",
        "/collaboration",
        "/ambassador-program",
        "/ambassador-programs",
        "/curious-community",
        "-com",
        "ambassadors",
        "ambassador",
        "community",
        "affiliates",
        "affiliate",
        "affiliate-program",
        "affiliate-programs",
        "collab",
        "collabs",
        "collaborators",
        "partnerships",
        "partners",
        "partner",
        "partner-program",
        "partner-programs",
        "collaborations",
        "collaboration",
        "ambassador-program",
        "ambassador-programs",
        "curious-community"
    )
    if any(p in u for p in signup_path_hints):
        return True
    if _collabs_html_has_outside_cta_phrases(text):
        return True
    score = 0
    if "collab" in text:
        score += 1
    if "affiliate" in text:
        score += 1
    if "apply" in text or "application" in text:
        score += 1
    return score >= 2


def _collabs_page_has_signup_cta(html: str) -> bool:
    """
    Tiêu chí đúng cho trang đăng ký Collabs:
    page phải có element: <div class="collabs-page__cta ..."> (hoặc single-quote).
    """
    h = html or ""
    # Bắt buộc là div.collabs-page__cta để tránh false-positive từ script/css hoặc element khác.
    return bool(
        re.search(
            r"""<div[^>]*\bclass\s*=\s*["'][^"']*\bcollabs-page__cta\b[^"']*["'][^>]*>""",
            h,
            flags=re.I,
        )
    )


def discover_collabs_signup_url(
    site_url: str, timeout_sec: int = 20, should_stop: Callable[[], bool] | None = None
) -> str:
    """
    Tìm link đăng ký collabs từ domain chính.
    Ưu tiên /pages/collab; fallback quét homepage để tìm href chứa collab/affiliate.
    Nếu should_stop trả True (vd. người dùng bấm hủy), trả chuỗi rỗng sớm.
    """
    raw = str(site_url or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw if "://" in raw else f"https://{raw}")
    if not parsed.netloc:
        return ""
    base = f"{parsed.scheme or 'https'}://{parsed.netloc}"
    default_signup = f"{base}/pages/collab"

    # Tổng thời gian tìm kiếm cho 1 domain (giây). Quá thời gian -> trả default và nhảy domain khác.
    # Yêu cầu mới: tối đa 30 giây (có thể override bằng env).
    max_domain_search_sec = int(os.getenv("COLLABS_SIGNUP_MAX_DOMAIN_SECONDS", "30") or "30")
    if max_domain_search_sec < 5:
        max_domain_search_sec = 5
    deadline = time.monotonic() + float(max_domain_search_sec)

    def _halt() -> bool:
        return should_stop is not None and bool(should_stop())

    def _remaining_sec() -> float:
        return max(0.0, deadline - time.monotonic())

    def _try_get(candidate_url: str) -> tuple[str, str]:
        if _halt() or _remaining_sec() <= 0:
            return "", ""
        try:
            # Mỗi request dùng timeout theo phần thời gian còn lại (không vượt quá timeout_sec).
            req_timeout = min(float(timeout_sec), max(0.5, _remaining_sec()))
            res = requests.get(candidate_url, allow_redirects=True, timeout=req_timeout)
            final_url = str(res.url or candidate_url).strip()
            if not res.ok:
                return "", ""
            ctype = str(res.headers.get("content-type") or "").lower()
            # Tránh tải nhầm file/binary; ưu tiên HTML.
            if ctype and ("text/html" not in ctype) and ("application/xhtml" not in ctype):
                return "", ""
            text = res.text or ""
            return final_url, text
        except Exception:
            return "", ""

    def _bases_to_try() -> list[str]:
        # Dù input là http/https, thử cả hai để giảm bỏ sót.
        return [f"https://{parsed.netloc}", f"http://{parsed.netloc}"]

    # 0) Quét nhanh homepage trước khi brute-force preferred paths
    # để bắt các slug tùy biến như /pages/westoncommunity.
    home_url, home_html = _try_get(base)
    if home_url:
        quick_keywords = ("community", "-com", "collab", "affiliate", "ambassador", "partner", "apply", "creator", "influencer", "ambassadors", "ambassador", "community", "affiliates", "affiliate", "affiliate-program", "affiliate-programs", "collab", "collabs", "collaborators", "partnerships", "partners", "partner", "partner-program", "partner-programs", "collaborations", "collaboration", "ambassador-program", "ambassador-programs", "curious-community")
        quick_links = []
        for href in re.findall(r"""href=["']([^"'#]+)["']""", home_html or "", flags=re.I):
            h = str(href or "").strip()
            if not h:
                continue
            low_h = h.lower()
            if low_h.startswith(("mailto:", "tel:", "javascript:", "data:")):
                continue
            cand = urljoin(home_url, h)
            cp = urlparse(cand)
            if cp.netloc != parsed.netloc:
                continue
            normalized = f"{cp.scheme or 'https'}://{cp.netloc}{cp.path or '/'}"
            low = normalized.lower()
            if any(k in low for k in quick_keywords):
                quick_links.append(normalized)
        # Ưu tiên community trước, sau đó collab/affiliate.
        quick_links = sorted(
            set(quick_links),
            key=lambda u: (0 if "community" in u.lower() else (1 if "collab" in u.lower() else 2)),
        )
        for cand in quick_links[:60]:
            if _halt() or _remaining_sec() <= 0:
                break
            final_url, html = _try_get(cand)
            if final_url and _collabs_page_has_signup_cta(html):
                return final_url

    preferred_paths = [
        "/pages/collab",
        "/pages/collabs",
        "/pages/collabs-signup",
        "/pages/partnerships",
        "/pages/ambassador",
        "/pages/affiliate",
        "/pages/affiliates",
        "/pages/collaborators",
        "/pages/affiliate-program",
        "/pages/affiliate-programs",
        "/pages/ambassadors",
        "/pages/ambassador-program",
        "/pages/ambassador-programs",
        "/pages/collaboration",
        "/pages/partners",
        "/pages/partner-program",
        "/pages/partner-programs",
        "/pages/collaborations",
        "/pages/curious-community",
        "/pages/shopify-collabs",
        "/ambassadors",
        "/ambassador",
        "/community",
        "/affiliates",
        "/affiliate",
        "/affiliate-program",
        "/affiliate-programs",
        "/collab",
        "/collabs",
        "/collaborators",
        "/partnerships",
        "/partners",
        "/partner",
        "/partner-program",
        "/partner-programs",
        "/collaborations",
        "/collaboration",
        "/ambassador-program",
        "/ambassador-programs",
        "/curious-community",
        "-com",
    ]
    # 1) Thử các path phổ biến trên cả https/http base. Tìm thấy là trả ngay.
    for b in _bases_to_try():
        for p in preferred_paths:
            if _halt():
                return ""
            if _remaining_sec() <= 0:
                return default_signup
            final_url, text = _try_get(f"{b}{p}")
            if final_url and _collabs_page_has_signup_cta(text):
                return final_url

    # 2) Nếu homepage chưa có từ bước quét nhanh thì lấy lại.
    if not home_url:
        home_url, home_html = _try_get(base)
    if _halt():
        return ""
    if not home_url:
        return default_signup

    # 2.05) Quick-hit từ homepage: thử ngay các link nội bộ có tín hiệu cộng đồng/collab
    # để tránh rơi fallback /pages/collab khi site dùng slug tùy biến như /pages/westoncommunity.
    def _extract_same_domain_links_quick(page_html: str, page_url: str) -> list[str]:
        out = []
        for href in re.findall(r"""href=["']([^"'#]+)["']""", page_html or "", flags=re.I):
            h = str(href).strip()
            if not h:
                continue
            low_h = h.lower()
            if low_h.startswith(("mailto:", "tel:", "javascript:", "data:")):
                continue
            cand = urljoin(page_url, h)
            cp = urlparse(cand)
            if cp.netloc != parsed.netloc:
                continue
            path = cp.path or "/"
            low_path = path.lower()
            if any(
                low_path.endswith(ext)
                for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg", ".js", ".css", ".pdf", ".xml")
            ):
                continue
            out.append(f"{cp.scheme or 'https'}://{cp.netloc}{path}")
        return out

    quick_candidates = _extract_same_domain_links_quick(home_html, home_url)
    # Ưu tiên community/collab trước để tăng độ chính xác cho Shopify Collabs page.
    quick_candidates.sort(
        key=lambda u: (
            0
            if "community" in u.lower()
            else (1 if "collab" in u.lower() else (2 if "affiliate" in u.lower() else 3))
        )
    )
    for cand in quick_candidates:
        if _halt() or _remaining_sec() <= 0:
            break
        low = cand.lower()
        if not any(k in low for k in ("community", "collab", "affiliate", "ambassador", "partner", "apply")):
            continue
        final_url, html = _try_get(cand)
        if final_url and _collabs_page_has_signup_cta(html):
            return final_url

    # 2.1) Thử lấy thêm candidate từ sitemap để bắt các slug tùy biến
    # kiểu /pages/westoncommunity (không nằm trong preferred_paths).
    def _try_sitemap_candidates() -> str:
        if _halt() or _remaining_sec() <= 0:
            return ""
        sitemap_urls = [f"{base}/sitemap.xml", f"{base}/sitemap_index.xml"]
        keywords = ("collab", "affiliate", "ambassador", "community", "partner", "creator", "influencer", "apply", "-com", "collaborations", "collaboration", "collaborator", "collaborators", "partnerships", "partnership", "partners", "partner", "partner-program", "partner-programs", "ambassadors", "ambassador", "ambassador-program", "ambassador-programs", "curious-community")
        seen_sitemaps = set()

        def _iter_locs_from_sitemap(sm_url: str, depth: int = 0) -> list[str]:
            if _halt() or _remaining_sec() <= 0:
                return []
            s = str(sm_url or "").strip()
            if not s or s in seen_sitemaps:
                return []
            seen_sitemaps.add(s)
            try:
                req_timeout = min(float(timeout_sec), max(0.5, _remaining_sec()))
                res = requests.get(s, allow_redirects=True, timeout=req_timeout)
                if not res.ok:
                    return []
                text = res.text or ""
            except Exception:
                return []
            locs = [str(x or "").strip() for x in re.findall(r"<loc>(.*?)</loc>", text, flags=re.I | re.S)]
            out: list[str] = []
            for loc in locs:
                if not loc:
                    continue
                cp = urlparse(loc)
                if cp.netloc != parsed.netloc:
                    continue
                low = loc.lower()
                # sitemap index -> đi sâu thêm 1 cấp vào sitemap con
                if low.endswith(".xml"):
                    if depth < 1:
                        out.extend(_iter_locs_from_sitemap(loc, depth=depth + 1))
                    continue
                out.append(loc)
            return out

        for sm in sitemap_urls:
            if _halt() or _remaining_sec() <= 0:
                return ""
            for loc in _iter_locs_from_sitemap(sm):
                if _halt() or _remaining_sec() <= 0:
                    return ""
                cand = str(loc or "").strip()
                if not cand:
                    continue
                cp = urlparse(cand)
                if cp.netloc != parsed.netloc:
                    continue
                low = cand.lower()
                if not any(k in low for k in keywords):
                    continue
                final_url, html = _try_get(cand)
                if final_url and _collabs_page_has_signup_cta(html):
                    return final_url
        return ""

    sm_url = _try_sitemap_candidates()
    if sm_url:
        return sm_url

    # Quét toàn site (có giới hạn) để tìm page apply.
    # Mặc định nâng lên để giảm bỏ sót; có thể chỉnh bằng env.
    # Đặt <=0 để "không giới hạn" nhưng vẫn bị chặn bởi hard cap an toàn.
    max_scan_pages = int(os.getenv("COLLABS_SIGNUP_SCAN_MAX_PAGES", "500") or "500")
    hard_cap = int(os.getenv("COLLABS_SIGNUP_SCAN_HARD_CAP", "5000") or "5000")
    if hard_cap < 50:
        hard_cap = 50

    def _extract_same_domain_links(page_html: str, page_url: str) -> list[str]:
        out = []
        for href in re.findall(r"""href=["']([^"'#]+)["']""", page_html or "", flags=re.I):
            h = str(href).strip()
            if not h:
                continue
            # Bỏ các scheme không phải http(s)
            low_h = h.lower()
            if low_h.startswith(("mailto:", "tel:", "javascript:", "data:")):
                continue
            cand = urljoin(page_url, h)
            cp = urlparse(cand)
            if cp.netloc != parsed.netloc:
                continue
            path = cp.path or "/"
            # Bỏ link static/media để không tốn lượt quét.
            low_path = path.lower()
            if any(low_path.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg", ".js", ".css", ".pdf", ".xml")):
                continue
            # Chuẩn hóa bỏ query/fragment để tránh trùng lặp vô hạn.
            out.append(f"{cp.scheme or 'https'}://{cp.netloc}{path}")
        return out

    queue = [home_url]
    seen = set()
    idx = 0
    def _can_scan_more() -> bool:
        if len(seen) >= hard_cap:
            return False
        if max_scan_pages <= 0:
            return True
        return len(seen) < max_scan_pages

    while idx < len(queue) and _can_scan_more():
        if _halt():
            return ""
        if _remaining_sec() <= 0:
            return default_signup
        cur = queue[idx]
        idx += 1
        if cur in seen:
            continue
        seen.add(cur)
        final_url, text = _try_get(cur)
        if not final_url:
            continue
        # Chỉ coi là trang đăng ký nếu có CTA thật.
        if _collabs_page_has_signup_cta(text):
            return final_url

        for nxt in _extract_same_domain_links(text, final_url):
            if nxt in seen:
                continue
            low = nxt.lower()
            # Ưu tiên đường dẫn có tín hiệu collab/apply/affiliate/community/-com/partnership.
            if any(k in low for k in ("collab", "affiliate", "ambassador", "apply", "community", "-com", "partnership")):
                queue.insert(idx, nxt)
            else:
                queue.append(nxt)

    # Không tìm thấy CTA trong ngân sách thời gian / số trang -> fallback default.
    if _halt():
        return ""
    return default_signup


def _extract_url_from_ddg_href(href: str) -> str:
    h = str(href or "").strip()
    if not h:
        return ""
    if h.startswith("http://") or h.startswith("https://"):
        return h
    # DuckDuckGo thường bọc URL trong tham số uddg.
    if "uddg=" in h:
        try:
            q = dict(parse_qsl(urlparse(h).query, keep_blank_values=True))
            raw = unquote(str(q.get("uddg") or "").strip())
            if raw.startswith("http://") or raw.startswith("https://"):
                return raw
        except Exception:
            return ""
    return ""


def _fetch_duckduckgo_result_urls(query: str, max_results: int = 80, timeout_sec: int = 20, delay_ms: int = 350) -> list[str]:
    q = str(query or "").strip()
    if not q:
        return []
    out: list[str] = []
    seen: set[str] = set()
    offset = 0
    page_size = 30
    while len(out) < max_results:
        search_url = f"https://html.duckduckgo.com/html/?q={quote_plus(q)}&s={offset}"
        try:
            res = requests.get(
                search_url,
                timeout=timeout_sec,
                headers={
                    "user-agent": os.getenv(
                        "COLLABS_OUTSIDE_UA",
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
                    )
                },
            )
            if not res.ok:
                break
            html = res.text or ""
        except Exception:
            break

        hrefs = re.findall(r"""class=["']result__a["'][^>]*href=["']([^"']+)["']""", html, flags=re.I)
        if not hrefs:
            # Fallback parser: lấy mọi anchor có link ra ngoài.
            hrefs = re.findall(r"""<a[^>]*href=["']([^"']+)["']""", html, flags=re.I)
        added = 0
        for href in hrefs:
            final_url = _extract_url_from_ddg_href(href)
            if not final_url:
                continue
            p = urlparse(final_url)
            if not p.netloc:
                continue
            norm = f"{p.scheme or 'https'}://{p.netloc}{p.path or '/'}"
            low = norm.lower()
            if any(x in low for x in ("facebook.com", "instagram.com", "tiktok.com", "youtube.com", "x.com/", "twitter.com/")):
                continue
            if norm in seen:
                continue
            seen.add(norm)
            out.append(norm)
            added += 1
            if len(out) >= max_results:
                break
        if added == 0:
            break
        offset += page_size
        if delay_ms > 0:
            time.sleep(delay_ms / 1000.0)
    return out


def _fetch_bing_rss_result_urls(
    query: str, max_results: int = 80, timeout_sec: int = 20, should_stop: Callable[[], bool] | None = None
) -> list[str]:
    q = str(query or "").strip()
    if not q:
        return []
    if should_stop and should_stop():
        return []
    try:
        res = requests.get(
            f"https://www.bing.com/search?q={quote_plus(q)}&format=rss",
            timeout=timeout_sec,
            headers={
                "user-agent": os.getenv(
                    "COLLABS_OUTSIDE_UA",
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
                )
            },
        )
        if not res.ok:
            return []
        xml = res.text or ""
    except Exception:
        return []

    out: list[str] = []
    seen: set[str] = set()
    links = re.findall(r"<link>(.*?)</link>", xml, flags=re.I | re.S)
    for lk in links:
        url = str(lk or "").strip()
        if not (url.startswith("http://") or url.startswith("https://")):
            continue
        p = urlparse(url)
        if not p.netloc:
            continue
        low = (p.netloc + (p.path or "")).lower()
        if "bing.com/search" in low:
            continue
        if any(x in low for x in ("facebook.com", "instagram.com", "tiktok.com", "youtube.com", "x.com/", "twitter.com/")):
            continue
        norm = f"{p.scheme or 'https'}://{p.netloc}{p.path or '/'}"
        if norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
        if len(out) >= max_results:
            break
    return out


def _google_cse_credentials() -> tuple[str, str]:
    """Trả về (api_key, cx) nếu đủ cấu hình Google Custom Search JSON API; ngược lại ('','')."""
    key = (os.getenv("GOOGLE_CUSTOM_SEARCH_API_KEY") or os.getenv("GOOGLE_CSE_API_KEY") or "").strip()
    cx = (os.getenv("GOOGLE_CUSTOM_SEARCH_ENGINE_ID") or os.getenv("GOOGLE_CSE_CX") or "").strip()
    return key, cx


def _outside_cursor_path() -> Path:
    return Path(os.getenv("COLLABS_OUTSIDE_CURSOR_FILE", str(BASE_DIR / ".collabs_outside_cursor.json")))


def _outside_dedup_path() -> Path:
    return Path(os.getenv("COLLABS_OUTSIDE_DEDUP_FILE", str(BASE_DIR / ".collabs_outside_seen_hosts.json")))


def _outside_cursor_key(provider: str, query: str) -> str:
    return f"{provider}::{(query or '').strip().lower()}"


def _load_outside_cursor(provider: str, query: str) -> int:
    p = _outside_cursor_path()
    if not p.exists():
        return 0
    try:
        payload = json.loads(p.read_text(encoding="utf-8-sig") or "{}")
    except Exception:
        return 0
    if not isinstance(payload, dict):
        return 0
    cursors = payload.get("cursors") or {}
    if not isinstance(cursors, dict):
        return 0
    raw = cursors.get(_outside_cursor_key(provider, query), 0)
    try:
        n = int(raw)
    except Exception:
        return 0
    return max(0, n)


def _save_outside_cursor(provider: str, query: str, next_offset: int) -> None:
    p = _outside_cursor_path()
    data: dict = {}
    try:
        if p.exists():
            loaded = json.loads(p.read_text(encoding="utf-8-sig") or "{}")
            if isinstance(loaded, dict):
                data = loaded
    except Exception:
        data = {}
    cursors = data.get("cursors")
    if not isinstance(cursors, dict):
        cursors = {}
    cursors[_outside_cursor_key(provider, query)] = max(0, int(next_offset))
    data["cursors"] = cursors
    data["updatedAt"] = time.time()
    try:
        p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    except Exception:
        return


def _fetch_google_result_urls_via_custom_search_api(
    query: str,
    max_results: int = 80,
    timeout_sec: int = 20,
    should_stop: Callable[[], bool] | None = None,
) -> list[str]:
    """
    Lấy URL organic qua Google Custom Search JSON API (hỗ trợ cú pháp Google: inurl, OR, ngoặc, …).
    Yêu cầu .env:
    - GOOGLE_CUSTOM_SEARCH_API_KEY (hoặc GOOGLE_CSE_API_KEY)
    - GOOGLE_CUSTOM_SEARCH_ENGINE_ID (hoặc GOOGLE_CSE_CX — Search engine ID)

    Mỗi request tối đa 10 kết quả; API không trả quá 100 kết quả cho một truy vấn (phân trang start 1..91).
    """
    api_key, cx = _google_cse_credentials()
    if not api_key or not cx:
        raise RuntimeError(
            "Thiếu GOOGLE_CUSTOM_SEARCH_API_KEY hoặc GOOGLE_CUSTOM_SEARCH_ENGINE_ID (cx) cho Google CSE."
        )
    q = str(query or "").strip()
    if not q:
        return []
    max_results = max(1, min(300, int(max_results)))
    # JSON API: tổng chỉ mục tối đa ~100; mỗi lần gọi num <= 10 và start + num <= 101.
    cap = min(max_results, 100)
    base = "https://www.googleapis.com/customsearch/v1"
    out: list[str] = []
    seen: set[str] = set()
    start = 1
    ua = (os.getenv("COLLABS_OUTSIDE_UA") or "").strip() or (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36"
    )
    headers = {"User-Agent": ua, "Accept": "application/json"}
    while len(out) < cap and start <= 100:
        if should_stop and should_stop():
            return out
        num = min(10, cap - len(out), max(0, 101 - start))
        if num < 1:
            break
        params = {"key": api_key, "cx": cx, "q": q, "num": num, "start": start}
        try:
            res = requests.get(base, params=params, headers=headers, timeout=int(timeout_sec))
        except requests.RequestException as e:
            raise RuntimeError(f"Google Custom Search API: lỗi mạng — {e}") from e
        if not res.ok:
            snippet = (res.text or "")[:400]
            if res.status_code == 429:
                raise RuntimeError("Google Custom Search API hết quota (429). Thử lại sau hoặc kiểm tra billing.")
            raise RuntimeError(f"Google Custom Search API lỗi HTTP {res.status_code}: {snippet}")
        try:
            data = res.json() if res.text else {}
        except json.JSONDecodeError:
            raise RuntimeError("Google Custom Search API trả nội dung không phải JSON.")
        items = data.get("items") or []
        if not items:
            break
        for it in items:
            if not isinstance(it, dict):
                continue
            u = it.get("link")
            if not isinstance(u, str):
                continue
            u = u.strip()
            if not u.startswith(("http://", "https://")):
                continue
            p = urlparse(u)
            if not p.netloc:
                continue
            low = (p.netloc + (p.path or "")).lower()
            if any(
                x in low
                for x in ("facebook.com", "instagram.com", "tiktok.com", "youtube.com", "x.com/", "twitter.com/")
            ):
                continue
            norm = f"{p.scheme or 'https'}://{p.netloc}{p.path or '/'}"
            if norm in seen:
                continue
            seen.add(norm)
            out.append(norm)
            if len(out) >= cap:
                break
        start += len(items)
        if len(items) < num:
            break
        time.sleep(0.12)
    return out


def _apify_run_actor(actor_id: str, run_input: dict, wait_secs: int = 180, token: str | None = None) -> str:
    """
    Run Apify actor và trả về defaultDatasetId.
    """
    tok = (str(token or "").strip() or (os.getenv("APIFY_TOKEN") or "").strip())
    if not tok:
        raise RuntimeError("Thiếu APIFY_TOKEN trong .env (cần để chạy Apify actor).")
    act = str(actor_id or "").strip()
    if not act:
        raise RuntimeError("Thiếu COLLABS_OUTSIDE_GOOGLE_ACTOR_ID (Apify actor id).")
    # Apify API expects "username~actorname" or actorId, not "username/actorname".
    if "/" in act and "~" not in act:
        parts = [p for p in act.split("/") if p]
        if len(parts) >= 2:
            act = f"{parts[0]}~{parts[1]}"
    url = f"https://api.apify.com/v2/acts/{act}/runs?token={tok}&waitForFinish={int(wait_secs)}"
    res = requests.post(url, json=run_input, timeout=max(30, int(wait_secs) + 30))
    if not res.ok:
        raise RuntimeError(f"Apify run actor lỗi HTTP {res.status_code}: {res.text[:300]}")
    body = res.json() if res.text else {}
    data = (body or {}).get("data") or {}
    status = str(data.get("status") or "").upper()
    run_id = str(data.get("id") or "").strip()
    # Một số lần waitForFinish trả READY/RUNNING (queue/tải cao). Poll thêm thay vì fail ngay.
    if status in {"READY", "RUNNING"} and run_id:
        deadline = time.monotonic() + max(10, int(wait_secs))
        while time.monotonic() < deadline:
            time.sleep(2.0)
            st_url = f"https://api.apify.com/v2/actor-runs/{run_id}?token={tok}"
            st = requests.get(st_url, timeout=30)
            if not st.ok:
                continue
            st_body = st.json() if st.text else {}
            st_data = (st_body or {}).get("data") or {}
            status = str(st_data.get("status") or "").upper()
            if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
                data = st_data
                break
    if status != "SUCCEEDED":
        raise RuntimeError(f"Apify actor chưa SUCCEEDED (status={status})")
    dataset_id = str(data.get("defaultDatasetId") or "").strip()
    if not dataset_id:
        raise RuntimeError("Apify actor không trả defaultDatasetId")
    return dataset_id


def _extract_urls_from_google_actor_items(items: list) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    if not isinstance(items, list):
        return []
    for it in items:
        if not isinstance(it, dict):
            continue
        # 1) Một số actor trả thẳng link/url
        for k in ("url", "link", "resultUrl", "result_url"):
            v = it.get(k)
            if isinstance(v, str) and v.strip().startswith(("http://", "https://")):
                u = v.strip()
                if u not in seen:
                    seen.add(u)
                    out.append(u)
        # 2) google-search-scraper thường trả organicResults:[{url,...}]
        org = it.get("organicResults") or it.get("organic_results") or []
        if isinstance(org, list):
            for r in org:
                if not isinstance(r, dict):
                    continue
                u = r.get("url") or r.get("link")
                if isinstance(u, str) and u.strip().startswith(("http://", "https://")):
                    uu = u.strip()
                    if uu not in seen:
                        seen.add(uu)
                        out.append(uu)
    return out


def _fetch_google_result_urls_via_apify(
    query: str,
    max_results: int = 80,
    timeout_sec: int = 180,
    start_page: int = 1,
    end_page: int | None = None,
    should_stop: Callable[[], bool] | None = None,
) -> list[str]:
    """
    Lấy URL từ Google Search theo đúng cú pháp (inurl/OR/ngoặc) bằng Apify actor.
    Yêu cầu env:
    - COLLABS_OUTSIDE_APIFY_TOKEN (khuyến nghị) hoặc APIFY_TOKEN
    - COLLABS_OUTSIDE_GOOGLE_ACTOR_ID (vd: apify/google-search-scraper)
    """
    q = str(query or "").strip()
    if not q:
        return []
    actor_id = (os.getenv("COLLABS_OUTSIDE_GOOGLE_ACTOR_ID") or "").strip()
    search_token = (os.getenv("COLLABS_OUTSIDE_APIFY_TOKEN") or "").strip()
    if not search_token:
        # Fallback về APIFY_TOKEN để không phá luồng cũ, nhưng khuyến nghị tách token.
        search_token = (os.getenv("APIFY_TOKEN") or "").strip()
    # Google actor hiện bị giới hạn 10 kết quả/trang (theo cập nhật Google), nên dùng maxPagesPerQuery để tăng tổng.
    cap = max(10, min(300, int(max_results)))
    per_page = 10
    if start_page < 1:
        start_page = 1
    if end_page is not None and end_page < start_page:
        end_page = start_page
    # Tối ưu chi phí theo range trang người dùng chọn:
    # - Nếu có end_page: chỉ cần số bản ghi tương ứng với số trang trong range.
    # - Actor vẫn crawl từ trang 1..maxPagesPerQuery, nhưng ta không kéo dư theo max_results toàn cục.
    if end_page is not None:
        # Cần fetch đủ tới cuối range để slice chính xác (vd trang 4-5 cần tối thiểu 50 kết quả thô).
        raw_needed_for_slice = max(10, int(end_page) * per_page)
        # Ưu tiên đúng phạm vi trang user chọn: nếu MAX_RESULTS thấp hơn mức cần thiết,
        # tự nâng lên để không bị rỗng/thiếu dữ liệu khi chọn page sâu (vd 6-8).
        cap = max(cap, raw_needed_for_slice)
        cap = min(cap, 300)
        max_pages = max(1, min(20, int(end_page)))
    else:
        max_pages = max(1, min(20, max(1, (cap + per_page - 1) // per_page)))
    # Actor hiện tại yêu cầu input.queries là string.
    run_input = {
        "queries": q,
        "maxPagesPerQuery": max_pages,
        "resultsPerPage": per_page,
        "includeUnfilteredResults": False,
        "languageCode": "en",
        "mobileResults": False,
    }
    if should_stop and should_stop():
        return []
    dataset_id = _apify_run_actor(actor_id, run_input, wait_secs=int(timeout_sec), token=search_token)
    if not dataset_id:
        return []
    items = apify_list_items(dataset_id, token=search_token)
    urls = _extract_urls_from_google_actor_items(items)
    # chuẩn hóa + giới hạn
    out: list[str] = []
    seen: set[str] = set()
    for u in urls:
        p = urlparse(u)
        if not p.netloc:
            continue
        norm = f"{p.scheme or 'https'}://{p.netloc}{p.path or '/'}"
        if norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
        if len(out) >= cap:
            break
    # giả lập phân trang Google: 10 kết quả / trang
    if start_page > 1 or end_page is not None:
        s = (start_page - 1) * 10
        e = (int(end_page) * 10) if end_page is not None else len(out)
        out = out[s:e]
    return out


# Gán trước khi gọi discover (vd. từ webapp) để log và tìm kiếm dùng cùng một chuỗi query.
EFFECTIVE_OUTSIDE_QUERY_KEY = "_effective_outside_query"
# Thống kê nội bộ (sau discover) — target batch vs URL ứng viên vs brand sau verify.
OUTSIDE_DISCOVERY_STATS_KEY = "_outside_discovery_stats"

# Query cố định cho Outside Discovery (Google); ghép với dạng random chữ (x OR y) trong build_random_outside_collabs_google_query.
_OUTSIDE_COLLABS_GOOGLE_FIXED_QUERIES: tuple[str, ...] = (
    'inurl:pages/collabs "apply"',
    'inurl:pages/collabs "apply now"',
    'inurl:pages/collabs "apply today"',
    'inurl:pages/collabs "apply to join"',
    'inurl:pages/collabs "join"',
    'inurl:pages/collabs "join now"',
    'inurl:pages/collabs "join today"',
    'inurl:pages/collabs "join our program"',
    'inurl:pages/collabs "join our community"',
    'inurl:pages/collabs "join the program"',
    'inurl:pages/collabs "become a creator"',
    'inurl:pages/collabs "become a partner"',
    'inurl:pages/collabs "become an affiliate"',
    'inurl:pages/collabs "become an ambassador"',
    'inurl:pages/collabs "creator program"',
    'inurl:pages/collabs "affiliate program"',
    'inurl:pages/collabs "ambassador program"',
    'inurl:pages/collabs "influencer program"',
    'inurl:pages/collabs "content creator"',
    'inurl:pages/collabs "ugc creator"',
    'inurl:pages/collab "apply"',
    'inurl:pages/collab "apply now"',
    'inurl:pages/collab "apply to join"',
    'inurl:pages/collab "join"',
    'inurl:pages/collab "join now"',
    'inurl:pages/collab "join our community"',
    'inurl:pages/collab "join the program"',
    'inurl:pages/collab "become an affiliate"',
    'inurl:pages/collab "become a creator"',
    'inurl:pages/collab "creator program"',
    'inurl:pages/collab "affiliate program"',
    'inurl:pages/collab "ambassador program"',
    'inurl:pages/collab "ugc creator"',
    'inurl:pages/collab "content creator"',
    'inurl:pages "collabs" "apply"',
    'inurl:pages "collabs" "apply now"',
    'inurl:pages "collabs" "apply today"',
    'inurl:pages "collabs" "apply to join"',
    'inurl:pages "collabs" "join"',
    'inurl:pages "collabs" "join now"',
    'inurl:pages "collabs" "join today"',
    'inurl:pages "collabs" "join our program"',
    'inurl:pages "collabs" "join our community"',
    'inurl:pages "collabs" "become an affiliate"',
    'inurl:pages "collabs" "become a creator"',
    'inurl:pages "collabs" "become a partner"',
    'inurl:pages "collabs" "creator application"',
    'inurl:pages "collabs" "creator program"',
    'inurl:pages "collabs" "affiliate program"',
    'inurl:pages "collabs" "ambassador program"',
    'inurl:pages "collabs" "influencer program"',
    'inurl:pages "collabs" "ugc creator"',
    'inurl:pages "collabs" "content creator"',
    'inurl:pages "collab" "apply"',
    'inurl:pages "collab" "apply now"',
    'inurl:pages "collab" "apply to join"',
    'inurl:pages "collab" "join"',
    'inurl:pages "collab" "join now"',
    'inurl:pages "collab" "join our community"',
    'inurl:pages "collab" "join the program"',
    'inurl:pages "collab" "become an affiliate"',
    'inurl:pages "collab" "become a creator"',
    'inurl:pages "collab" "creator application"',
    'inurl:pages "collab" "creator program"',
    'inurl:pages "collab" "affiliate program"',
    'inurl:pages "collab" "ambassador program"',
    'inurl:pages "collab" "ugc creator"',
    'inurl:pages "collab" "content creator"',
    'inurl:pages "collabs" "sign up"',
    'inurl:pages "collabs" "get started"',
    'inurl:pages "collabs" "start now"',
    'inurl:pages "collabs" "start today"',
    'inurl:pages "collabs" "register"',
    'inurl:pages "collabs" "register now"',
    'inurl:pages "collab" "sign up"',
    'inurl:pages "collab" "get started"',
    'inurl:pages "collab" "start now"',
    'inurl:pages "collab" "register"',
    'inurl:pages "collab" "register now"',
    'inurl:pages/collabs "sign up"',
    'inurl:pages/collabs "get started"',
    'inurl:pages/collabs "start now"',
    'inurl:pages/collabs "register"',
    'inurl:pages/collabs "register now"',
    'inurl:pages/collab "sign up"',
    'inurl:pages/collab "get started"',
    'inurl:pages/collab "start now"',
    'inurl:pages/collab "register"',
    'inurl:pages/collab "register now"',
)


def _build_random_letter_pair_outside_collabs_query() -> str:
    """inurl:pages/collab(s) ("x" OR "y") với x≠y ngẫu nhiên."""
    x, y = random.sample(string.ascii_lowercase, 2)
    path = random.choice(("pages/collab", "pages/collabs"))
    return f'inurl:{path} ("{x}" OR "{y}")'


def build_random_outside_collabs_google_query() -> str:
    """
    Mỗi lần lọc: **một** query — ngẫu nhiên **hoặc** dạng chữ (x OR y) **hoặc** một chuỗi cố định.
    Hai họ có xác suất ~50/50 để dạng chữ vẫn được dùng thường xuyên.
    """
    if random.random() < 0.5:
        return _build_random_letter_pair_outside_collabs_query()
    return random.choice(_OUTSIDE_COLLABS_GOOGLE_FIXED_QUERIES)


def resolve_outside_discovery_query_string(filters: dict | None) -> str:
    """Chuỗi query web search: ưu tiên filters.outside_query, rồi COLLABS_OUTSIDE_QUERY, rồi query ngẫu nhiên."""
    f = filters if isinstance(filters, dict) else {}
    raw = str(f.get("outside_query") or os.getenv("COLLABS_OUTSIDE_QUERY") or "").strip()
    if raw:
        return raw
    return build_random_outside_collabs_google_query()


def _outside_discovery_verify_and_build_offer(
    url: str,
    verify_timeout_sec: int,
    should_stop: Callable[[], bool] | None,
) -> dict | None:
    """Verify một URL candidate ngoài Discovery; trả dict offer hoặc None."""
    if should_stop and should_stop():
        return None
    hk = host_key(url)
    if not hk:
        return None
    try:
        res = requests.get(url, timeout=verify_timeout_sec)
        if not res.ok:
            return None
        html = res.text or ""
    except Exception:
        return None

    low_url = url.lower()
    must_path = ("/pages/collab" in low_url) or ("/pages/collabs" in low_url)
    low_html = (html or "").lower()
    has_cta_phrase = _collabs_html_has_outside_cta_phrases(low_html)
    if not must_path:
        return None
    if not has_cta_phrase:
        return None

    looks_like = _collabs_page_looks_like_signup(html, url)
    weak_signal_url = any(
        k in low_url for k in ("/pages/collab", "/pages/collabs", "affiliate", "ambassador", "partner")
    )
    if not looks_like and not weak_signal_url:
        return None

    apply_url = discover_collabs_signup_url(url, timeout_sec=verify_timeout_sec, should_stop=should_stop)
    if not apply_url:
        return None

    brand_guess = hk.split(".")[0].replace("-", " ").replace("_", " ").title()
    return {
        "brand": brand_guess,
        "url": f"https://{hk}",
        "offer": "",
        "cookieDays": "",
        "client_url": apply_url,
        "offer_id": hk,
        "shop_id": "",
        "program_id": "",
        "mkp_listing_id": hk,
        "commission_type": "",
        "category": "Outside Discovery",
        "collabs_product_category_code": "",
        "epc": "",
        "payments": "",
        "currency": "",
        "payout_rate": "",
        "approval_rate": "",
        "offer_score": "",
        "recommend_score": "",
        "application_review": "",
        "promotion_details": [],
        "target_audience_customer_channels": [],
        "target_audience_locations": [],
        "target_audience_ages": [],
        "target_audience_genders": [],
        "can_apply_offer": None,
        "is_applied_offer": None,
        "collabs_partnership_status": "",
        "collabs_partnership_state": "",
        "collabs_saved": None,
        "collabs_previously_purchased": None,
        "collabs_target_countries": "",
        "collabs_logo_url": "",
        "collabs_images": "",
        "collabs_shopify_store_id": "",
        "collabs_holding_period": "",
    }


def discover_collabs_outside_discovery_offers(
    filters: dict | None = None, should_stop: Callable[[], bool] | None = None
) -> list[dict]:
    """
    Lọc Collabs ngoài Discovery từ web search:
    - Tìm candidate URL theo query.
    - Verify trang có tín hiệu collab/apply.
    - Tìm apply URL cuối bằng discover_collabs_signup_url.
    """
    f = filters if isinstance(filters, dict) else {}
    query_text = str(f.get(EFFECTIVE_OUTSIDE_QUERY_KEY) or "").strip() or resolve_outside_discovery_query_string(f)
    # Cú pháp Google (inurl/OR/ngoặc): dùng Apify actor, Google CSE JSON API, hoặc Bing (hỗ trợ một phần).
    query_list = [query_text] if query_text else []

    # outside_target_results (UI mới) ưu tiên hơn .env; fallback về COLLABS_OUTSIDE_MAX_RESULTS.
    raw_target = f.get("outside_target_results")
    if raw_target is None or str(raw_target).strip() == "":
        raw_target = os.getenv("COLLABS_OUTSIDE_MAX_RESULTS", "80") or "80"
    try:
        max_results = int(raw_target)
    except Exception:
        max_results = int(os.getenv("COLLABS_OUTSIDE_MAX_RESULTS", "80") or "80")
    max_results = max(10, min(30, max_results))
    max_results = (max_results // 10) * 10
    if max_results < 10:
        max_results = 10
    # Pool để cursor quay vòng; cho phép >300 nếu cần.
    raw_pool = os.getenv("COLLABS_OUTSIDE_CURSOR_POOL", "300") or "300"
    try:
        cursor_pool = int(raw_pool)
    except Exception:
        cursor_pool = 300
    cursor_pool = max(max_results, min(5000, cursor_pool))
    verify_timeout_sec = int(os.getenv("COLLABS_OUTSIDE_VERIFY_TIMEOUT_SEC", "30") or "30")
    if not query_list:
        return []
    if should_stop and should_stop():
        return []
    provider = (
        "apify"
        if (os.getenv("COLLABS_OUTSIDE_GOOGLE_ACTOR_ID") or "").strip()
        else ("google_cse" if _google_cse_credentials()[0] and _google_cse_credentials()[1] else "bing")
    )
    # Google CSE có trần ~100 kết quả/query -> giới hạn pool hiệu lực để tránh batch rỗng.
    effective_pool = min(cursor_pool, 100) if provider == "google_cse" else cursor_pool
    cursor_offset = _load_outside_cursor(provider, query_text)
    # Mỗi lần chạy lấy batch kế tiếp đúng bằng outside_target_results (max_results).
    start_idx = cursor_offset % effective_pool
    if start_idx + max_results > effective_pool:
        # Chạm giới hạn pool thì quay về page đầu cho lần này.
        start_idx = 0
    end_idx = start_idx + max_results
    start_page = (start_idx // 10) + 1
    end_page = max(start_page, (end_idx + 9) // 10)
    candidates: list[str] = []
    seen_candidate: set[str] = set()
    # Thứ tự ưu tiên: Apify Google actor → Google Custom Search JSON API → Bing RSS.
    # CSE hỗ trợ cú pháp Google (inurl/OR/…) và không cần Apify; Bing chỉ fallback khi không cấu hình Google.
    cse_key, cse_cx = _google_cse_credentials()
    if provider == "apify":
        fetch_budget = max(max_results, end_page * 10)
        urls = _fetch_google_result_urls_via_apify(
            query_text,
            max_results=fetch_budget,
            timeout_sec=180,
            start_page=start_page,
            end_page=end_page,
            should_stop=should_stop,
        )
    elif provider == "google_cse" and cse_key and cse_cx:
        fetch_budget = max(max_results, min(100, end_idx))
        urls = _fetch_google_result_urls_via_custom_search_api(
            query_text,
            max_results=fetch_budget,
            timeout_sec=verify_timeout_sec,
            should_stop=should_stop,
        )
        urls = urls[start_idx:end_idx]
    elif cse_key or cse_cx:
        raise RuntimeError(
            "Collabs ngoài Discovery: thiếu một trong hai — GOOGLE_CUSTOM_SEARCH_API_KEY và "
            "GOOGLE_CUSTOM_SEARCH_ENGINE_ID (hoặc GOOGLE_CSE_API_KEY / GOOGLE_CSE_CX)."
        )
    else:
        fetch_budget = max(max_results, end_idx)
        urls = _fetch_bing_rss_result_urls(
            query_text,
            max_results=fetch_budget,
            timeout_sec=verify_timeout_sec,
            should_stop=should_stop,
        )
        urls = urls[start_idx:end_idx]
    for u in urls:
        if u in seen_candidate:
            continue
        seen_candidate.add(u)
        candidates.append(u)

    # Dedupe giữa các lần chạy (theo host). Bật bằng env:
    # - COLLABS_OUTSIDE_DEDUP_PERSIST=1
    # - (tuỳ chọn) COLLABS_OUTSIDE_DEDUP_FILE=.collabs_outside_seen_hosts.json
    dedup_enabled = str(os.getenv("COLLABS_OUTSIDE_DEDUP_PERSIST", "")).strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
        "on",
        "enable",
        "enabled",
    }
    seen_hosts_persist: set[str] = set()
    dedup_path = _outside_dedup_path()
    if dedup_enabled:
        try:
            if dedup_path.exists():
                payload = json.loads(dedup_path.read_text(encoding="utf-8-sig") or "{}")
                arr = payload.get("hosts") if isinstance(payload, dict) else payload
                if isinstance(arr, list):
                    seen_hosts_persist = {host_key(x) for x in arr if host_key(x)}
        except Exception:
            seen_hosts_persist = set()

    offers: list[dict] = []
    seen_host: set[str] = set()
    added_hosts: set[str] = set()
    work_urls: list[str] = []
    stopped_early = False
    for url in candidates:
        if should_stop and should_stop():
            stopped_early = True
            break
        hk = host_key(url)
        if not hk or hk in seen_host:
            continue
        if dedup_enabled and hk in seen_hosts_persist:
            continue
        seen_host.add(hk)
        work_urls.append(url)

    nw = int(os.getenv("COLLABS_OUTSIDE_VERIFY_WORKERS", "8") or "8")
    nw = max(1, min(16, nw))

    if work_urls:
        if nw <= 1 or len(work_urls) <= 1:
            for u in work_urls:
                if should_stop and should_stop():
                    stopped_early = True
                    break
                off = _outside_discovery_verify_and_build_offer(u, verify_timeout_sec, should_stop)
                if off:
                    offers.append(off)
                    if dedup_enabled:
                        added_hosts.add(host_key(off.get("url", "") or ""))
        else:
            ex = ThreadPoolExecutor(max_workers=nw)
            try:
                futures = {
                    ex.submit(_outside_discovery_verify_and_build_offer, u, verify_timeout_sec, should_stop): i
                    for i, u in enumerate(work_urls)
                }
                pending = set(futures.keys())
                slots: list[dict | None] = [None] * len(work_urls)
                while pending:
                    done, pending = wait(pending, timeout=0.25, return_when=FIRST_COMPLETED)
                    for fut in done:
                        idx = futures.pop(fut, -1)
                        if idx < 0:
                            continue
                        try:
                            slots[idx] = fut.result()
                        except Exception:
                            slots[idx] = None
                    if should_stop and should_stop():
                        stopped_early = True
                        break
                for slot in slots:
                    if not slot:
                        continue
                    offers.append(slot)
                    if dedup_enabled:
                        added_hosts.add(host_key(slot.get("url", "") or ""))
            finally:
                ex.shutdown(wait=False, cancel_futures=True)

    if dedup_enabled and added_hosts:
        try:
            merged = sorted(set(seen_hosts_persist) | set(added_hosts))
            # giữ tối đa 50k hosts để tránh file quá lớn
            merged = merged[-50000:]
            dedup_path.write_text(
                json.dumps({"updatedAt": time.time(), "hosts": merged}, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass
    # Chỉ cập nhật cursor khi chạy xong batch (không hủy giữa chừng) để không bỏ lỡ URL kế tiếp.
    if not stopped_early and not (should_stop and should_stop()):
        _save_outside_cursor(provider, query_text, end_idx % effective_pool)
    if isinstance(f, dict):
        f[OUTSIDE_DISCOVERY_STATS_KEY] = {
            "target_batch": max_results,
            "candidate_urls": len(candidates),
            "offers_verified": len(offers),
        }
    return offers


# Trạng thái traffic so với ngưỡng (CSV + log)
STATUS_TRAFFIC_OK = "ĐẠT"
STATUS_TRAFFIC_FAIL = "CHƯA ĐẠT"


def format_uppromote_cookie_days_cell(raw) -> str:
    """Excel: ví dụ 30 → \"30 day\"; giữ nguyên nếu đã có chữ day/days."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    low = s.lower()
    if "day" in low:
        return s
    m = re.match(r"^(\d+(?:[.,]\d+)?)", s.replace(",", "."))
    if m:
        num = m.group(1).replace(",", ".")
        return f"{num} day"
    return f"{s} day"


def format_percent_cell(raw) -> str:
    """Excel: hiển thị tỷ lệ kèm %, ví dụ 20%. Số float trong (0,1) coi như phân số → nhân 100."""
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return ""
    if isinstance(raw, (int, float)):
        x = float(raw)
        if isinstance(raw, float) and 0 < x < 1:
            x *= 100
        if abs(x - round(x)) < 1e-6:
            return f"{int(round(x))}%"
        t = f"{x:.4f}".rstrip("0").rstrip(".")
        return f"{t}%"
    s = str(raw).strip()
    if not s:
        return ""
    if "%" in s:
        return s
    s2 = s.replace(",", ".").replace("%", "")
    try:
        x = float(s2)
        if 0 < x < 1:
            x *= 100
        if abs(x - round(x)) < 1e-6:
            return f"{int(round(x))}%"
        t = f"{x:.4f}".rstrip("0").rstrip(".")
        return f"{t}%"
    except ValueError:
        return f"{s}%"


# CSV Uppromote (tiếng Việt)
UPPROMOTE_CSV_HEADER_VI = [
    "Trạng thái",
    "Thương hiệu",
    "Website",
    "URL apply",
    "Hoa hồng",
    "Ngày cookie",
    "Danh mục",
    "Traffic (hiển thị)",
    "Traffic ước tính/tháng",
    "Trang/lượt xem",
    "Tỷ lệ thanh toán",
    "Tỷ lệ duyệt",
    "Tỷ lệ thoát",
    "Top quốc gia",
    "Top từ khóa",
    "Tiền tệ",
    "Điểm offer",
    "Điểm gợi ý",
    "Duyệt đơn",
    "Chu kỳ thanh toán",
    "Chi tiết khuyến mãi",
    "Kênh được phép",
    "Đối tượng vị trí",
    "Đối tượng độ tuổi",
    "Đối tượng giới",
    "Có thể apply",
    "Đã apply",
    "Offer ID",
    "Shop ID",
    "Program ID",
    "Marketplace Listing ID",
    "EPC (TB/đơn)",
]

# CSV Goaff: chỉ cột Similarweb (Apify) + trường có trong API Goaff (không cột Uppromote rỗng)
GOAFF_CSV_HEADER = [
    "Trạng thái",
    "Thương hiệu",
    "Website",
    "Số tiền HH",
    "Ngày cookie",
    "Traffic (hiển thị)",
    "Traffic ước tính/tháng",
    "Link đăng ký",
    "Trang/lượt xem",
    "Duyệt tự động",
    "Top từ khóa",
    "Tỷ lệ thoát",
    "Top quốc gia",
    "Tiền tệ",
    "Tên (API)",
    "Đăng ký mở",
    "Loại hoa hồng",
    "Hoa hồng trên",
    "ID cửa hàng",
]

REFERSION_CSV_HEADER = [
    "Trạng thái",
    "Thương hiệu",
    "Website",
    "URL apply",
    "Hoa hồng",
    "Ngày cookie",
    "Danh mục",
    "Hoạt động",
    "Đang tạm dừng",
    "Traffic (hiển thị)",
    "Traffic ước tính/tháng",
    "Trang/lượt xem",
    "Tỷ lệ thoát",
    "Top quốc gia",
    "Top từ khóa",
    "Loại hoa hồng",
    "EPC",
    "Thanh toán",
    "Tiền tệ",
    "ID Refersion",
    "Client ID",
    "Offer ID",
]

COLLABS_CSV_HEADER = [
    "Trạng thái",
    "Thương hiệu",
    "Website",
    "Link Apply",
    "Hoa hồng mạng lưới",
    "Thời gian giữ đơn",
    "Danh mục",
    "Đã lưu",
    "Quốc gia mục tiêu",
    "Traffic (hiển thị)",
    "Traffic ước tính/tháng",
    "Trang/lượt xem",
    "Tỷ lệ thoát",
    "Top quốc gia",
    "Top từ khóa",
    "Link logo",
    "Offer ID",
    "ID Shopify Store",
]


def build_uppromote_csv_row_vi(offer: dict, item: dict, status: str) -> list:
    eng = engagement_from_item(item)
    estimated_monthly = estimated_monthly_visits_formatted(item, eng)
    return [
        status,
        offer.get("brand", ""),
        offer.get("url", ""),
        offer.get("client_url", ""),
        offer.get("offer", ""),
        format_uppromote_cookie_days_cell(offer.get("cookieDays", "")),
        offer.get("category", ""),
        eng.get("VisitsFormatted", ""),
        estimated_monthly,
        eng.get("PagePerVisit", ""),
        format_percent_cell(offer.get("payout_rate", "")),
        format_percent_cell(offer.get("approval_rate", "")),
        format_percent_cell(eng.get("BounceRate", "")),
        top_countries_csv(item.get("TopCountryShares") or []),
        top_keywords_csv(keyword_shares_from_item(item)),
        offer.get("currency", ""),
        offer.get("offer_score", ""),
        offer.get("recommend_score", ""),
        offer.get("application_review", ""),
        offer.get("payments", ""),
        join_list(offer.get("promotion_details")),
        join_list(offer.get("target_audience_customer_channels")),
        join_list(offer.get("target_audience_locations")),
        join_list(offer.get("target_audience_ages")),
        join_list(offer.get("target_audience_genders")),
        offer.get("can_apply_offer", ""),
        offer.get("is_applied_offer", ""),
        offer.get("offer_id", ""),
        offer.get("shop_id", ""),
        offer.get("program_id", ""),
        offer.get("mkp_listing_id", ""),
        offer.get("epc", ""),
    ]


def build_goaff_csv_row(offer: dict, item: dict, status: str) -> list:
    url = offer.get("url", "")
    eng = engagement_from_item(item)
    estimated_monthly = estimated_monthly_visits_formatted(item, eng)
    cookie_days = offer.get("cookieDays", "")
    cookie_days_text = str(cookie_days).strip() if cookie_days is not None else ""
    if cookie_days_text and cookie_days_text.replace(".", "", 1).isdigit():
        cookie_days_text = f"{cookie_days_text} day"
    return [
        status,
        offer.get("brand", ""),
        url,
        format_goaff_commission_amount_display(offer),
        cookie_days_text,
        eng.get("VisitsFormatted", ""),
        estimated_monthly,
        goaff_create_account_url(offer),
        eng.get("PagePerVisit", ""),
        fmt_yes_no_01(offer.get("goaff_is_approved_automatically")),
        top_keywords_csv(keyword_shares_from_item(item)),
        format_percent_cell(eng.get("BounceRate", "")),
        top_countries_csv(item.get("TopCountryShares") or []),
        offer.get("currency", ""),
        offer.get("goaff_name", ""),
        fmt_yes_no_01(offer.get("goaff_are_registrations_open")),
        offer.get("goaff_commission_type", ""),
        offer.get("goaff_commission_on", ""),
        offer.get("goaff_id", ""),
    ]


def build_refersion_csv_row(offer: dict, item: dict, status: str) -> list:
    eng = engagement_from_item(item)
    estimated_monthly = estimated_monthly_visits_formatted(item, eng)
    return [
        status,
        offer.get("brand", ""),
        offer.get("url", ""),
        offer.get("client_url", ""),
        offer.get("offer", ""),
        format_uppromote_cookie_days_cell(offer.get("cookieDays", "")),
        offer.get("category", ""),
        refersion_active_from_denied(offer.get("refersion_denied")),
        fmt_yes_no_bool_like(offer.get("refersion_pending")),
        eng.get("VisitsFormatted", ""),
        estimated_monthly,
        eng.get("PagePerVisit", ""),
        format_percent_cell(eng.get("BounceRate", "")),
        top_countries_csv(item.get("TopCountryShares") or []),
        top_keywords_csv(keyword_shares_from_item(item)),
        offer.get("commission_type", ""),
        offer.get("epc", ""),
        offer.get("payments", ""),
        offer.get("currency", ""),
        offer.get("refersion_id", ""),
        offer.get("refersion_client_id", ""),
        offer.get("refersion_offer_id", ""),
    ]


def build_collabs_csv_row(offer: dict, item: dict, status: str) -> list:
    eng = engagement_from_item(item)
    estimated_monthly = estimated_monthly_visits_formatted(item, eng)
    visits_display = visits_formatted_from_engagement(eng)
    website = str(offer.get("url", "") or "").strip()
    apply_url = str(offer.get("client_url", "") or "").strip()
    if not apply_url and website:
        p = urlparse(website if "://" in website else f"https://{website}")
        if p.netloc:
            base = f"{p.scheme or 'https'}://{p.netloc}"
            apply_url = f"{base}/pages/collab"
    return [
        status,
        offer.get("brand", ""),
        website,
        apply_url,
        offer.get("offer", ""),
        format_collabs_holding_period(offer.get("collabs_holding_period", "")),
        offer.get("category", ""),
        fmt_yes_no_bool_like(offer.get("collabs_saved")),
        offer.get("collabs_target_countries", ""),
        visits_display,
        estimated_monthly,
        eng.get("PagePerVisit", ""),
        format_percent_cell(eng.get("BounceRate", "")),
        top_countries_csv(item.get("TopCountryShares") or []),
        top_keywords_csv(keyword_shares_from_item(item)),
        offer.get("collabs_logo_url", ""),
        offer.get("offer_id", ""),
        offer.get("collabs_shopify_store_id", ""),
    ]


def write_xlsx_highlight_status(path: Path, header: list, rows: list, status_col: int = 0) -> None:
    """Ghi file Excel: dòng có trạng thái ĐẠT (hoặc GET) được tô nền xanh lá nhạt."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.append(list(header))
    header_list = list(header)
    hyperlink_columns = {
        i + 1
        for i, name in enumerate(header_list)
        if str(name).strip() in {"Website", "URL apply", "Link đăng ký", "Link Apply"}
    }
    header_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for c in range(1, len(header_list) + 1):
        ws.cell(row=1, column=c).fill = header_fill
    ok_values = {STATUS_TRAFFIC_OK, "GET", "ĐẠT"}
    for row in rows:
        cells = [
            ("N/A" if (v is None or (isinstance(v, str) and not v.strip())) else v)
            for v in list(row)
        ]
        ws.append(cells)
        r = ws.max_row
        for c in hyperlink_columns:
            if c <= len(cells):
                cell = ws.cell(row=r, column=c)
                value = str(cell.value or "").strip()
                if value.startswith("http://") or value.startswith("https://"):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
        if len(cells) > status_col and str(cells[status_col]).strip() in ok_values:
            for c in range(1, len(cells) + 1):
                ws.cell(row=r, column=c).fill = green
    wb.save(path)


def fetch_goaffpro_page(base_url: str, offset: int, limit: int) -> dict:
    request_url = with_goaffpro_paging(base_url, offset, limit)
    res = requests.get(request_url, headers=build_goaffpro_headers(), timeout=60)
    text = res.text
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError(f"Goaffpro parse JSON lỗi (HTTP {res.status_code}): {text[:180]}") from exc
    if not res.ok:
        raise RuntimeError(f"Goaffpro HTTP {res.status_code}: {text[:300]}")
    return body if isinstance(body, dict) else {}


def with_refersion_page(url: str, page: int) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["page"] = str(page)
    new_query = urlencode(query, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def fetch_refersion_page(base_url: str, page: int) -> dict:
    request_url = with_refersion_page(base_url, page)
    res = requests.get(request_url, headers=build_refersion_headers(), timeout=60)
    text = res.text
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError(f"Refersion parse JSON lỗi (HTTP {res.status_code}): {text[:180]}") from exc
    if not res.ok:
        raise RuntimeError(f"Refersion HTTP {res.status_code}: {text[:300]}")
    if str(body.get("status") or "").lower() != "success":
        raise RuntimeError(f"Refersion API lỗi: {text[:300]}")
    return body if isinstance(body, dict) else {}


def fetch_collabs_page(base_url: str, first: int, after: str | None = None) -> dict:
    payload = {
        "operationName": "BrandsQuery",
        "query": COLLABS_BRANDS_QUERY,
        "variables": {
            "first": int(first),
            "productCategories": [],
            "brandValues": [],
        },
    }
    if after:
        payload["variables"]["after"] = str(after)
    res = requests.post(base_url, headers=build_collabs_headers(), json=payload, timeout=60)
    text = res.text
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError(f"Collabs parse JSON lỗi (HTTP {res.status_code}): {text[:180]}") from exc
    if not res.ok:
        raise RuntimeError(f"Collabs HTTP {res.status_code}: {text[:300]}")
    if body.get("errors"):
        raise RuntimeError(f"Collabs GraphQL lỗi: {json.dumps(body.get('errors'), ensure_ascii=False)[:300]}")
    return body if isinstance(body, dict) else {}


def fetch_collabs_brand_profile(base_url: str, shopify_store_gid: str) -> dict:
    gid = str(shopify_store_gid or "").strip()
    if not gid:
        return {}
    payload = {
        "operationName": "DiscoverBrandProfileQuery",
        "query": COLLABS_BRAND_PROFILE_QUERY,
        "variables": {
            "shopifyStoreId": gid,
        },
    }
    res = requests.post(base_url, headers=build_collabs_headers(), json=payload, timeout=60)
    text = res.text
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError(f"Collabs detail parse JSON lỗi (HTTP {res.status_code}): {text[:180]}") from exc
    if not res.ok:
        raise RuntimeError(f"Collabs detail HTTP {res.status_code}: {text[:300]}")
    if body.get("errors"):
        raise RuntimeError(f"Collabs detail GraphQL lỗi: {json.dumps(body.get('errors'), ensure_ascii=False)[:300]}")
    data = body.get("data") if isinstance(body, dict) else {}
    return data if isinstance(data, dict) else {}


def fetch_all_goaffpro_offers() -> list:
    base_url = (os.getenv("GOAFFPRO_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu GOAFFPRO_API_URL trong .env")

    enforce_fixed_fetch_defaults()
    limit = int(os.getenv("GOAFFPRO_LIMIT", str(DEFAULT_OFFERS_PER_PAGE)) or str(DEFAULT_OFFERS_PER_PAGE))
    max_pages_cap = goaffpro_max_pages_cap()
    delay_ms = int(
        os.getenv("GOAFFPRO_PAGE_DELAY_MS", str(DEFAULT_GOAFFPRO_PAGE_DELAY_MS))
        or str(DEFAULT_GOAFFPRO_PAGE_DELAY_MS)
    )

    all_stores = []
    page = 1
    while True:
        offset = (page - 1) * limit
        print(f"Goaffpro: tải offset={offset} (trang {page})...")
        body = fetch_goaffpro_page(base_url, offset, limit)
        page_items = body.get("stores") or []
        if not isinstance(page_items, list):
            page_items = []

        if not page_items:
            print(f"Goaffpro: không còn store — kết thúc phân trang.")
            break

        all_stores.extend(page_items)
        print(f"Goaffpro: +{len(page_items)} store (lũy kế {len(all_stores)})")

        if max_pages_cap is not None and page >= max_pages_cap:
            print(f"Goaffpro: dừng vì GOAFFPRO_MAX_PAGES={max_pages_cap}")
            break

        total_count = body.get("count")
        try:
            total_n = int(total_count) if total_count is not None else None
        except Exception:
            total_n = None
        if total_n is not None and offset + len(page_items) >= total_n:
            break

        if len(page_items) < limit:
            break

        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)

    return [map_goaffpro_store(s) for s in all_stores]


def fetch_all_refersion_offers() -> list:
    base_url = (os.getenv("REFERSION_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu REFERSION_API_URL trong .env")
    enforce_fixed_fetch_defaults()
    max_pages_cap = refersion_max_pages_cap()
    delay_ms = int(
        os.getenv("REFERSION_PAGE_DELAY_MS", str(DEFAULT_REFERSION_PAGE_DELAY_MS))
        or str(DEFAULT_REFERSION_PAGE_DELAY_MS)
    )
    all_offers = []
    page = 1
    while True:
        print(f"Refersion: tải trang {page}...")
        body = fetch_refersion_page(base_url, page)
        payload = body.get("data") or {}
        page_items = payload.get("offers") or []
        if not isinstance(page_items, list):
            page_items = []
        if not page_items:
            print("Refersion: không còn offer — kết thúc phân trang.")
            break
        all_offers.extend(page_items)
        print(f"Refersion: +{len(page_items)} offer (lũy kế {len(all_offers)})")
        if max_pages_cap is not None and page >= max_pages_cap:
            print(f"Refersion: dừng vì REFERSION_MAX_PAGES={max_pages_cap}")
            break
        total_results = payload.get("total_results")
        try:
            total_n = int(total_results) if total_results is not None else None
        except Exception:
            total_n = None
        if total_n is not None and len(all_offers) >= total_n:
            break
        if len(page_items) == 0:
            break
        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)
    return [map_refersion_offer(o) for o in all_offers]


def map_uppromote_offer(offer: dict, detail: dict | None = None) -> dict:
    detail = detail or {}
    website = detail.get("website")
    url = website.strip() if isinstance(website, str) and website.strip() else offer_url_from_uppromote(offer)
    return {
        "brand": detail.get("name") or offer.get("name") or "",
        "url": url,
        "offer": detail.get("commission") or offer.get("commission") or "",
        "cookieDays": detail.get("cookie") or offer.get("cookie") or "",
        "client_url": detail.get("apply_url") or offer.get("apply_url") or "",
        "offer_id": offer.get("id") or "",
        "shop_id": offer.get("shop_id") or detail.get("shop_id") or "",
        "program_id": offer.get("program_id") or detail.get("program_id") or "",
        "mkp_listing_id": detail.get("mkp_listing_id") or offer.get("id") or "",
        "commission_type": detail.get("commission_type") or offer.get("commissionText") or offer.get("commission_type") or "",
        "category": detail.get("categories") or offer.get("categories") or "",
        "epc": detail.get("average_earning_per_sale") or offer.get("average_earning_per_sale") or "",
        "payments": detail.get("payout_period") or offer.get("payout_period") or "",
        "currency": detail.get("currency") or offer.get("currency") or "",
        "payout_rate": detail.get("payout_rate") or offer.get("payout_rate") or "",
        "approval_rate": detail.get("approval_rate") or offer.get("approval_rate") or "",
        "offer_score": detail.get("offer_score") or offer.get("offer_score") or "",
        "recommend_score": detail.get("recommend_score") or offer.get("recommend_score") or "",
        "application_review": detail.get("application_review") or offer.get("application_review") or "",
        "promotion_details": detail.get("promotion_details") or offer.get("promotion_details") or [],
        "target_audience_customer_channels": detail.get("target_audience_customer_channels")
        or offer.get("target_audience_customer_channels")
        or [],
        "target_audience_locations": detail.get("target_audience_locations") or offer.get("target_audience_locations") or [],
        "target_audience_ages": detail.get("target_audience_ages") or offer.get("target_audience_ages") or [],
        "target_audience_genders": detail.get("target_audience_genders") or offer.get("target_audience_genders") or [],
        "can_apply_offer": detail.get("can_apply_offer") if "can_apply_offer" in detail else offer.get("can_apply_offer"),
        "is_applied_offer": detail.get("is_applied_offer") if "is_applied_offer" in detail else offer.get("is_applied_offer"),
        "payment_support": detail.get("payment_support") or offer.get("payment_support"),
    }


def fetch_uppromote_page(base_url: str, page: int) -> dict:
    request_url = with_page(base_url, page)
    res = requests.get(request_url, headers=build_uppromote_headers(), timeout=60)
    text = res.text
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError(f"Uppromote parse JSON lỗi (HTTP {res.status_code}): {text[:180]}") from exc
    if not res.ok:
        raise RuntimeError(f"Uppromote HTTP {res.status_code}: {text[:300]}")
    if body.get("status") not in (200, "200"):
        raise RuntimeError(f"Uppromote API lỗi: {text[:300]}")
    return body


def fetch_all_uppromote_offers() -> list:
    base_url = (os.getenv("UPPROMOTE_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu UPPROMOTE_API_URL trong .env")

    enforce_fixed_fetch_defaults()
    max_pages_cap = uppromote_max_pages_cap()
    delay_ms = int(os.getenv("UPPROMOTE_PAGE_DELAY_MS", str(DEFAULT_UPPROMOTE_PAGE_DELAY_MS)) or str(DEFAULT_UPPROMOTE_PAGE_DELAY_MS))

    all_offers = []
    page = 1
    while True:
        print(f"Uppromote: tải trang {page}...")
        body = fetch_uppromote_page(base_url, page)
        payload = body.get("data") or {}
        page_items = payload.get("data") or []
        if not isinstance(page_items, list):
            page_items = []

        if not page_items:
            print(f"Uppromote: trang {page} không còn offer — kết thúc phân trang.")
            break

        all_offers.extend(page_items)
        print(f"Uppromote: +{len(page_items)} offer (lũy kế {len(all_offers)})")

        if max_pages_cap is not None and page >= max_pages_cap:
            print(f"Uppromote: dừng vì UPPROMOTE_MAX_PAGES={max_pages_cap}")
            break

        next_page = payload.get("next_page_url")
        if not next_page:
            break

        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)

    detail_delay_ms = int(os.getenv("UPPROMOTE_DETAIL_DELAY_MS", "50") or "50")
    mapped = []
    for i, offer in enumerate(all_offers, start=1):
        detail = {}
        shop_id = offer.get("shop_id")
        if shop_id:
            try:
                detail = fetch_uppromote_offer_detail(shop_id)
            except Exception as exc:
                print(f"Cảnh báo: không lấy được detail cho shop_id={shop_id}: {exc}")
            if detail_delay_ms > 0:
                time.sleep(detail_delay_ms / 1000)
        mapped_offer = map_uppromote_offer(offer, detail)
        mapped.append(mapped_offer)
        if i % 20 == 0 or i == len(all_offers):
            print(f"Uppromote detail: {i}/{len(all_offers)}")
    return mapped


def read_domains_from_txt(path: Path) -> list:
    if not path.exists():
        return []
    out = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            t = line.strip()
            if t:
                out.append(t)
    return out


def unique_hosts(items: list) -> list:
    return sorted({host_key(x) for x in items if host_key(x)})


def top_countries_csv(countries, limit=3):
    if not isinstance(countries, list):
        return ""
    values = []
    for c in countries[:limit]:
        code = c.get("CountryCode", "")
        val = c.get("Value", 0)
        try:
            pct = round(float(val), 2)
        except Exception:
            pct = 0
        values.append(f"{code} ({pct}%)")
    return ", ".join(values)


def keyword_label(k):
    return k.get("Name") or k.get("Keyword") or k.get("keyword") or k.get("Key") or k.get("key") or ""


def keyword_volume(k):
    """Khối lượng tìm kiếm (Volume: …); không dùng Traffic/Visits — traffic nằm ở EstimatedValue."""
    candidates = [
        k.get("Volume"),
        k.get("SearchVolume"),
        k.get("MonthlyVolume"),
        k.get("EstimatedMonthlySearchVolume"),
    ]
    for raw in candidates:
        if isinstance(raw, (int, float)):
            return int(round(raw))
        if isinstance(raw, str) and raw.strip():
            v = parse_visits_value(raw)
            if v > 0:
                return int(round(v))
    return 0


def keyword_traffic_from_estimated(k):
    """Traffic hiển thị trong Top Keywords: ưu tiên EstimatedValue (Apify), rồi Traffic/EstTraffic…"""
    if not isinstance(k, dict):
        return 0
    candidates = [
        k.get("EstimatedValue"),
        k.get("estimatedValue"),
        k.get("Traffic"),
        k.get("traffic"),
        k.get("EstTraffic"),
        k.get("Visits"),
    ]
    for raw in candidates:
        if raw is None:
            continue
        if isinstance(raw, str) and not raw.strip():
            continue
        v = parse_visits_value(raw)
        if abs(v - round(v)) < 1e-9:
            return int(round(v))
        return round(v, 4)
    return 0


def keyword_cpc_number_str(k):
    """Phần số CPC (không kèm $) cho định dạng Cpc:1.07$."""
    if not isinstance(k, dict):
        return "0"
    raw = k.get("Cpc") or k.get("CPC") or k.get("cpc") or k.get("EstimatedCpc") or k.get("estimatedCpc")
    if raw is None or raw == "":
        return "0"
    if isinstance(raw, (int, float)):
        x = float(raw)
    else:
        s = str(raw).strip().rstrip("$").replace(",", "").strip()
        try:
            x = float(s)
        except ValueError:
            return "0"
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    t = f"{x:.4f}".rstrip("0").rstrip(".")
    return t or "0"


def top_keywords_csv(keywords, limit=TOP_KEYWORDS_COUNT):
    if not isinstance(keywords, list):
        return ""
    values = []
    for k in keywords[:limit]:
        if isinstance(k, str):
            s = k.strip()
            if s:
                values.append(s)
            continue
        if not isinstance(k, dict):
            continue
        label = keyword_label(k)
        if not label:
            continue
        vol = keyword_volume(k)
        traf = keyword_traffic_from_estimated(k)
        cpc = keyword_cpc_number_str(k)
        values.append(f"{label} (Volume: {vol}, Traffic: {traf}, Cpc:{cpc}$)")
    return ", ".join(values)


def join_list(value, sep="; "):
    if value is None:
        return ""
    if isinstance(value, list):
        return sep.join([str(x).strip() for x in value if str(x).strip()])
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return ""


def keyword_shares_from_item(item):
    eng = item.get("Engagments") or item.get("Engagements") or {}
    candidates = [
        item.get("TopKeywordShares"),
        item.get("TopKeywords"),
        eng.get("TopKeywordShares"),
        eng.get("TopKeywords"),
        item.get("TopOrganicKeywordShares"),
        item.get("TopOrganicKeywords"),
        eng.get("TopOrganicKeywordShares"),
        eng.get("TopOrganicKeywords"),
    ]
    for arr in candidates:
        if isinstance(arr, list) and arr:
            return arr
        if isinstance(arr, str) and arr.strip():
            return [arr.strip()]
    return []


def load_brand_average_prices() -> dict:
    path = BASE_DIR / "brand-prices.csv"
    if not path.exists():
        print("Không tìm thấy brand-prices.csv — cột avg_price và est_commission_per_sale sẽ để trống.")
        return {}
    result = {}
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            t = line.strip()
            if not t or t.startswith("#"):
                continue
            parts = [x.strip() for x in t.split(",")]
            if len(parts) < 2:
                continue
            brand = parts[0].lower()
            num = parts[1].replace(",", "")
            try:
                result[brand] = float(num)
            except Exception:
                continue
    print(f"Đã load {len(result)} brand avg price từ brand-prices.csv")
    return result


def _normalize_apify_actor_id(actor_id: str) -> str:
    act = str(actor_id or "").strip()
    if "/" in act and "~" not in act:
        parts = [p for p in act.split("/") if p]
        if len(parts) >= 2:
            act = f"{parts[0]}~{parts[1]}"
    return act


def _apify_store_actor_dataset_id(actor_id: str, run_input: dict, *, wait_secs: int | None = None) -> str:
    """
    Chạy actor Apify Store (POST + poll), trả về defaultDatasetId.
    Dùng chung cho actor Similarweb mặc định và actor fallback outside discovery.
    """
    token = (os.getenv("APIFY_TOKEN") or "").strip()
    if not token:
        raise RuntimeError("Thiếu APIFY_TOKEN trong .env")
    ws = wait_secs if wait_secs is not None else int(os.getenv("APIFY_WAIT_FOR_FINISH_SECS", "240") or "240")
    http_retries = int(os.getenv("APIFY_HTTP_RETRIES", "3") or "3")
    if http_retries < 1:
        http_retries = 1
    connect_timeout_sec = float(os.getenv("APIFY_CONNECT_TIMEOUT_SECS", "20") or "20")
    if connect_timeout_sec < 3:
        connect_timeout_sec = 3.0

    def _request_with_retry(method: str, req_url: str, *, json_payload=None, read_timeout_sec: float = 60.0):
        last_exc: Exception | None = None
        for attempt in range(1, http_retries + 1):
            try:
                return requests.request(
                    method=method,
                    url=req_url,
                    json=json_payload,
                    timeout=(connect_timeout_sec, max(5.0, float(read_timeout_sec))),
                )
            except requests.RequestException as exc:
                last_exc = exc
                if attempt >= http_retries:
                    break
                sleep_sec = min(6.0, 1.2 * attempt)
                print(f"Apify request retry {attempt}/{http_retries} sau lỗi: {exc}")
                time.sleep(sleep_sec)
        raise RuntimeError(f"Apify request thất bại sau {http_retries} lần thử: {last_exc}")

    act = _normalize_apify_actor_id(actor_id)
    if not act:
        raise RuntimeError("Thiếu actor id Apify.")
    url = f"https://api.apify.com/v2/acts/{act}/runs?token={token}&waitForFinish={int(ws)}"
    res = _request_with_retry("POST", url, json_payload=run_input, read_timeout_sec=float(ws) + 30.0)
    if not res.ok:
        raise RuntimeError(f"Apify call actor lỗi HTTP {res.status_code}: {res.text[:300]}")
    body = res.json()
    data = body.get("data") or {}
    run_id = data.get("id")
    if not run_id:
        raise RuntimeError("Apify không trả về run id")

    status = data.get("status")
    wait_poll_secs = int(os.getenv("APIFY_POLL_WAIT_SECS", "120") or "120")
    max_polls = int(os.getenv("APIFY_MAX_POLLS", "10") or "10")
    poll_idx = 0
    while status not in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
        poll_idx += 1
        if poll_idx > max_polls:
            raise RuntimeError(f"Apify run chưa hoàn tất sau {max_polls} lần poll (status={status})")
        poll_url = f"https://api.apify.com/v2/actor-runs/{run_id}?token={token}&waitForFinish={wait_poll_secs}"
        poll_res = _request_with_retry("GET", poll_url, read_timeout_sec=wait_poll_secs + 30)
        if not poll_res.ok:
            raise RuntimeError(f"Apify poll lỗi HTTP {poll_res.status_code}: {poll_res.text[:300]}")
        data = (poll_res.json() or {}).get("data") or data
        status = data.get("status")
        print(f"Apify poll {poll_idx}: status={status}")

    if status != "SUCCEEDED":
        raise RuntimeError(f"Apify run không thành công (status={status})")

    dataset_id = data.get("defaultDatasetId")
    if not dataset_id:
        raise RuntimeError("Apify không trả về defaultDatasetId")
    return str(dataset_id)


def apify_call_actor(domains: list) -> str:
    if not domains:
        raise RuntimeError("Danh sách domain Apify rỗng")
    return _apify_store_actor_dataset_id(
        ACTOR_ID,
        {"domains": list(domains), "proxyConfiguration": {"useApifyProxy": False}},
    )


def _similarweb_monthly_text_has_positive_visits(text: str) -> bool:
    """Chuỗi dạng T1(577), T2(0) — chỉ coi là có traffic nếu có ít nhất một số trong ngoặc > 0."""
    if not text or not str(text).strip():
        return False
    for m in re.finditer(r"\(([^)]+)\)", str(text)):
        if parse_visits_value(m.group(1)) > 0:
            return True
    return False


def _similarweb_item_has_traffic(item: dict | None) -> bool:
    """
    Có dữ liệu traffic *dùng được* (số visits > 0 hoặc monthly có tháng > 0).
    Tránh coi chuỗi monthly chỉ toàn 0 / VisitsFormatted \"0\" là \"đã có traffic\" — khi đó vẫn cần fallback Radeance.
    """
    if not item:
        return False
    eng = engagement_from_item(item)
    if parse_visits_from_engagement(eng) > 0:
        return True
    em = str(estimated_monthly_visits_formatted(item, eng) or "").strip()
    if em and _similarweb_monthly_text_has_positive_visits(em):
        return True
    vf = str(visits_formatted_from_engagement(eng) or "").strip()
    if vf and parse_visits_value(vf) > 0:
        return True
    return False


def _format_visits_integer_display(v: float) -> str:
    try:
        n = int(round(float(v)))
    except Exception:
        return ""
    if n <= 0:
        return ""
    return f"{n:,}"


def _radeance_dataset_row_host_keys(raw: dict) -> set[str]:
    """Các host chuẩn hoá có thể suy ra từ một dòng dataset radeance/similarweb-scraper."""
    low = {str(k).strip().lower(): v for k, v in raw.items()}
    keys: set[str] = set()
    for cand in (
        raw.get("searchUrl"),
        raw.get("url"),
        raw.get("domain"),
        raw.get("website"),
        low.get("searchurl"),
        low.get("url"),
        low.get("domain"),
        low.get("website"),
    ):
        if isinstance(cand, str) and cand.strip():
            hk = host_key(cand.strip())
            if hk:
                keys.add(hk)
    return keys


def normalize_radeance_similarweb_item(raw: dict) -> dict:
    """
    Chuẩn hoá một dòng dataset từ radeance/similarweb-scraper về shape gần actor Similarweb cũ
    (Engagements, TopCountryShares, TopKeywordShares, EstimatedMonthlyVisits) để build_collabs_csv_row / lọc traffic hoạt động.
    """
    if not isinstance(raw, dict) or not raw:
        return {}
    low = {str(k).strip().lower(): v for k, v in raw.items()}
    domain = str(raw.get("domain") or low.get("domain") or "").strip()
    url = str(raw.get("url") or raw.get("searchUrl") or low.get("url") or low.get("searchurl") or "").strip()
    if not url and domain:
        d = domain.lstrip("/")
        url = f"https://{d}" if d else ""
    eng_in = raw.get("engagement") or raw.get("Engagement") or {}
    if not isinstance(eng_in, dict):
        eng_in = {}
    total_vis = (
        raw.get("totalVisits")
        or raw.get("TotalVisits")
        or low.get("totalvisits")
        or raw.get("Visits")
        or low.get("visits")
    )
    if total_vis is None:
        total_vis = eng_in.get("visits") or eng_in.get("Visits")
    visits_f = parse_visits_value(total_vis)
    if visits_f <= 0:
        for ak in (
            "traffic",
            "Traffic",
            "estimatedTraffic",
            "EstimatedTraffic",
            "estimatedVisits",
            "EstimatedVisits",
            "lastMonthVisits",
            "LastMonthVisits",
            "pageViews",
            "PageViews",
            "globalVisits",
            "GlobalVisits",
        ):
            v = raw.get(ak)
            if v is None:
                v = low.get(ak.lower())
            if v is None:
                continue
            pv = parse_visits_value(v)
            if pv > visits_f:
                visits_f = pv
    mv_arr = raw.get("monthlyVisits") or low.get("monthlyvisits")
    if visits_f <= 0 and isinstance(mv_arr, list) and mv_arr:
        best_mv = 0.0
        for row in mv_arr:
            if isinstance(row, dict):
                lv = row.get("visits") or row.get("Visits")
                if lv is not None:
                    best_mv = max(best_mv, parse_visits_value(lv))
        if best_mv > visits_f:
            visits_f = best_mv
    ppg = raw.get("pagesPerVisit")
    if ppg is None:
        ppg = eng_in.get("pagesPerVisit")
    br = raw.get("bounceRate")
    if br is None:
        br = eng_in.get("bounceRate")
    monthly_df = raw.get("monthlyVisitsDateFormat") or low.get("monthlyvisitsdateformat") or {}
    if isinstance(monthly_df, str) and monthly_df.strip().startswith("{"):
        try:
            monthly_df = json.loads(monthly_df)
        except Exception:
            monthly_df = {}
    if not isinstance(monthly_df, dict):
        monthly_df = {}
    if visits_f <= 0 and monthly_df:
        try:
            mx = max((parse_visits_value(v) for v in monthly_df.values()), default=0.0)
            if mx > visits_f:
                visits_f = mx
        except Exception:
            pass

    top_country_shares: list[dict] = []
    tcc_raw = raw.get("website_traffic_by_country") or raw.get("countryShare") or []
    if isinstance(tcc_raw, list):
        for row in tcc_raw[:12]:
            if not isinstance(row, dict):
                continue
            code = str(row.get("country") or row.get("CountryCode") or "").strip().upper()
            if not code:
                continue
            sh = row.get("share")
            if sh is None:
                sh = row.get("Value")
            try:
                fv = float(sh)
            except Exception:
                fv = 0.0
            if 0 <= fv <= 1.0:
                fv *= 100.0
            top_country_shares.append({"CountryCode": code, "Value": fv})

    top_kw: list[dict] = []
    tk_raw = raw.get("topKeywords") or []
    if isinstance(tk_raw, list):
        for row in tk_raw:
            if not isinstance(row, dict):
                continue
            kw = row.get("keyword") or row.get("Keyword") or row.get("Name")
            if not kw:
                continue
            vol = row.get("searchVolume") or row.get("search_volume") or row.get("Volume")
            est = row.get("estimatedValue") or row.get("estimated_value") or row.get("EstimatedValue")
            cpc = row.get("cpc") if row.get("cpc") is not None else row.get("CPC")
            try:
                vol_i = int(round(float(vol))) if vol is not None and str(vol).strip() else 0
            except Exception:
                vol_i = 0
            top_kw.append(
                {
                    "Name": str(kw),
                    "Keyword": str(kw),
                    "Volume": vol_i,
                    "EstimatedValue": est,
                    "Cpc": cpc,
                }
            )

    hk_canon = host_key(url) if url else (host_key(domain) if domain else "")
    site_field = hk_canon or (domain or "").strip() or host_key(str(raw.get("searchUrl") or ""))
    return {
        "SiteName": site_field,
        "Domain": domain,
        "Url": url or (f"https://{hk_canon}" if hk_canon else ""),
        "Engagements": {
            "Visits": visits_f,
            "VisitsFormatted": _format_visits_integer_display(visits_f),
            "PagePerVisit": ppg,
            "BounceRate": br,
        },
        "EstimatedMonthlyVisits": monthly_df,
        "TopCountryShares": top_country_shares,
        "TopKeywordShares": top_kw,
        "TopKeywords": top_kw,
    }


def merge_outside_similarweb_fallback_batch(part_domains: list, batch_items: list) -> tuple[list, int]:
    """
    Sau actor Similarweb mặc định: với mỗi domain trong batch, nếu không có traffic hợp lệ
    thì gọi actor fallback (radeance/similarweb-scraper), chuẩn hoá và ghi đè vào map theo host.

    Trả về (danh sách item theo thứ tự part_domains, số domain đã gọi fallback).
    """
    if not isinstance(part_domains, list) or not part_domains:
        return (batch_items if isinstance(batch_items, list) else []), 0
    batch_items = batch_items if isinstance(batch_items, list) else []

    env_fb = os.getenv("COLLABS_OUTSIDE_SIMILARWEB_FALLBACK_ACTOR")
    if env_fb is not None and not str(env_fb).strip():
        fallback_actor = ""
    else:
        fallback_actor = (str(env_fb).strip() if env_fb else DEFAULT_OUTSIDE_SIMILARWEB_FALLBACK_ACTOR)

    by_h: dict[str, dict] = {}
    for it in batch_items:
        if not isinstance(it, dict):
            continue
        k = host_key(apify_site_field(it))
        if k:
            by_h[k] = it

    need: list[str] = []
    seen_need: set[str] = set()
    for d in part_domains:
        hk = host_key(d)
        if not hk:
            continue
        it = by_h.get(hk, {})
        if _similarweb_item_has_traffic(it):
            continue
        if hk not in seen_need:
            seen_need.add(hk)
            need.append(hk)

    if not need or not fallback_actor:
        out = []
        for d in part_domains:
            hk = host_key(d)
            if not hk:
                continue
            out.append(by_h.get(hk, {}))
        return out, 0

    urls = []
    for hk in need:
        urls.append(f"https://{hk}")

    fb_wait = int(os.getenv("COLLABS_OUTSIDE_SIMILARWEB_WAIT_SECS", "360") or "360")
    run_input = {
        "urls": urls,
        "include_base_data": True,
        "include_similar_sites": False,
        "include_indepth_data": False,
        "output_mode": "individual",
    }
    try:
        ds = _apify_store_actor_dataset_id(fallback_actor, run_input, wait_secs=fb_wait)
        fb_items = apify_list_items(ds)
    except Exception as exc:
        print(f"Cảnh báo Similarweb fallback ({fallback_actor}): {exc}")
        out = []
        for d in part_domains:
            hk = host_key(d)
            if not hk:
                continue
            out.append(by_h.get(hk, {}))
        return out, 0

    for i, raw in enumerate(fb_items):
        if not isinstance(raw, dict):
            continue
        norm = normalize_radeance_similarweb_item(raw)
        if not norm:
            continue
        row_hosts = _radeance_dataset_row_host_keys(raw)
        matched: str | None = None
        for hk in need:
            if hk in row_hosts:
                matched = hk
                break
        if matched is None and i < len(need):
            matched = need[i]
        nk = host_key(apify_site_field(norm))
        if matched:
            by_h[matched] = norm
        if nk:
            by_h[nk] = norm

    out = []
    for d in part_domains:
        hk = host_key(d)
        if not hk:
            continue
        out.append(by_h.get(hk, {}))
    return out, len(need)


def check_apify_connection() -> str:
    """Kiểm tra kết nối/tính hợp lệ APIFY_TOKEN trước khi chạy lọc."""
    token = (os.getenv("APIFY_TOKEN") or "").strip()
    if not token:
        raise RuntimeError("Thiếu APIFY_TOKEN trong .env")
    url = f"https://api.apify.com/v2/users/me?token={token}"
    try:
        res = requests.get(url, timeout=20)
    except requests.RequestException as exc:
        raise RuntimeError(f"Không kết nối được Apify: {exc}") from exc
    if not res.ok:
        raise RuntimeError(f"Apify connection check lỗi HTTP {res.status_code}: {res.text[:300]}")
    try:
        body = res.json()
    except Exception as exc:
        raise RuntimeError("Apify connection check trả về dữ liệu không hợp lệ.") from exc
    data = body.get("data") or {}
    username = (
        str(data.get("username") or "").strip()
        or str(data.get("email") or "").strip()
        or str(data.get("id") or "").strip()
        or "unknown-user"
    )
    return username


def apify_list_items(dataset_id: str, token: str | None = None) -> list:
    tok = (str(token or "").strip() or (os.getenv("APIFY_TOKEN") or "").strip())
    http_retries = int(os.getenv("APIFY_HTTP_RETRIES", "3") or "3")
    if http_retries < 1:
        http_retries = 1
    connect_timeout_sec = float(os.getenv("APIFY_CONNECT_TIMEOUT_SECS", "20") or "20")
    if connect_timeout_sec < 3:
        connect_timeout_sec = 3.0

    def _get_with_retry(req_url: str, read_timeout_sec: float):
        last_exc: Exception | None = None
        for attempt in range(1, http_retries + 1):
            try:
                return requests.get(req_url, timeout=(connect_timeout_sec, max(5.0, float(read_timeout_sec))))
            except requests.RequestException as exc:
                last_exc = exc
                if attempt >= http_retries:
                    break
                time.sleep(min(6.0, 1.2 * attempt))
        raise RuntimeError(f"Apify list items request thất bại sau {http_retries} lần thử: {last_exc}")

    items = []
    offset = 0
    limit = 1000
    while True:
        url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={tok}&offset={offset}&limit={limit}&clean=true"
        res = _get_with_retry(url, read_timeout_sec=120)
        if not res.ok:
            raise RuntimeError(f"Apify list items lỗi HTTP {res.status_code}: {res.text[:300]}")
        chunk = res.json()
        if not isinstance(chunk, list) or not chunk:
            break
        items.extend(chunk)
        if len(chunk) < limit:
            break
        offset += limit
    return items


def chunked(items: list, size: int):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main():
    out_path = BASE_DIR / "result.csv"

    offers = fetch_all_uppromote_offers()
    print(f"Loaded {len(offers)} offers (Uppromote)")

    snapshot_path = BASE_DIR / "uppromote-offers-last.json"
    with snapshot_path.open("w", encoding="utf-8") as f:
        json.dump({"fetchedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()), "count": len(offers), "offers": offers}, f, ensure_ascii=False)
    print(f"Đã lưu snapshot offers -> {snapshot_path}")

    from_offers = unique_hosts([o.get("url", "") for o in offers])
    domain_file = Path(os.getenv("DOMAIN_FILE", str(BASE_DIR / "domain.txt")))
    from_file = read_domains_from_txt(domain_file)
    domains = sorted(set(from_offers + from_file))
    if not domains:
        raise RuntimeError("Không có domain: thêm url trong offers Uppromote hoặc tạo domain.txt.")
    print(f"Apify sẽ chạy trên {len(domains)} domain (từ offer + domain.txt nếu có).")

    print(f"Chạy Actor {ACTOR_ID} với {len(domains)} domain...")
    items = []
    max_domains_per_run = int(
        os.getenv("APIFY_MAX_DOMAINS_PER_RUN", str(DEFAULT_APIFY_MAX_DOMAINS_PER_RUN))
        or str(DEFAULT_APIFY_MAX_DOMAINS_PER_RUN)
    )
    for idx, part in enumerate(chunked(domains, max_domains_per_run), start=1):
        print(f"Apify batch {idx}: {len(part)} domains")
        dataset_id = apify_call_actor(part)
        items.extend(apify_list_items(dataset_id))

    by_host = {}
    for item in items:
        site = apify_site_field(item)
        key = host_key(site)
        if key:
            by_host[key] = item

    header = [
        "Status",
        "Brand",
        "Website",
        "Domain",
        "Traffic Formatted",
        "Traffic Raw",
        "Pages Per Visit",
        "Bounce Rate",
        "Top Countries",
        "Top Keywords",
        "Commission",
        "Commission Type",
        "Cookie Days",
        "Currency",
        "Category",
        "Payout Rate",
        "Approval Rate",
        "Offer Score",
        "Recommend Score",
        "Application Review",
        "Payout Period",
        "Promotion Details",
        "Allowed Channels",
        "Target Locations",
        "Target Ages",
        "Target Genders",
        "Can Apply",
        "Is Applied",
        "Apply URL",
        "Offer ID",
        "Shop ID",
        "Program ID",
        "Marketplace Listing ID",
        "EPC Average Earning Per Sale",
    ]

    try:
        csv_handle = out_path.open("w", encoding="utf-8", newline="")
    except PermissionError:
        fallback = BASE_DIR / f"uppromote_{int(time.time())}.csv"
        print(f"Cảnh báo: {out_path.name} đang bị khóa, ghi sang {fallback.name}")
        out_path = fallback
        csv_handle = out_path.open("w", encoding="utf-8", newline="")

    with csv_handle as f:
        writer = csv.writer(f)
        writer.writerow(header)

        rows = 0
        for offer in offers:
            brand = offer.get("brand", "")
            url = offer.get("url", "")
            key = host_key(url)
            item = lookup_apify_item(url, by_host)
            eng = engagement_from_item(item)
            visits = parse_visits_from_engagement(eng)
            status = "GET" if visits >= float(MIN_VISITS) else "NO"

            commission_str = str(offer.get("offer", "") or "")

            countries = item.get("TopCountryShares") or []
            keyword_shares = keyword_shares_from_item(item)

            writer.writerow(
                [
                    status,
                    brand,
                    url,
                    key,
                    eng.get("VisitsFormatted", ""),
                    int(visits) if visits.is_integer() else visits,
                    eng.get("PagePerVisit", ""),
                    eng.get("BounceRate", ""),
                    top_countries_csv(countries),
                    top_keywords_csv(keyword_shares),
                    commission_str,
                    offer.get("commission_type", ""),
                    offer.get("cookieDays", ""),
                    offer.get("currency", ""),
                    offer.get("category", ""),
                    offer.get("payout_rate", ""),
                    offer.get("approval_rate", ""),
                    offer.get("offer_score", ""),
                    offer.get("recommend_score", ""),
                    offer.get("application_review", ""),
                    offer.get("payments", ""),
                    join_list(offer.get("promotion_details")),
                    join_list(offer.get("target_audience_customer_channels")),
                    join_list(offer.get("target_audience_locations")),
                    join_list(offer.get("target_audience_ages")),
                    join_list(offer.get("target_audience_genders")),
                    offer.get("can_apply_offer", ""),
                    offer.get("is_applied_offer", ""),
                    offer.get("client_url", ""),
                    offer.get("offer_id", ""),
                    offer.get("shop_id", ""),
                    offer.get("program_id", ""),
                    offer.get("mkp_listing_id", ""),
                    offer.get("epc", ""),
                ]
            )

            rows += 1
            visits_show = eng.get("VisitsFormatted") or (int(visits) if visits.is_integer() else visits)
            print(f"{status} {brand} | {key} | visits={visits_show}")

    print(f"Done. {rows} dòng -> {out_path} (status GET nếu traffic>{MIN_VISITS})")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[Nhận SIGINT] Thoát.")
        sys.exit(0)
    except Exception as exc:
        print(exc)
        sys.exit(1)
