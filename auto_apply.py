from __future__ import annotations

from pathlib import Path
from typing import Callable
import re
import time
from datetime import datetime
from urllib.parse import urlparse

from openpyxl import load_workbook


APPLY_HEADER_CANDIDATES = ("URL apply", "Link đăng ký", "Link Apply")


def extract_apply_links_from_xlsx(
    path: Path,
    *,
    apply_mode: str = "only_dat",
    row_start: int | None = None,
    row_end: int | None = None,
) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows, None) or ()
        header = [str(v or "").strip() for v in header_row]
        idx = -1
        for i, name in enumerate(header):
            if name in APPLY_HEADER_CANDIDATES:
                idx = i
                break
        if idx < 0:
            return []
        # Status col (để lọc "ĐẠT" khi apply_mode=only_dat)
        status_idx = -1
        for i, name in enumerate(header):
            n = str(name or "").strip()
            if n in ("Trạng thái", "Status"):
                status_idx = i
                break
        out: list[str] = []
        seen: set[str] = set()
        start_n = int(row_start) if row_start is not None else 1
        if start_n < 1:
            start_n = 1
        end_n = int(row_end) if row_end is not None else None
        if end_n is not None and end_n < start_n:
            end_n = start_n

        apply_all = str(apply_mode or "").strip().lower() == "all"
        ok_values = {"đạt", "get", "ok", "pass"}

        data_n = 0  # 1-based index over data rows (excluding header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_n += 1
            if data_n < start_n:
                continue
            if end_n is not None and data_n > end_n:
                break
            if idx >= len(row):
                continue
            if not apply_all and status_idx >= 0 and status_idx < len(row):
                st = str(row[status_idx] or "").strip().lower()
                if st and st not in ok_values:
                    continue
            raw = str(row[idx] or "").strip()
            if not raw:
                continue
            if not (raw.startswith("http://") or raw.startswith("https://")):
                continue
            if raw in seen:
                continue
            seen.add(raw)
            out.append(raw)
        return out
    finally:
        wb.close()


def run_auto_apply(
    links: list[str],
    profile: dict[str, str],
    auto_submit: bool,
    cdp_url: str | None = None,
    login_first: bool = True,
    login_timeout_sec: int = 300,
    should_stop: Callable[[], bool] | None = None,
    log: Callable[[str], None] | None = None,
    brand_timeout_sec: int = 60,
) -> dict:
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        raise RuntimeError(
            "Thiếu Playwright. Cài bằng: pip install playwright && playwright install chromium"
        ) from exc

    def _log(msg: str) -> None:
        if log:
            log(msg)

    def _check_stop(where: str = "") -> None:
        if should_stop and should_stop():
            suffix = f" ({where})" if where else ""
            _log(f"Đã nhận lệnh hủy{suffix}.")
            raise RuntimeError("Đã hủy Auto Apply.")

    class _SkipBrand(Exception):
        pass

    INPUT_FILL_DELAY_MS = 2000

    def _contexts(page):
        out = [page]
        try:
            for fr in page.frames:
                if fr is not page.main_frame:
                    out.append(fr)
        except Exception:
            pass
        return out

    def _is_collabs_page(page) -> bool:
        try:
            u = (page.url or "").lower()
        except Exception:
            return False
        return "collabs.shopify.com" in u

    def _click_apply_now_only(page) -> bool:
        """
        User rule: nếu link đăng ký không có CTA dẫn tới Collabs apply form => bỏ qua brand ngay.
        Chấp nhận các CTA phổ biến: Apply/Apply now/Collab with us/Sign me up/Sign up/Affiliate...
        """
        # Ưu tiên cao nhất: CTA nằm trong vùng `div.collabs-page__cta`.
        for ctx in _contexts(page):
            try:
                clicked_in_cta = bool(
                    ctx.evaluate(
                        """() => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const ctaRoots = Array.from(document.querySelectorAll('div.collabs-page__cta'));
                          if (!ctaRoots.length) return false;
                          const textHints = [
                            "click here to set up shopify collabs",
                            "apply",
                            "join",
                            "sign up",
                            "set up shopify collabs",
                            "shopify collabs",
                          ];
                          let best = null;
                          let bestScore = -1;
                          for (const root of ctaRoots) {
                            const clickable = Array.from(
                              root.querySelectorAll('a,button,[role="button"],input[type="button"],input[type="submit"]')
                            ).filter((el) => isVisible(el) && !el.disabled);
                            for (const el of clickable) {
                              const txt = norm(
                                [
                                  el.innerText || el.textContent || "",
                                  el.getAttribute("aria-label") || "",
                                  el.getAttribute("title") || "",
                                  el.getAttribute("value") || "",
                                ].join(" ")
                              );
                              let score = 1;
                              for (const h of textHints) if (txt.includes(h)) score += 10;
                              if (score > bestScore) {
                                best = el;
                                bestScore = score;
                              }
                            }
                          }
                          if (!best) return false;
                          try { best.click(); return true; } catch (_) {}
                          return false;
                        }"""
                    )
                )
                if clicked_in_cta:
                    page.wait_for_timeout(900)
                    return True
            except Exception:
                pass

        selectors = [
            'button:has-text("Click Here to Set Up Shopify Collabs")',
            'a:has-text("Click Here to Set Up Shopify Collabs")',
            'button:has-text("Apply now")',
            'a:has-text("Apply now")',
            'button:has-text("Apply Now")',
            'a:has-text("Apply Now")',
            'button:has-text("Apply today")',
            'a:has-text("Apply today")',
            'button:has-text("Apply")',
            'a:has-text("Apply")',
            'button:has-text("Apply here")',
            'a:has-text("Apply here")',
            'button:has-text("Apply Here")',
            'a:has-text("Apply Here")',
            'button:has-text("Apply to join")',
            'a:has-text("Apply to join")',
            'button:has-text("Apply to Join")',
            'a:has-text("Apply to Join")',
            'button:has-text("Collab with us")',
            'a:has-text("Collab with us")',
            'button:has-text("Collaborate with us")',
            'a:has-text("Collaborate with us")',
            'button:has-text("Collaborate")',
            'a:has-text("Collaborate")',
            'button:has-text("Work with us")',
            'a:has-text("Work with us")',
            'button:has-text("Work With Us")',
            'a:has-text("Work With Us")',
            'button:has-text("Join the community")',
            'a:has-text("Join the community")',
            'button:has-text("Join Community")',
            'a:has-text("Join Community")',
            'button:has-text("Join")',
            'a:has-text("Join")',
            'button:has-text("Join now")',
            'a:has-text("Join now")',
            'button:has-text("Join our community")',
            'a:has-text("Join our community")',
            'button:has-text("Sign Me Up")',
            'a:has-text("Sign Me Up")',
            'button:has-text("Sign me up")',
            'a:has-text("Sign me up")',
            'button:has-text("Sign up")',
            'a:has-text("Sign up")',
            'button:has-text("Sign Up")',
            'a:has-text("Sign Up")',
            'button:has-text("Affiliate")',
            'a:has-text("Affiliate")',
            'button:has-text("Affiliate program")',
            'a:has-text("Affiliate program")',
            'button:has-text("Affiliate Program")',
            'a:has-text("Affiliate Program")',
            'button:has-text("Become an affiliate")',
            'a:has-text("Become an affiliate")',
            'button:has-text("Ambassador")',
            'a:has-text("Ambassador")',
            'button:has-text("Ambassador program")',
            'a:has-text("Ambassador program")',
            'button:has-text("Creator")',
            'a:has-text("Creator")',
            'button:has-text("Creator program")',
            'a:has-text("Creator program")',
            'button:has-text("Become a creator")',
            'a:has-text("Become a creator")',
            'button:has-text("Become an ambassador")',
            'a:has-text("Become an ambassador")',
            'button:has-text("Partner With Us")',
            'a:has-text("Partner With Us")',
            'button:has-text("Partner with us")',
            'a:has-text("Partner with us")',
            'button:has-text("Become a partner")',
            'a:has-text("Become a partner")',
            'button:has-text("Get Started")',
            'a:has-text("Get Started")',
            'button:has-text("Get started")',
            'a:has-text("Get started")',
            # Một số site dùng tiếng Việt
            'button:has-text("Đăng ký")',
            'a:has-text("Đăng ký")',
            'button:has-text("Ứng tuyển")',
            'a:has-text("Ứng tuyển")',
            'button:has-text("Hợp tác")',
            'a:has-text("Hợp tác")',
        ]
        for ctx in _contexts(page):
            for sel in selectors:
                try:
                    loc = ctx.locator(sel).first
                    if loc.count() and loc.is_visible() and loc.is_enabled():
                        loc.click(timeout=2500)
                        page.wait_for_timeout(900)
                        return True
                except Exception:
                    continue
            # Fallback: click CTA theo text (phòng khi DOM bọc span/div khiến :has-text miss)
            try:
                clicked = bool(
                    ctx.evaluate(
                        """() => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const keywords = [
                            "click here to set up shopify collabs",
                            "set up shopify collabs",
                            "shopify collabs",
                            // Apply
                            "apply now",
                            "apply today",
                            "apply",
                            "application",
                            "apply here",
                            "apply to join",
                            "start application",
                            "submit application",
                            // Collab/partner/community
                            "collab with us",
                            "collaborate with us",
                            "collab",
                            "collabs",
                            "collaborate",
                            "collaboration",
                            "partner with us",
                            "work with us",
                            "join our community",
                            "join the community",
                            "join community",
                            "join now",
                            "join",
                            // Affiliate/Ambassador/Creator
                            "affiliate program",
                            "affiliate",
                            "ambassador program",
                            "ambassador",
                            "creator program",
                            "creator",
                            "become a creator",
                            "become an affiliate",
                            "become an ambassador",
                            "influencer",
                            "become an influencer",
                            // VN
                            "đăng ký",
                            "ứng tuyển",
                            "hợp tác",
                          ];
                          const bad = [
                            "sign in",
                            "log in",
                            "login",
                            "subscribe",
                            "newsletter",
                            "account",
                          ];
                          const candidates = Array.from(document.querySelectorAll('a,button,[role="button"],input[type="button"],input[type="submit"]'))
                            .filter((el) => isVisible(el) && !el.disabled);

                          const attrText = (el) => {
                            const parts = [];
                            try { parts.push(el.innerText || el.textContent || ""); } catch (_) {}
                            try { parts.push(el.getAttribute("aria-label") || ""); } catch (_) {}
                            try { parts.push(el.getAttribute("title") || ""); } catch (_) {}
                            try { parts.push(el.getAttribute("value") || ""); } catch (_) {}
                            return norm(parts.filter(Boolean).join(" "));
                          };

                          const hrefText = (el) => {
                            try {
                              const h = (el.getAttribute && el.getAttribute("href")) ? String(el.getAttribute("href") || "") : "";
                              return norm(h);
                            } catch (_) {
                              return "";
                            }
                          };

                          const hasBad = (s) => bad.some((b) => s.includes(b));
                          const hrefSignals = [
                            "pages/collab",
                            "pages/collabs",
                            "pages/collabor",
                            "pages/collaborators",
                            "pages/creator",
                            "pages/creators",
                            "pages/influencer",
                            "pages/influencers",
                            "pages/ambassador",
                            "pages/ambassadors",
                            "pages/affiliate-program",
                            "pages/affiliates",
                            "pages/partner",
                            "pages/partners",
                            "collabs",
                            "collab",
                            "affiliate",
                            "ambassador",
                            "creator",
                            "influencer",
                            "partner",
                            "partnership",
                            "community",
                          ];

                          const scoreEl = (el) => {
                            const t = attrText(el);
                            const h = hrefText(el);
                            if (hasBad(t)) return -1;
                            let s = 0;
                            for (const k of keywords) if (t.includes(k)) s += 10;
                            for (const sig of hrefSignals) if (h.includes(sig)) s += 6;
                            // Bonus: nếu là <a> có href rõ ràng
                            try {
                              if (String(el.tagName || "").toLowerCase() === "a" && h) s += 2;
                            } catch (_) {}
                            // Bonus: element trong vùng CTA thường là button
                            try {
                              const tag = String(el.tagName || "").toLowerCase();
                              if (tag === "button") s += 1;
                            } catch (_) {}
                            return s;
                          };

                          let best = null;
                          let bestScore = 0;
                          for (const el of candidates) {
                            const sc = scoreEl(el);
                            if (sc > bestScore) {
                              bestScore = sc;
                              best = el;
                            }
                          }
                          if (best && bestScore >= 10) {
                            try { best.click(); return true; } catch (_) {}
                          }
                          return false;
                        }"""
                    )
                )
                if clicked:
                    page.wait_for_timeout(900)
                    return True
            except Exception:
                pass
        return False

    def _resolve_collabs_page_after_apply(context, current_page):
        """
        Sau khi bấm Apply now, collabs có thể mở ở:
        - cùng tab hiện tại
        - tab mới
        Hàm này trả về tab collabs để tiếp tục fill; nếu không có thì trả current_page.
        """
        try:
            if _is_collabs_page(current_page):
                return current_page
        except Exception:
            pass
        # Ưu tiên tab mới mở gần nhất
        try:
            pages = list(context.pages)
            for p in reversed(pages):
                try:
                    if _is_collabs_page(p):
                        return p
                except Exception:
                    continue
        except Exception:
            pass
        return current_page

    def _fill_first_visible(page, selectors: list[str], value: str) -> bool:
        if not value:
            return False
        for ctx in _contexts(page):
            for sel in selectors:
                try:
                    loc = ctx.locator(sel).first
                    if loc.count() and loc.is_visible():
                        loc.fill(value, timeout=1800)
                        try:
                            page.wait_for_timeout(INPUT_FILL_DELAY_MS)
                        except Exception:
                            pass
                        return True
                except Exception:
                    continue
        return False

    def _norm(text: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()

    def _tokens(text: str) -> set[str]:
        return {t for t in _norm(text).split(" ") if t}

    def _sim_score(label: str, pattern: str) -> float:
        lt = _tokens(label)
        pt = _tokens(pattern)
        if not lt or not pt:
            return 0.0
        inter = len(lt & pt)
        if inter == 0:
            return 0.0
        # Ưu tiên pattern được phủ token tốt hơn (recall theo pattern)
        recall = inter / max(1, len(pt))
        jacc = inter / max(1, len(lt | pt))
        phrase_bonus = 0.15 if _norm(pattern) in _norm(label) else 0.0
        # Bổ sung mức giống nhau theo chuỗi ký tự để bắt biến thể label.
        ls = _norm(label).replace(" ", "")
        ps = _norm(pattern).replace(" ", "")
        char_sim = 0.0
        if ls and ps:
            common = sum(1 for ch in set(ps) if ch in set(ls))
            char_sim = common / max(1, len(set(ps)))
        return (0.65 * recall) + (0.2 * jacc) + (0.15 * char_sim) + phrase_bonus

    def _pick_profile_value(label: str, fallback: str = "") -> str:
        def _clean_value(v: str, long_text: bool = False) -> str:
            s = str(v or "").strip()
            # Tránh dùng chuỗi số vô nghĩa đã lưu cache trước đó cho câu trả lời tự nhiên.
            if long_text and re.fullmatch(r"[0-9\\s.,-]+", s or ""):
                return ""
            return s

        # Một số trường hợp cần điền câu trả lời "chỉ định" theo label.
        raw_label = str(label or "").strip()
        low_label = raw_label.lower()
        if low_label:
            # Phone: luôn ưu tiên mẫu số điện thoại nếu label có chữ "phone"
            if "phone" in low_label:
                ph = str(profile.get("phone") or "").strip()
                if ph:
                    return ph
            # 1) Social / affiliate link: ưu tiên Instagram rồi TikTok
            if ("social" in low_label) or ("share your affiliate link" in low_label):
                insta = str(profile.get("instagram") or "").strip()
                tt = str(profile.get("tiktok") or "").strip()
                if insta:
                    return insta
                if tt:
                    return tt
            # 2) Website
            if "website" in low_label:
                site = str(profile.get("website") or "").strip()
                if site:
                    return site
            # 2.5) Product / products -> mẫu trải nghiệm sản phẩm
            if ("product" in low_label) or ("products" in low_label):
                purchase_love = str(profile.get("purchase_love") or "").strip()
                if purchase_love:
                    return purchase_love
            # 3) Why -> dùng mẫu "lý do muốn tham gia"
            if "why" in low_label:
                why_join = str(profile.get("why_join") or "").strip()
                if why_join:
                    return why_join
            # 4) Followers / how many / how much -> mẫu chỉ định mới
            if ("followers" in low_label) or ("how many" in low_label) or ("how much" in low_label):
                fe = str(profile.get("followers_engagement") or "").strip()
                return fe or "Bạn có bao nhiêu người theo dõi và bạn nhận được bao nhiêu tương tác trên các bài đăng riêng lẻ?"

        mapping: list[tuple[list[str], str]] = [
            (["date of birth", "dob", "birthdate"], profile.get("dob", "")),
            (["shipping location", "shipping country", "country"], profile.get("shipping_location", "")),
            (["business type"], profile.get("business_type", "")),
            (["in 1 or 2 sentences tell us why you d love to be in our community"], profile.get("why_join", "")),
            (["why would you like to collaborate with me"], profile.get("why_join", "")),
            (["what is the key to a successful collaboration"], profile.get("successful_partnership", "")),
            (["which brands have you worked with", "worked with brands"], profile.get("brands_worked", "")),
            (["most successful partnership", "successful partnership"], profile.get("successful_partnership", "")),
            (["what piece of your content inspires you", "content inspires"], profile.get("content_inspires", "")),
            (["what do you hope to gain from this partnership", "hope to gain"], profile.get("hope_gain", "")),
            (["how did you find us", "how did you first hear"], profile.get("how_found", "")),
            (["where do you live city country", "city country"], profile.get("city_country", "")),
            (["who is your demographic", "demographic"], profile.get("demographic", "")),
            (["what strategies do you use to grow your account", "grow your account"], profile.get("growth_strategy", "")),
            (["what content ideas do you have for our product", "content ideas"], profile.get("content_ideas", "")),
            (["why you d be a great fit", "great fit"], profile.get("why_fit", "")),
            (["have you made a purchase from us before", "purchase from us before"], "Yes"),
            (["what did you love about the product"], profile.get("purchase_love", "")),
            (["why you d love to be in our community", "why would you love to be in our community"], profile.get("why_join", "")),
            (["amazon creator", "amazon page"], profile.get("website", "")),
            (["instagram handle", "instagram"], profile.get("instagram", "")),
            (["tik tok handle", "tiktok"], profile.get("tiktok", "")),
            (["youtube", "youtube handle"], profile.get("youtube", "")),
            (["email"], profile.get("email", "")),
            (["phone"], profile.get("phone", "")),
            (["website"], profile.get("website", "")),
            (["full name", "name"], profile.get("full_name", "")),
        ]
        generic_short = str(profile.get("generic_short") or "").strip()
        generic_long = str(profile.get("generic_long") or "").strip()
        best_val = ""
        best_score = 0.0
        for patterns, val in mapping:
            is_long_text = any(
                k in " ".join(patterns)
                for k in (
                    "successful partnership",
                    "content inspires",
                    "hope to gain",
                    "how did you find",
                    "great fit",
                    "why",
                    "love about the product",
                )
            )
            v = _clean_value(str(val or ""), long_text=is_long_text)
            if not v:
                continue
            s = 0.0
            for ptn in patterns:
                s = max(s, _sim_score(label, ptn))
            if s > best_score:
                best_score = s
                best_val = v
        if best_score >= 0.45 and best_val:
            return best_val
        if fallback.strip():
            return fallback.strip()
        # Ưu tiên mẫu chung theo loại câu hỏi để hạn chế bỏ sót input.
        nlabel = _norm(label)
        is_long_question = any(
            k in nlabel
            for k in (
                "why",
                "how",
                "experience",
                "strategy",
                "partner",
                "community",
                "collaborate",
                "tell us",
                "describe",
                "explain",
            )
        )
        if is_long_question:
            if generic_long:
                return generic_long
            # Nếu không có mẫu dài, trả N/A để giảm lỗi validate với câu hỏi dài/khó.
            if generic_short:
                return generic_short
            return "N/A"
        if generic_short:
            return generic_short
        # Không ép câu trả lời mặc định cứng; nếu thiếu mẫu thì để trống.
        return str(profile.get("message") or "").strip()

    def _fill_date_of_birth(page) -> bool:
        dob_raw = str(profile.get("dob") or "").strip()
        if not dob_raw:
            return False

        def _to_iso_date(value: str) -> str:
            s = str(value or "").strip()
            if not s:
                return ""
            s = s.replace(".", "/").replace("-", "/").strip()
            # Try common formats: MM/DD/YYYY, DD/MM/YYYY, YYYY/MM/DD
            for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
            # Already ISO?
            try:
                return datetime.strptime(str(value).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
            except ValueError:
                return ""

        iso = _to_iso_date(dob_raw)

        # Ưu tiên field date kiểu Polaris theo HTML thực tế (ID có thể đổi: PolarisTextField1/2/...).
        if iso:
            for ctx in _contexts(page):
                try:
                    ok = bool(
                        ctx.evaluate(
                            """(iso) => {
                              const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                              const isVisible = (el) => {
                                if (!el) return false;
                                const r = el.getBoundingClientRect();
                                const st = getComputedStyle(el);
                                return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                              };
                              const fire = (el) => {
                                try { el.dispatchEvent(new Event('input', { bubbles: true })); } catch (_) {}
                                try { el.dispatchEvent(new Event('change', { bubbles: true })); } catch (_) {}
                                try { el.dispatchEvent(new Event('blur', { bubbles: true })); } catch (_) {}
                              };
                              const lockDobValue = (inp, lockedIso) => {
                                if (!inp || !lockedIso) return;
                                try {
                                  inp.dataset.lockDobIso = String(lockedIso);
                                } catch (_) {}
                                const enforce = () => {
                                  try {
                                    if ((inp.value || '') !== lockedIso) {
                                      inp.value = lockedIso;
                                      fire(inp);
                                    }
                                  } catch (_) {}
                                };
                                try {
                                  inp.addEventListener('input', enforce, true);
                                  inp.addEventListener('change', enforce, true);
                                  inp.addEventListener('blur', enforce, true);
                                  inp.addEventListener('focus', enforce, true);
                                } catch (_) {}
                                try {
                                  const prev = Number(inp.dataset.lockDobTimerId || 0);
                                  if (prev) clearInterval(prev);
                                } catch (_) {}
                                try {
                                  const timerId = setInterval(enforce, 180);
                                  inp.dataset.lockDobTimerId = String(timerId);
                                  setTimeout(() => {
                                    try { clearInterval(timerId); } catch (_) {}
                                  }, 7000);
                                } catch (_) {}
                                enforce();
                              };
                              const candidates = Array.from(document.querySelectorAll('input[type="date"]'))
                                .filter((i) => isVisible(i) && !i.disabled);
                              let inp = candidates.find((i) => {
                                const id = (i.id || '').toLowerCase();
                                const aria = norm(i.getAttribute('aria-labelledby') || '');
                                return id.startsWith('polaristextfield') || aria.includes('polaristextfield');
                              }) || null;
                              if (!inp) {
                                inp = candidates.find((i) => {
                                  const aria = norm(i.getAttribute('aria-labelledby') || '');
                                  if (aria) {
                                    const first = aria.split(/\\s+/).filter(Boolean)[0];
                                    const lb = first ? document.getElementById(first) : null;
                                    const txt = norm((lb && lb.innerText) || '');
                                    return (
                                      txt.includes("what’s your date of birth") ||
                                      txt.includes("what's your date of birth") ||
                                      txt.includes("date of birth") ||
                                      txt.includes("birthdate") ||
                                      txt.includes("birth date") ||
                                      txt.includes("dob") ||
                                      txt.includes("ngày sinh") ||
                                      txt.includes("ngay sinh")
                                    );
                                  }
                                  return false;
                                }) || null;
                              }
                              // Nếu trang chỉ có 1 input[type=date] thì chọn trực tiếp để giảm miss.
                              if (!inp && candidates.length === 1) inp = candidates[0];
                              if (!inp) return false;
                              try {
                                inp.setAttribute('autocomplete', 'off');
                                inp.value = '';
                                fire(inp);
                                try { inp.valueAsDate = new Date(iso + 'T00:00:00'); } catch (_) {}
                                inp.value = iso;
                                fire(inp);
                                lockDobValue(inp, iso);
                                return (inp.value || '') === iso;
                              } catch (_) {
                                return false;
                              }
                            }""",
                            iso,
                        )
                    )
                    if ok:
                        try:
                            page.wait_for_timeout(INPUT_FILL_DELAY_MS)
                        except Exception:
                            pass
                        return True
                except Exception:
                    continue

        # Collabs DOB thường là input type="date" với label "What’s your date of birth?".
        for ctx in _contexts(page):
            try:
                ok = bool(
                    ctx.evaluate(
                        """(raw, iso) => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const labelTextFor = (inp) => {
                            const aria = inp.getAttribute('aria-labelledby') || '';
                            if (aria) {
                              const first = aria.split(/\\s+/).filter(Boolean)[0];
                              const lb = first ? document.getElementById(first) : null;
                              if (lb && lb.innerText) return norm(lb.innerText);
                            }
                            const id = inp.id || '';
                            if (id) {
                              const lb = document.querySelector(`label[for="${id}"]`);
                              if (lb && lb.innerText) return norm(lb.innerText);
                            }
                            const near = inp.closest('label, fieldset, .Polaris-FormLayout__Item, div');
                            return near && near.innerText ? norm(near.innerText) : '';
                          };
                          const blocks = Array.from(document.querySelectorAll('fieldset, .Polaris-FormLayout__Item, div, form'));
                          const fire = (el) => {
                            try { el.dispatchEvent(new Event('input', { bubbles: true })); } catch (_) {}
                            try { el.dispatchEvent(new Event('change', { bubbles: true })); } catch (_) {}
                          };
                          const lockDobValue = (inp, lockedIso) => {
                            if (!inp || !lockedIso) return;
                            try {
                              inp.dataset.lockDobIso = String(lockedIso);
                            } catch (_) {}
                            const enforce = () => {
                              try {
                                if ((inp.value || '') !== lockedIso) {
                                  inp.value = lockedIso;
                                  fire(inp);
                                }
                              } catch (_) {}
                            };
                            try {
                              inp.addEventListener('input', enforce, true);
                              inp.addEventListener('change', enforce, true);
                              inp.addEventListener('blur', enforce, true);
                              inp.addEventListener('focus', enforce, true);
                            } catch (_) {}
                            try {
                              const prev = Number(inp.dataset.lockDobTimerId || 0);
                              if (prev) clearInterval(prev);
                            } catch (_) {}
                            try {
                              const timerId = setInterval(enforce, 180);
                              inp.dataset.lockDobTimerId = String(timerId);
                              setTimeout(() => {
                                try { clearInterval(timerId); } catch (_) {}
                              }, 7000);
                            } catch (_) {}
                            enforce();
                          };
                          const setDateValue = (inp, isoValue, rawValue) => {
                            const typ = norm(inp.getAttribute('type') || '');
                            const val = (typ === 'date' && isoValue) ? isoValue : rawValue;
                            try {
                              inp.setAttribute('autocomplete', 'off');
                                  inp.setAttribute('autocorrect', 'off');
                                  inp.setAttribute('data-form-type', 'other');
                                  inp.setAttribute('data-lpignore', 'true');
                              inp.value = '';
                              fire(inp);
                              if (typ === 'date' && isoValue) {
                                try { inp.valueAsDate = new Date(isoValue + 'T00:00:00'); } catch (_) {}
                                inp.value = isoValue;
                                lockDobValue(inp, isoValue);
                              } else {
                                inp.value = String(val || '');
                              }
                              fire(inp);
                              try { inp.focus(); } catch (_) {}
                              return true;
                            } catch (_) {
                              return false;
                            }
                          };

                          // 1) Prefer matching input[type=date] by label text
                          const dateInputs = Array.from(document.querySelectorAll('input[type="date"]'))
                            .filter((i) => isVisible(i) && !i.disabled);
                          for (const inp of dateInputs) {
                            const lt = labelTextFor(inp);
                            if (
                              lt.includes("what’s your date of birth") ||
                              lt.includes("what's your date of birth") ||
                              lt.includes("date of birth") ||
                              lt.includes("birthdate") ||
                              lt.includes("birth date") ||
                              lt.includes("dob") ||
                              lt.includes("ngày sinh") ||
                              lt.includes("ngay sinh")
                            ) {
                              return setDateValue(inp, iso, raw);
                            }
                          }
                          if (dateInputs.length === 1) {
                            return setDateValue(dateInputs[0], iso, raw);
                          }
                          for (const b of blocks) {
                            const text = norm(b.innerText || '');
                            if (
                              !text.includes("what’s your date of birth") &&
                              !text.includes("what's your date of birth") &&
                              !text.includes("date of birth") &&
                              !text.includes("birthdate") &&
                              !text.includes("birth date") &&
                              !text.includes("dob") &&
                              !text.includes("ngày sinh") &&
                              !text.includes("ngay sinh")
                            ) continue;
                            const inp = Array.from(b.querySelectorAll('input')).find((i) => isVisible(i) && !i.disabled);
                            if (!inp) return false;
                            return setDateValue(inp, iso, raw);
                          }
                          return false;
                        }""",
                        dob_raw,
                        iso,
                    )
                )
                if ok:
                    return True
            except Exception:
                continue

        # Fallback: thử fill theo selectors (cho trường hợp input text thường)
        selectors = [
            'input[type="date"]',
            'input[aria-label*="date of birth" i]',
            'input[placeholder*="date of birth" i]',
            'input[placeholder*="MM" i][placeholder*="DD" i][placeholder*="YYYY" i]',
            'input[aria-label*="ngày sinh" i]',
            'input[placeholder*="ngày sinh" i]',
            'input[name*="birth" i]',
            'input[id*="birth" i]',
            'input[name*="dob" i]',
            'input[id*="dob" i]',
        ]
        return _fill_first_visible(page, selectors, iso or dob_raw)

    def _select_shipping_location(page) -> bool:
        target = str(profile.get("shipping_location") or "").strip() or "United States"
        for ctx in _contexts(page):
            try:
                ok = bool(
                    ctx.evaluate(
                        """(target) => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const selects = Array.from(document.querySelectorAll('select')).filter((s) => isVisible(s) && !s.disabled);
                          const wanted = norm(target);
                          const aliases = {
                            "uae": "united arab emirates",
                            "uk": "united kingdom",
                            "usa": "united states",
                            "gemany": "germany",
                          };
                          const wantedNorm = aliases[wanted] || wanted;
                          const countryToCode = {
                            "united states": "us",
                            "united kingdom": "gb",
                            "germany": "de",
                            "france": "fr",
                            "canada": "ca",
                            "australia": "au",
                            "netherlands": "nl",
                            "sweden": "se",
                            "italy": "it",
                            "spain": "es",
                            "united arab emirates": "ae",
                            "saudi arabia": "sa",
                            "singapore": "sg",
                            "india": "in",
                            "japan": "jp",
                            "vietnam": "vn",
                          };
                          const wantedCode = (wantedNorm.length === 2 ? wantedNorm : (countryToCode[wantedNorm] || "")).toLowerCase();
                          const isShippingSelect = (sel) => {
                            const id = sel.id || '';
                            let labelText = '';
                            if (id) {
                              const lb = document.querySelector(`label[for="${id}"]`);
                              if (lb && lb.innerText) labelText = norm(lb.innerText);
                            }
                            const wrapText = norm((sel.closest('label, fieldset, .Polaris-FormLayout__Item, div') || {}).innerText || '');
                            const combined = `${labelText} ${wrapText}`.trim();
                            return combined.includes('shipping') && combined.includes('location');
                          };
                          const pickOption = (opts) => {
                            // 1) exact by visible country name
                            let hit = opts.find((o) => norm(o.textContent || '') === wantedNorm);
                            if (hit) return hit;
                            // 2) exact by option code (US/DE/VN...)
                            if (wantedCode) {
                              hit = opts.find((o) => norm(o.value || '') === wantedCode);
                              if (hit) return hit;
                            }
                            // 3) include by text/value
                            hit = opts.find((o) => {
                              const t = norm(o.textContent || '');
                              const v = norm(o.value || '');
                              return t.includes(wantedNorm) || wantedNorm.includes(t) || (wantedCode && v === wantedCode);
                            });
                            return hit || null;
                          };
                          for (const sel of selects) {
                            if (!isShippingSelect(sel)) continue;
                            const opts = Array.from(sel.options || []);
                            let hit = pickOption(opts);
                            if (hit) {
                              sel.value = hit.value;
                              sel.dispatchEvent(new Event('input', { bubbles: true }));
                              sel.dispatchEvent(new Event('change', { bubbles: true }));
                              // Chỉ coi là thành công khi selected match đúng text hoặc code.
                              const curText = norm((sel.selectedOptions && sel.selectedOptions[0] ? sel.selectedOptions[0].textContent : '') || '');
                              const curVal = norm(sel.value || '');
                              if (curText === wantedNorm || curText.includes(wantedNorm) || (wantedCode && curVal === wantedCode)) {
                                return true;
                              }
                            }
                          }
                          return false;
                        }""",
                        target,
                    )
                )
                if ok:
                    return True
            except Exception:
                continue
        return False

    def _set_identify_choice(page) -> bool:
        """
        Collabs hiện tại render "How do you identify?" dưới dạng radio.
        Rule:
        - Nếu user chọn 1 giá trị -> click đúng radio đó.
        - Nếu user chọn nhiều (từ UI checkbox) -> ưu tiên cái đầu.
        - Nếu trống -> Prefer not to say.
        """
        desired = profile.get("identify") or []
        if not isinstance(desired, list):
            desired = []
        desired = [str(x or "").strip() for x in desired if str(x or "").strip()]
        target = desired[0] if desired else "Prefer not to say"
        for ctx in _contexts(page):
            try:
                ok = bool(
                    ctx.evaluate(
                        """(target) => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const want = norm(target);
                          const blocks = Array.from(document.querySelectorAll('.Polaris-FormLayout__Item, fieldset, .Polaris-ChoiceList, form, div'));
                          const labelOf = (inp) => {
                            const id = inp.id || '';
                            if (id) {
                              const lb = document.querySelector(`label[for="${id}"]`);
                              if (lb && lb.innerText) return norm(lb.innerText);
                            }
                            const near = inp.closest('label');
                            if (near && near.innerText) return norm(near.innerText);
                            return '';
                          };
                          const isIdentifyBlock = (b) => {
                            const t = norm(b.innerText || '');
                            return t.includes('how do you identify');
                          };
                          for (const b of blocks) {
                            if (!isIdentifyBlock(b)) continue;
                            const radios = Array.from(b.querySelectorAll('input[type="radio"]')).filter((i) => isVisible(i) && !i.disabled);
                            if (!radios.length) continue;
                            let chosen = radios.find((r) => {
                              const txt = labelOf(r);
                              return txt.includes(want);
                            });
                            if (!chosen) {
                              chosen = radios.find((r) => labelOf(r).includes('prefer not to say')) || radios[0];
                            }
                            if (!chosen) continue;
                            try {
                              const id = chosen.id || '';
                              if (id) {
                                const lb = document.querySelector(`label[for="${id}"]`);
                                if (lb) lb.click();
                                else chosen.click();
                              } else {
                                chosen.click();
                              }
                            } catch (_) {
                              try { chosen.click(); } catch (_) {}
                            }
                            return true;
                          }
                          return false;
                        }""",
                        target,
                    )
                )
                if ok:
                    return True
            except Exception:
                continue
        return False

    def _label_for_field(page, el) -> str:
        try:
            return str(
                el.evaluate(
                    """(node) => {
                        const clean = (s) => (s || '').trim();
                        const aria = clean(node.getAttribute('aria-label'));
                        if (aria) return aria;
                        const ph = clean(node.getAttribute('placeholder'));
                        if (ph) return ph;
                        const id = clean(node.id || '');
                        if (id) {
                          const lb = document.querySelector(`label[for="${id}"]`);
                          if (lb && clean(lb.innerText)) return clean(lb.innerText);
                        }
                        const near = node.closest('label, fieldset, .Polaris-FormLayout__Item, div');
                        if (near && clean(near.innerText)) return clean(near.innerText).split('\\n')[0];
                        const nm = clean(node.getAttribute('name'));
                        if (nm) return nm;
                        return '';
                    }"""
                )
                or ""
            ).strip()
        except Exception:
            return ""

    def _fill_visible_text_fields(page) -> int:
        filled = 0
        selectors = [
            "input:not([type=hidden]):not([type=radio]):not([type=checkbox]):not([type=submit]):not([type=button])",
            "textarea",
        ]
        for ctx in _contexts(page):
            for sel in selectors:
                fields = ctx.locator(sel)
                cnt = fields.count()
                for i in range(cnt):
                    el = fields.nth(i)
                    try:
                        if not el.is_visible() or not el.is_enabled():
                            continue
                        label = _label_for_field(ctx, el)
                        typ = ""
                        try:
                            typ = (el.get_attribute("type") or "").lower().strip()
                        except Exception:
                            typ = ""
                        # Date inputs (DOB etc.) have strict ISO value format; handled separately.
                        if typ == "date":
                            continue
                        if typ == "email":
                            value = profile.get("email", "").strip() or _pick_profile_value(label, profile.get("email", ""))
                        elif typ == "tel":
                            value = profile.get("phone", "").strip() or _pick_profile_value(label, profile.get("phone", ""))
                        else:
                            value = _pick_profile_value(label, "")
                        if not str(value).strip():
                            continue
                        # Không dùng dữ liệu cache/autofill từ browser: luôn clear rồi điền lại.
                        try:
                            el.evaluate(
                                """(node) => {
                                    node.setAttribute('autocomplete', 'off');
                                    if ('value' in node) node.value = '';
                                }"""
                            )
                        except Exception:
                            pass
                        try:
                            el.fill("", timeout=900)
                        except Exception:
                            pass
                        try:
                            el.fill(str(value), timeout=1800)
                        except Exception:
                            # fallback: force set value + fire events
                            el.evaluate(
                                """(node, val) => {
                                  node.value = val;
                                  node.dispatchEvent(new Event('input', { bubbles: true }));
                                  node.dispatchEvent(new Event('change', { bubbles: true }));
                                }""",
                                str(value),
                            )
                        filled += 1
                        try:
                            page.wait_for_timeout(INPUT_FILL_DELAY_MS)
                        except Exception:
                            pass
                    except Exception:
                        continue
        return filled

    def _select_purchase_before_choice(page) -> bool:
        """
        Mẫu riêng cho câu:
        "Have you made a purchase from us before?"
        Hỗ trợ cả radio và checkbox, chọn theo profile.purchase_before_choice.
        """
        preferred = str(profile.get("purchase_before_choice") or "Yes").strip() or "Yes"
        for ctx in _contexts(page):
            try:
                ok = bool(
                    ctx.evaluate(
                        """(preferredRaw) => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const preferred = norm(preferredRaw || 'yes');
                          const blocks = Array.from(document.querySelectorAll('fieldset, .Polaris-ChoiceList, .Polaris-FormLayout__Item, form, div'));
                          const labelOf = (input) => {
                            const id = input.id || '';
                            if (id) {
                              const lb = document.querySelector(`label[for="${id}"]`);
                              if (lb && lb.innerText) return norm(lb.innerText);
                            }
                            const near = input.closest('label');
                            if (near && near.innerText) return norm(near.innerText);
                            return '';
                          };
                          const isQuestionBlock = (text) => {
                            const t = norm(text || '');
                            return t.includes('have you made a purchase from us before') || t.includes('made a purchase from us before');
                          };
                          const pickPreferred = (arr) => {
                            return arr.find((x) => {
                              const t = labelOf(x);
                              if (!t) return false;
                              if (preferred.includes('not yet')) return t.includes('not yet') || t.includes('plan to');
                              if (preferred === 'no') return /\\bno\\b/.test(t);
                              return /\\byes\\b/.test(t);
                            }) || null;
                          };

                          for (const b of blocks) {
                            if (!isQuestionBlock(b.innerText || '')) continue;
                            const radios = Array.from(b.querySelectorAll('input[type="radio"]')).filter((i) => isVisible(i) && !i.disabled);
                            if (radios.length) {
                              const target = pickPreferred(radios) || radios[0];
                              if (target && !target.checked) target.click();
                              return true;
                            }
                            const boxes = Array.from(b.querySelectorAll('input[type="checkbox"]')).filter((i) => isVisible(i) && !i.disabled);
                            if (boxes.length) {
                              for (const bx of boxes) if (bx.checked) bx.click();
                              const target = pickPreferred(boxes) || boxes[0];
                              if (target && !target.checked) target.click();
                              return true;
                            }
                          }
                          return false;
                        }""",
                        preferred,
                    )
                )
                if ok:
                    return True
            except Exception:
                continue
        return False

    def _stable_page_before_actions(page) -> None:
        # Hạn chế hiệu ứng cuộn mượt gây cảm giác kéo lên/xuống liên tục.
        for ctx in _contexts(page):
            try:
                ctx.evaluate(
                    """() => {
                      try { document.documentElement.style.scrollBehavior = 'auto'; } catch (_) {}
                      try { document.body.style.scrollBehavior = 'auto'; } catch (_) {}
                      try { document.documentElement.style.overscrollBehavior = 'none'; } catch (_) {}
                      try { document.body.style.overscrollBehavior = 'none'; } catch (_) {}
                      try {
                        if (!window.__autoApplyScrollLocked) {
                          window.__autoApplyScrollLocked = true;
                          const y = window.scrollY;
                          window.addEventListener('scroll', () => window.scrollTo(0, y), { passive: true });
                        }
                      } catch (_) {}
                    }"""
                )
            except Exception:
                pass

    def _pick_choices_yes_or_first(page) -> int:
        """
        Rule:
        - Ưu tiên chọn option có text "Yes" nếu tồn tại.
        - Nếu không có Yes thì chọn option đầu.
        - Áp cho cả radio/checkbox theo từng block câu hỏi.
        """
        total = 0
        for ctx in _contexts(page):
            try:
                total += int(
                    ctx.evaluate(
                        """() => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const textOf = (input) => {
                            const id = input.id || '';
                            if (id) {
                              const lb = document.querySelector(`label[for="${id}"]`);
                              if (lb) return norm(lb.innerText);
                            }
                            const near = input.closest('label');
                            if (near) return norm(near.innerText);
                            const wrap = input.closest('fieldset, .Polaris-ChoiceList, .Polaris-FormLayout__Item, div');
                            return wrap ? norm(wrap.innerText) : '';
                          };
                          const isPurchaseQuestionInput = (input) => {
                            const wrap = input.closest('fieldset, .Polaris-ChoiceList, .Polaris-FormLayout__Item, form, div');
                            const t = norm((wrap && wrap.innerText) || '');
                            return t.includes('have you made a purchase from us before') || t.includes('made a purchase from us before');
                          };
                          let picked = 0;
                          // 1) Radios by group name (fix trường hợp nhiều câu trong cùng block)
                          const radios = Array.from(document.querySelectorAll('input[type="radio"]'))
                            .filter((i) => isVisible(i) && !i.disabled && !isPurchaseQuestionInput(i));
                          const radioGroups = new Map();
                          for (const r of radios) {
                            const k = r.name || `__single_${radios.indexOf(r)}`;
                            if (!radioGroups.has(k)) radioGroups.set(k, []);
                            radioGroups.get(k).push(r);
                          }
                          for (const arr of radioGroups.values()) {
                            if (!arr.length) continue;
                            if (arr.some((r) => r.checked)) continue;
                            let target = arr.find((r) => /\\byes\\b/.test(textOf(r))) || arr[0];
                            if (target) {
                              target.click();
                              picked += 1;
                            }
                          }
                          // 2) Checkboxes: mỗi cụm câu hỏi chỉ giữ checkbox đầu tiên.
                          const allBoxes = Array.from(document.querySelectorAll('input[type="checkbox"]'))
                            .filter((i) => isVisible(i) && !i.disabled && !isPurchaseQuestionInput(i));
                          const checkboxGroups = new Map();
                          const groupRootOf = (el) => {
                            return el.closest('fieldset, .Polaris-LegacyStack, .Polaris-ChoiceList, [role="group"], .Polaris-FormLayout__Item, form')
                              || el.closest('div')
                              || document.body;
                          };
                          for (const cb of allBoxes) {
                            const root = groupRootOf(cb);
                            const key = root;
                            if (!checkboxGroups.has(key)) checkboxGroups.set(key, []);
                            checkboxGroups.get(key).push(cb);
                          }
                          for (const arr of checkboxGroups.values()) {
                            if (!arr.length) continue;
                            const target = arr[0];
                            let changed = false;
                            const hasChecked = arr.some((x) => x.checked);
                            if (hasChecked) {
                              // Có ô đã tick: clear toàn bộ rồi tick lại ô đầu.
                              for (let i = 0; i < arr.length; i += 1) {
                                if (arr[i].checked) {
                                  arr[i].click();
                                  changed = true;
                                }
                              }
                            }
                            // Không có ô nào tick hoặc vừa clear xong: chỉ tick ô đầu.
                            if (target && !target.checked) {
                              target.click();
                              changed = true;
                            }
                            if (changed) picked += 1;
                          }
                          return picked;
                        }""",
                    )
                    or 0
                )
            except Exception:
                continue
        return total

    def _select_first_for_unknown_selects(page) -> int:
        """
        Với select lạ chưa có rule: nếu chưa chọn giá trị thì chọn option đầu tiên hợp lệ.
        Giảm tối đa lỗi validate required.
        """
        total = 0
        for ctx in _contexts(page):
            try:
                total += int(
                    ctx.evaluate(
                        """() => {
                          const norm = (s) => (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                          const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const st = getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
                          };
                          const skipIfMatches = (txt) => {
                            const t = norm(txt);
                            return t.includes('shipping location') || t.includes('date of birth') || t.includes('identify');
                          };
                          let changed = 0;
                          const selects = Array.from(document.querySelectorAll('select'))
                            .filter((s) => isVisible(s) && !s.disabled);
                          for (const sel of selects) {
                            const wrapText = norm((sel.closest('label, fieldset, .Polaris-FormLayout__Item, div') || {}).innerText || '');
                            if (skipIfMatches(wrapText)) continue;
                            const current = norm(sel.value || '');
                            if (current) continue;
                            const opts = Array.from(sel.options || []).filter((o) => !o.disabled);
                            if (!opts.length) continue;
                            let target = null;
                            if (!target) target = opts.find((o) => norm(o.value || '') !== "" && !norm(o.textContent || '').includes('select'));
                            if (!target) target = opts[0];
                            sel.value = target.value;
                            sel.dispatchEvent(new Event('input', { bubbles: true }));
                            sel.dispatchEvent(new Event('change', { bubbles: true }));
                            changed += 1;
                          }
                          return changed;
                        }"""
                    )
                    or 0
                )
            except Exception:
                continue
        return total

    def _fill_profile(page) -> int:
        filled = 0
        full_name = profile.get("full_name", "")
        first_name = profile.get("first_name", "") or (full_name.split(" ")[0] if full_name else "")
        last_name = profile.get("last_name", "") or (" ".join(full_name.split(" ")[1:]) if full_name else "")
        if _fill_first_visible(
            page,
            [
                'input[name*="full" i]',
                'input[id*="full" i]',
                'input[placeholder*="full name" i]',
                'input[aria-label*="full name" i]',
            ],
            full_name,
        ):
            filled += 1
        if _fill_first_visible(
            page,
            ['input[name*="first" i]', 'input[id*="first" i]', 'input[placeholder*="first" i]'],
            first_name,
        ):
            filled += 1
        if _fill_first_visible(
            page,
            ['input[name*="last" i]', 'input[id*="last" i]', 'input[placeholder*="last" i]'],
            last_name,
        ):
            filled += 1
        field_map = {
            "email": ['input[type="email"]', 'input[name*="email" i]', 'input[id*="email" i]'],
            "phone": ['input[type="tel"]', 'input[name*="phone" i]', 'input[id*="phone" i]'],
            "website": ['input[name*="website" i]', 'input[id*="website" i]', 'input[placeholder*="website" i]'],
            "instagram": ['input[name*="instagram" i]', 'input[id*="instagram" i]'],
            "tiktok": ['input[name*="tiktok" i]', 'input[id*="tiktok" i]'],
            "youtube": ['input[name*="youtube" i]', 'input[id*="youtube" i]'],
        }
        for key, selectors in field_map.items():
            if _fill_first_visible(page, selectors, profile.get(key, "")):
                filled += 1
        if _fill_first_visible(
            page,
            ['textarea[name*="message" i]', 'textarea[id*="message" i]', "textarea"],
            profile.get("message", ""),
        ):
            filled += 1
        return filled

    def _submit_if_needed(page) -> bool:
        selectors = [
            'button:has-text("Apply")',
            'button:has-text("Submit")',
            'button:has-text("Send")',
            'button[type="submit"]',
            'input[type="submit"]',
        ]
        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if loc.count() and loc.is_visible() and loc.is_enabled():
                    txt = (loc.inner_text(timeout=1200) or "").strip().lower()
                    if "apply" in txt or "submit" in txt or "send" in txt or sel in (
                        'button[type="submit"]',
                        'input[type="submit"]',
                    ):
                        loc.click(timeout=2500)
                        return True
            except Exception:
                continue
        return False

    def _click_next_if_visible(page) -> bool:
        selectors = [
            'button:has-text("Next")',
            'button:has-text("Continue")',
            'button:has-text("Tiếp")',
            'input[type="submit"][value*="Next" i]',
        ]
        for ctx in _contexts(page):
            for sel in selectors:
                try:
                    loc = ctx.locator(sel).first
                    if loc.count() and loc.is_visible() and loc.is_enabled():
                        loc.click(timeout=2000)
                        page.wait_for_timeout(1000)
                        return True
                except Exception:
                    continue
        return False

    def _click_send_application(page) -> bool:
        # Strict: chỉ coi là "hoàn thành" khi bấm được Send application / Submit application
        selectors = [
            'button:has-text("Send application")',
            'button:has-text("Submit application")',
        ]
        for ctx in _contexts(page):
            for sel in selectors:
                try:
                    loc = ctx.locator(sel).first
                    if loc.count() and loc.is_visible() and loc.is_enabled():
                        txt = (loc.inner_text(timeout=1000) or "").lower()
                        if ("send application" in txt) or ("submit application" in txt):
                            loc.click(timeout=2500)
                            page.wait_for_timeout(1200)
                            return True
                except Exception:
                    continue
        return False

    def _click_submit_fallback(page) -> bool:
        selectors = [
            'button:has-text("Submit")',
            'button:has-text("Apply")',
            'button:has-text("Send")',
            'button[type="submit"]',
            'input[type="submit"]',
        ]
        for ctx in _contexts(page):
            for sel in selectors:
                try:
                    loc = ctx.locator(sel).first
                    if loc.count() and loc.is_visible() and loc.is_enabled():
                        loc.click(timeout=2500)
                        page.wait_for_timeout(1200)
                        return True
                except Exception:
                    continue
        return False

    # (moved) _wait_shopify_login_if_needed is implemented later with max_wait_sec support.

    def _url_path(url: str) -> str:
        try:
            p = urlparse(url)
            return (p.path or "/").lower()
        except Exception:
            return "/"

    def _is_shopify_accounts(page) -> bool:
        try:
            u = (page.url or "").lower()
        except Exception:
            return False
        return "accounts.shopify.com" in u

    def _is_shopify_signup(page) -> bool:
        try:
            u = (page.url or "").lower()
        except Exception:
            return False
        if "accounts.shopify.com" not in u:
            return False
        # Match /signup, /signup/, /signup?... (bỏ query)
        return _url_path(u).startswith("/signup")

    def _wait_shopify_login_if_needed(page, max_wait_sec: float | None = None) -> bool:
        """
        Chờ user login thủ công nhưng không vượt quá max_wait_sec (nếu có).
        Trả về True nếu đã thoát khỏi accounts.shopify.com, False nếu timeout.
        """
        try:
            u = page.url or ""
        except Exception:
            u = ""
        if "accounts.shopify.com" not in u:
            return True
        limit = float(max_wait_sec) if max_wait_sec is not None else float(login_timeout_sec)
        limit = max(1.0, limit)
        _log(
            "Đang ở trang Shopify Accounts (chưa login). "
            f"Vui lòng đăng nhập trong cửa sổ vừa mở. Tối đa {int(limit)}s…"
        )
        deadline = time.monotonic() + float(limit)
        while time.monotonic() < deadline:
            page.wait_for_timeout(600)
            try:
                cur = page.url or ""
            except Exception:
                cur = ""
            if "accounts.shopify.com" not in cur:
                _log("Đã login xong, tiếp tục Auto Apply.")
                return True
        _log("Timeout chờ login Shopify.")
        return False

    def _block_until_logged_in_or_fail(page, reason: str = "", brand_deadline: float | None = None) -> None:
        """
        Nếu đang ở Shopify Accounts thì chặn cứng tại đây, chờ user login xong mới cho chạy tiếp.
        Không được chuyển sang link khác trong lúc chờ.
        """
        if not _is_shopify_accounts(page):
            return
        # Rule mới:
        # - Nếu bị chuyển sang accounts.shopify.com/signup => phải chờ user login để tiếp tục.
        # - Nếu chưa login được thì tối đa 200s sẽ bỏ brand và nhảy sang brand khác.
        # Lưu ý: không ràng buộc vào brand_deadline (vì per-brand timeout thường chỉ ~30s).
        max_wait = 200.0
        suffix = f" ({reason})" if reason else ""
        _log(
            "Phát hiện chưa login Shopify. Đang tạm dừng Auto Apply"
            f"{suffix}. Vui lòng login trong cửa sổ này..."
        )
        ok = _wait_shopify_login_if_needed(page, max_wait_sec=max_wait)
        if not ok:
            # Không treo: quá ngân sách brand thì nhảy brand khác.
            raise _SkipBrand("login-timeout")
        # Sau khi login xong, chờ thêm chút để redirect ổn định.
        page.wait_for_timeout(800)

    def _ensure_logged_in_before_run(page) -> None:
        """
        Luôn check login trước khi chạy Auto Apply:
        - Mở Collabs
        - Nếu bị chuyển sang accounts.shopify.com thì bắt buộc user login xong mới chạy tiếp
        """
        _log("Kiểm tra trạng thái login Shopify Collabs…")
        try:
            page.goto("https://collabs.shopify.com/", wait_until="domcontentloaded", timeout=45000)
        except Exception:
            pass
        if not _wait_shopify_login_if_needed(page):
            raise RuntimeError("Chưa login Shopify Collabs (timeout chờ đăng nhập).")
        _log("Đã xác nhận login Shopify Collabs.")

    def _can_open_collabs(context) -> bool:
        probe = context.new_page()
        try:
            probe.goto("https://collabs.shopify.com/", wait_until="domcontentloaded", timeout=30000)
        except Exception:
            try:
                probe.close()
            except Exception:
                pass
            return False
        try:
            u = (probe.url or "").lower()
        except Exception:
            u = ""
        try:
            probe.close()
        except Exception:
            pass
        return "collabs.shopify.com" in u or "accounts.shopify.com" in u

    total = len(links)
    ok_count = 0
    submit_count = 0
    submitted_items: list[dict] = []
    attempted_items: list[dict] = []

    def _domain_from_link(raw_link: str) -> str:
        try:
            u = str(raw_link or "").strip()
            if not u:
                return ""
            p = urlparse(u)
            host = (p.netloc or "").strip().lower()
            if host.startswith("www."):
                host = host[4:]
            return host
        except Exception:
            return ""

    def _brand_name_from_page_or_link(p, raw_link: str) -> str:
        name = ""
        try:
            if p and hasattr(p, "title"):
                name = str(p.title() or "").strip()
        except Exception:
            name = ""
        if name.strip().lower() in ("shopify collabs", "collabs", "shopify"):
            name = ""
        if not name:
            try:
                up = urlparse(str(raw_link or "").strip())
                host = (up.netloc or "").strip().lower()
                if host.startswith("www."):
                    host = host[4:]
                name = host or str(raw_link or "").strip()
            except Exception:
                name = str(raw_link or "").strip()
        return name or str(raw_link or "").strip()

    def _record_attempt(*, p=None, link: str, submitted: bool, note: str = "") -> None:
        attempted_items.append(
            {
                "brand": _brand_name_from_page_or_link(p, link),
                "domain": _domain_from_link(link),
                "email": str(profile.get("email") or "").strip(),
                "link": str(link or "").strip(),
                "submitted": bool(submitted),
                "note": str(note or "").strip(),
            }
        )
    with sync_playwright() as p:
        browser = None
        context = None
        page = None
        using_cdp = bool(cdp_url)
        if not using_cdp:
            raise RuntimeError(
                "Chế độ browser ảo đã tắt. Vui lòng dùng trình duyệt thật đã login qua CDP "
                "(ví dụ http://127.0.0.1:9222)."
            )
        if using_cdp:
            _log(f"Kết nối tới trình duyệt đang mở qua CDP: {cdp_url}")
            browser = p.chromium.connect_over_cdp(cdp_url)
            # Dùng context hiện có (giữ session/login). Nếu không có thì tạo mới.
            context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.new_page()
        try:
            # Bắt buộc check login trước khi chạy bất kỳ link nào.
            # Không chạy song song trong lúc user đang đăng nhập.
            if login_first:
                _check_stop("trước khi kiểm tra login")
                _ensure_logged_in_before_run(page)

            for i, link in enumerate(links, start=1):
                _check_stop(f"trước khi mở link {i}/{total}")
                _log(f"[{i}/{total}] Mở: {link}")
                brand_page = context.new_page()
                # Timeout tối đa cho 1 brand: quá thời gian -> đóng tab và nhảy brand khác.
                per_brand_deadline = time.monotonic() + float(max(20, int(brand_timeout_sec or 120)))
                try:
                    _check_stop(f"trước khi goto link {i}/{total}")
                    brand_page.goto(link, wait_until="domcontentloaded", timeout=45000)
                    # Link đầu tiên thường load chậm widget form -> đợi ổn định thêm.
                    if i == 1:
                        try:
                            brand_page.wait_for_load_state("networkidle", timeout=12000)
                        except Exception:
                            brand_page.wait_for_timeout(1200)
                except PlaywrightTimeoutError:
                    _log("  - Timeout mở trang, bỏ qua.")
                    _record_attempt(link=link, submitted=False, note="timeout_open_page")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                    continue
                except Exception as exc:
                    _log(f"  - Lỗi mở trang: {exc}")
                    _record_attempt(link=link, submitted=False, note=f"error_open_page: {exc}")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                    continue
                # Nếu ngay sau goto bị redirect sang signup thì bỏ brand.
                try:
                    _block_until_logged_in_or_fail(
                        brand_page, reason="sau khi mở link apply", brand_deadline=per_brand_deadline
                    )
                except _SkipBrand:
                    _record_attempt(p=brand_page, link=link, submitted=False, note="login_or_signup_blocked")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                    continue
                # Redirect login có thể xảy ra chậm sau khi goto.
                _check_stop("sau khi mở trang")
                brand_page.wait_for_timeout(900)
                try:
                    _block_until_logged_in_or_fail(
                        brand_page, reason="sau khi mở link apply", brand_deadline=per_brand_deadline
                    )
                except _SkipBrand:
                    _record_attempt(p=brand_page, link=link, submitted=False, note="login_or_signup_blocked")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                    continue
                # Chỉ điền trên collabs.shopify.com. Nếu link ngoài collabs thì phải có Apply now.
                if not _is_collabs_page(brand_page):
                    pages_before_apply = set()
                    try:
                        pages_before_apply = {id(p) for p in context.pages}
                    except Exception:
                        pages_before_apply = set()
                    clicked = _click_apply_now_only(brand_page)
                    if not clicked:
                        _log("  - Không tìm thấy nút Apply now, đóng tab và bỏ qua brand.")
                        _record_attempt(p=brand_page, link=link, submitted=False, note="no_apply_cta")
                        try:
                            brand_page.close()
                        except Exception:
                            pass
                        continue
                    _log("  - Đã bấm Apply now.")
                    _check_stop("sau khi bấm Apply now")
                    brand_page.wait_for_timeout(900)
                    # Nếu mở tab mới thì chuyển sang tab collabs mới đó.
                    target_page = _resolve_collabs_page_after_apply(context, brand_page)
                    if target_page is not brand_page:
                        _log("  - Apply mở tab mới, chuyển sang tab collabs để điền form.")
                        try:
                            # Đóng tab cũ nếu vẫn còn và không cần nữa.
                            if not _is_collabs_page(brand_page):
                                brand_page.close()
                        except Exception:
                            pass
                        brand_page = target_page
                    else:
                        # Có thể tab mới mở chậm, thử quét lại 1 lần.
                        try:
                            if pages_before_apply:
                                for p in context.pages:
                                    if id(p) not in pages_before_apply and _is_collabs_page(p):
                                        _log("  - Phát hiện tab collabs mới (trễ), chuyển sang tab mới.")
                                        try:
                                            if not _is_collabs_page(brand_page):
                                                brand_page.close()
                                        except Exception:
                                            pass
                                        brand_page = p
                                        break
                        except Exception:
                            pass
                    try:
                        _block_until_logged_in_or_fail(
                            brand_page, reason="sau khi bấm Apply now", brand_deadline=per_brand_deadline
                        )
                    except _SkipBrand:
                        _record_attempt(p=brand_page, link=link, submitted=False, note="login_or_signup_blocked")
                        try:
                            brand_page.close()
                        except Exception:
                            pass
                        continue
                if not _is_collabs_page(brand_page):
                    _log("  - Sau Apply now vẫn không vào collabs.shopify.com, bỏ qua brand.")
                    _record_attempt(p=brand_page, link=link, submitted=False, note="not_in_collabs_after_apply")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                    continue
                _check_stop("trước khi bắt đầu điền form")
                brand_page.wait_for_timeout(700)

                # Yêu cầu: phải mở được collabs.shopify.com mới bắt đầu điền.
                if not _can_open_collabs(context):
                    _log("  - Không mở được collabs.shopify.com sau khi Apply, bỏ qua brand này.")
                    _record_attempt(p=brand_page, link=link, submitted=False, note="cannot_open_collabs")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                    continue

                # Multi-step apply: fill page -> click Next (if any) -> repeat -> send application.
                total_filled = 0
                sent = False
                timed_out = False
                skip_brand = False
                for step in range(1, 8):
                    # Per-brand hard timeout: không treo quá 2 phút
                    if time.monotonic() > per_brand_deadline:
                        _log("  - Quá 2 phút cho brand này => đóng tab và nhảy sang brand khác.")
                        timed_out = True
                        try:
                            brand_page.close()
                        except Exception:
                            pass
                        break
                    _check_stop(f"trong form step {step} ({i}/{total})")
                    try:
                        _block_until_logged_in_or_fail(
                            brand_page, reason=f"trước bước form {step}", brand_deadline=per_brand_deadline
                        )
                    except _SkipBrand:
                        _log("  - Bị chuyển sang Shopify signup => bỏ qua brand này.")
                        _record_attempt(p=brand_page, link=link, submitted=False, note="login_or_signup_blocked")
                        try:
                            brand_page.close()
                        except Exception:
                            pass
                        skip_brand = True
                        break
                    _stable_page_before_actions(brand_page)
                    _fill_date_of_birth(brand_page)
                    _select_shipping_location(brand_page)
                    choice_picked = _pick_choices_yes_or_first(brand_page)
                    purchase_picked = _select_purchase_before_choice(brand_page)
                    selects_picked = _select_first_for_unknown_selects(brand_page)
                    # Sau khi auto-pick radio "Yes/First", override lại riêng cho Identify để không bị chọn nhầm.
                    _set_identify_choice(brand_page)
                    filled_step = _fill_visible_text_fields(brand_page)
                    total_filled += filled_step
                    _log(
                        f"  - Bước {step}: điền thêm {filled_step} ô, chọn {choice_picked} nhóm lựa chọn, "
                        f"purchase-picked={purchase_picked}, "
                        f"chọn mặc định {selects_picked} ô select."
                    )

                    # Điền xong + chọn xong mới bấm Next/Submit
                    _check_stop("trước khi bấm Next/Send")
                    brand_page.wait_for_timeout(500)

                    if _click_send_application(brand_page):
                        sent = True
                        submit_count += 1
                        _log("  - Đã bấm Send application (hoàn thành).")
                        break
                    if _click_next_if_visible(brand_page):
                        _log("  - Đã bấm Next, sang bước kế.")
                        continue
                    # No Next and no Send; try legacy submit when allowed.
                    if auto_submit and _click_submit_fallback(brand_page):
                        _log("  - Đã bấm submit/apply (fallback) nhưng chưa chắc là 'Send application'.")
                    break

                if timed_out or skip_brand:
                    if timed_out:
                        _record_attempt(p=brand_page, link=link, submitted=False, note="brand_timeout")
                    continue

                if total_filled <= 0:
                    if i == 1:
                        # Link đầu tiên hay render chậm, thử 1 lần bổ sung trước khi bỏ.
                        brand_page.wait_for_timeout(1200)
                        _pick_choices_yes_or_first(brand_page)
                        _select_purchase_before_choice(brand_page)
                        retry_filled = _fill_visible_text_fields(brand_page)
                        total_filled += retry_filled
                        _log(f"  - Retry link đầu: điền thêm {retry_filled} ô.")
                    if total_filled <= 0:
                        _log("  - Không nhận diện được ô để điền (có thể do custom widget/frame lạ).")
                        _record_attempt(p=brand_page, link=link, submitted=False, note="no_fields_detected")
                        try:
                            brand_page.close()
                        except Exception:
                            pass
                        continue
                ok_count += 1
                _log(f"  - Tổng đã điền: {total_filled} ô. Trạng thái gửi: {'đã gửi' if sent else 'chưa gửi'}.")
                # Rule mới:
                # - Chỉ đóng tab khi bấm được "Send application"/"Submit application" (sent=True)
                # - Nếu chưa hoàn thành thì giữ tab mở, chuyển sang link khác để apply tiếp
                if sent:
                    # Lưu lịch sử submit thành công cho từng brand/link.
                    brand_name = ""
                    try:
                        brand_name = (
                            (brand_page.title() or "").strip()
                            if brand_page and hasattr(brand_page, "title")
                            else ""
                        )
                    except Exception:
                        brand_name = ""
                    # Title "Shopify Collabs" là tên chung, không phải tên brand thực.
                    # Khi gặp title chung thì fallback sang domain từ link gốc.
                    if brand_name.strip().lower() in (
                        "shopify collabs",
                        "collabs",
                        "shopify",
                    ):
                        brand_name = ""
                    if not brand_name:
                        try:
                            p = urlparse(str(link or "").strip())
                            host = (p.netloc or "").strip().lower()
                            if host.startswith("www."):
                                host = host[4:]
                            brand_name = host or str(link or "").strip()
                        except Exception:
                            brand_name = str(link or "").strip()
                    submitted_items.append(
                        {
                            "brand": brand_name,
                            "email": str(profile.get("email") or "").strip(),
                            "link": str(link or "").strip(),
                        }
                    )
                    _record_attempt(p=brand_page, link=link, submitted=True, note="submitted")
                    try:
                        brand_page.close()
                    except Exception:
                        pass
                else:
                    _log("  - Chưa bấm được 'Send application' => GIỮ tab này mở và chuyển sang brand tiếp theo.")
                    _record_attempt(p=brand_page, link=link, submitted=False, note="not_submitted_kept_open")
            # end for link
        finally:
            try:
                if page:
                    page.close()
            except Exception:
                pass
            # Nếu dùng CDP thì không được đóng browser đang chạy của user.
            if not using_cdp:
                try:
                    if context:
                        context.close()
                except Exception:
                    pass
                try:
                    if browser:
                        browser.close()
                except Exception:
                    pass
            else:
                try:
                    if browser:
                        browser.close()
                except Exception:
                    pass
    return {
        "total": total,
        "filled": ok_count,
        "submitted": submit_count,
        "submitted_items": submitted_items,
        "attempted_items": attempted_items,
    }

