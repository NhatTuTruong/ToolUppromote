import os
import re
import multiprocessing
import queue as pyqueue
import subprocess
import sys
import time
import json
import uuid
from io import BytesIO
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

from edge_cdp import (
    EDGE_CDP_ACCOUNT_MAX,
    cdp_url_host_port,
    default_user_data_dir_for_port,
    ensure_edge_cdp_running,
    port_for_collabs_account,
    cdp_url_for_collabs_account,
)

from auto_apply_child import run_auto_apply_in_subprocess

from flask import Flask, jsonify, render_template, request, send_file

import filter as core
import auto_apply as auto_apply_core
from app import (
    ENV_PATH,
    apply_settings_for_run,
    filter_web_settings_payload,
    load_env_defaults,
    offer_passes_filters,
    row_is_dat,
    save_env,
)

import license_guard
from runtime_paths import app_dir, bundle_dir


BASE_DIR = app_dir()
license_guard.set_paths(BASE_DIR)


def _env_flag(name: str, default: bool = True) -> bool:
    raw = str(os.getenv(name, "")).strip().lower()
    if raw == "":
        return bool(default)
    if raw in {"1", "true", "yes", "y", "on", "enable", "enabled"}:
        return True
    if raw in {"0", "false", "no", "n", "off", "disable", "disabled"}:
        return False
    return bool(default)


def auto_apply_collabs_enabled() -> bool:
    # Feature flag: bật/tắt Auto Apply Collabs trên server + theo license key.
    if not _env_flag("ENABLE_AUTO_APPLY_COLLABS", True):
        return False
    try:
        lic = license_guard.license_status_payload()
        return bool(lic.get("auto_apply_collabs_enabled", True))
    except Exception:
        return True


class RunControl:
    def __init__(self):
        self.stop_event = threading.Event()
        self.pause_event = threading.Event()

    def stop(self):
        self.stop_event.set()

    def toggle_pause(self):
        if self.pause_event.is_set():
            self.pause_event.clear()
            return False
        self.pause_event.set()
        return True

    def wait_if_paused(self):
        while self.pause_event.is_set() and not self.stop_event.is_set():
            time.sleep(0.2)

    def should_stop(self):
        return self.stop_event.is_set()


class AppState:
    def __init__(self):
        self.lock = threading.Lock()
        self.ack_lock = threading.Lock()
        self.running = False
        self.paused = False
        self.status = "Sảnh"
        self.progress = 0.0
        self.logs = []
        self.log_ack_seen = 0
        self.control = None
        self.worker = None
        self.output_file = ""
        self.output_files = []

    def add_log(self, message: str):
        with self.lock:
            self.logs.append(message)

    def log_count(self) -> int:
        with self.lock:
            return len(self.logs)

    def notify_log_displayed(self, seen_total: int) -> None:
        """Client đã vẽ xong tới mốc len(logs) == seen_total (sau poll /api/logs)."""
        if seen_total < 0:
            return
        with self.ack_lock:
            if seen_total > self.log_ack_seen:
                self.log_ack_seen = seen_total

    def wait_log_displayed(self, target_total: int, deadline_sec: float = 300.0) -> str:
        """
        Chờ client xác nhận đã hiển thị tới target_total (len logs sau add_log).
        Trả về 'ok' | 'stop' | 'timeout'.
        """
        deadline = time.monotonic() + deadline_sec
        ctrl = self.control
        while time.monotonic() < deadline:
            if ctrl and ctrl.should_stop():
                return "stop"
            with self.ack_lock:
                if self.log_ack_seen >= target_total:
                    return "ok"
            time.sleep(0.012)
        return "timeout"

    def get_logs(self, since: int):
        with self.lock:
            total = len(self.logs)
            if since < 0:
                since = 0
            return self.logs[since:], total


STATE = AppState()


class AutoApplyJob:
    """Một slot song song: subprocess + meta (log chi tiết trong logs của job)."""

    def __init__(
        self,
        job_id: str,
        *,
        cdp_url: str,
        edge_user_data_dir: str,
        profile_email: str,
        links: list[str],
        process: multiprocessing.Process,
        stop_event: multiprocessing.synchronize.Event,
    ):
        self.job_id = job_id
        self.cdp_url = cdp_url
        self.edge_user_data_dir = edge_user_data_dir
        self.profile_email = profile_email
        self.links = links
        self.process = process
        self.stop_event = stop_event
        self.logs: list[str] = []
        self.status = "Running"
        self.result: dict | None = None
        self.error: str = ""


class AutoApplyState:
    """Nhiều job song song (multiprocessing); tương thích API cũ (logs/file/running)."""

    def __init__(self):
        self.lock = threading.Lock()
        self.running = False
        self.status = "Idle"
        self.logs: list[str] = []
        self.result: dict | None = None
        self.error: str = ""
        self.stop_event = threading.Event()
        self.worker: threading.Thread | None = None
        self.current_file: str = ""
        # --- multi ---
        self.batch_id: str = ""
        self.mp_ctx: multiprocessing.context.BaseContext | None = None
        self.shared_log_queue: multiprocessing.Queue | None = None
        self.shared_result_queue: multiprocessing.Queue | None = None
        self.jobs: dict[str, AutoApplyJob] = {}
        self.parallel_link_split: str = ""
        self.account_mode: str = ""
        self.sequential_multi: bool = False

    def reset(self):
        self.running = False
        self.status = "Idle"
        self.logs = []
        self.result = None
        self.error = ""
        self.stop_event = threading.Event()
        self.worker = None
        self.current_file = ""
        self.batch_id = ""
        self.mp_ctx = None
        self.shared_log_queue = None
        self.shared_result_queue = None
        self.jobs = {}
        self.parallel_link_split = ""
        self.account_mode = ""
        self.sequential_multi = False

    def add_log(self, msg: str):
        with self.lock:
            self.logs.append(str(msg))


AUTO_APPLY_STATE = AutoApplyState()
AUTO_APPLY_HISTORY_PATH = BASE_DIR / "auto-apply-history.json"
AUTO_APPLY_HISTORY_LOCK = threading.Lock()


def _load_auto_apply_history() -> list[dict]:
    if not AUTO_APPLY_HISTORY_PATH.exists():
        return []
    try:
        raw = AUTO_APPLY_HISTORY_PATH.read_text(encoding="utf-8")
    except OSError:
        return []
    try:
        data = json.loads(raw or "[]")
    except Exception:
        return []
    return data if isinstance(data, list) else []


def _save_auto_apply_history(items: list[dict]) -> None:
    try:
        AUTO_APPLY_HISTORY_PATH.write_text(
            json.dumps(items, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except OSError:
        pass


def _append_auto_apply_history(entry: dict) -> None:
    with AUTO_APPLY_HISTORY_LOCK:
        items = _load_auto_apply_history()
        items.insert(0, entry)
        # Giữ tối đa 200 phiên để file không phình to.
        items = items[:200]
        _save_auto_apply_history(items)


_root = bundle_dir()
app = Flask(
    __name__,
    template_folder=str(_root / "templates"),
    static_folder=str(_root / "static"),
)


def _refresh_license_env_from_file() -> None:
    """
    Đồng bộ các biến license từ .env vào process hiện tại.
    Cần vì core.load_env_file() không overwrite biến đã tồn tại trong os.environ.
    """
    disk = core.parse_env_file(ENV_PATH)
    for key in (
        "AFF_LICENSE_API_BASE_URL",
        "AFF_LICENSE_SERVER_URL",
        "AFF_LICENSE_API_TOKEN",
        "AFF_LICENSE_DAILY_LIMIT",
    ):
        if key in disk:
            os.environ[key] = str(disk.get(key) or "").strip()


def _resolve_export_page_range(filters: dict, source: str) -> tuple[int, int]:
    try:
        start_page = int(filters.get("start_page") or 1)
    except (TypeError, ValueError):
        start_page = 1
    if start_page < 1:
        start_page = 1

    end_raw = filters.get("end_page")
    try:
        if end_raw is None:
            end_page = start_page
        elif str(end_raw).strip() == "":
            end_page = start_page
        else:
            end_page = int(end_raw)
    except (TypeError, ValueError):
        end_page = start_page

    if end_page < start_page:
        end_page = start_page

    max_pages_cap = None
    if source == "goaffpro":
        max_pages_cap = core.goaffpro_max_pages_cap()
    elif source == "refersion":
        max_pages_cap = core.refersion_max_pages_cap()
    elif source == "collabs":
        max_pages_cap = core.collabs_max_pages_cap()
    else:
        max_pages_cap = core.uppromote_max_pages_cap()

    if max_pages_cap is not None:
        end_page = min(end_page, max_pages_cap)

    return start_page, end_page


def fetch_offers_uppromote(filters: dict) -> list:
    base_url = (os.getenv("UPPROMOTE_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu UPPROMOTE_API_URL trong cài đặt")
    core.enforce_fixed_fetch_defaults()
    max_pages_cap = core.uppromote_max_pages_cap()
    start_page = int(filters.get("start_page") or 1)
    end_raw = filters.get("end_page")
    if end_raw is None:
        end_page = 1
    elif str(end_raw).strip() == "":
        end_page = None
    else:
        end_page = int(end_raw)
    if start_page < 1:
        start_page = 1
    if end_page is not None and end_page < start_page:
        end_page = start_page
    if end_page is not None and max_pages_cap is not None:
        end_page = min(end_page, max_pages_cap)
    delay_ms = int(
        os.getenv("UPPROMOTE_PAGE_DELAY_MS", str(core.DEFAULT_UPPROMOTE_PAGE_DELAY_MS))
        or str(core.DEFAULT_UPPROMOTE_PAGE_DELAY_MS)
    )

    raw_offers = []
    page = start_page
    while True:
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return []
        STATE.add_log(f"Uppromote trang {page}: đang tải...")
        body = core.fetch_uppromote_page(base_url, page)
        payload = body.get("data") or {}
        page_items = payload.get("data") or []
        if not isinstance(page_items, list):
            page_items = []
        if not page_items:
            STATE.add_log(f"Uppromote trang {page}: hết dữ liệu, dừng phân trang.")
            break
        raw_offers.extend(page_items)
        STATE.add_log(f"Uppromote trang {page}: +{len(page_items)} offer (tổng {len(raw_offers)})")
        if end_page is not None and page >= end_page:
            STATE.add_log(f"Uppromote: đã tới trang kết thúc đã chọn: {end_page}")
            break
        if max_pages_cap is not None and page >= max_pages_cap:
            STATE.add_log(f"Uppromote: đã tới giới hạn trang trong cài đặt: {max_pages_cap}")
            break
        next_page = payload.get("next_page_url")
        if not next_page:
            break
        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)

    offers = []
    detail_delay_ms = int(os.getenv("UPPROMOTE_DETAIL_DELAY_MS", "50") or "50")
    total_raw = len(raw_offers)
    for idx, raw_offer in enumerate(raw_offers, start=1):
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return []
        detail = {}
        shop_id = raw_offer.get("shop_id")
        if shop_id:
            try:
                detail = core.fetch_uppromote_offer_detail(shop_id)
            except Exception as exc:
                STATE.add_log(f"Lỗi chi tiết shop_id={shop_id}: {exc}")
        offers.append(core.map_uppromote_offer(raw_offer, detail))
        if idx % 10 == 0 or idx == total_raw:
            STATE.add_log(f"Tiến độ chi tiết offer: {idx}/{total_raw}")
        if detail_delay_ms > 0:
            time.sleep(detail_delay_ms / 1000)
    return offers


def fetch_offers_goaffpro(filters: dict) -> list:
    base_url = (os.getenv("GOAFFPRO_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu GOAFFPRO_API_URL trong cài đặt")
    core.enforce_fixed_fetch_defaults()
    limit = int(
        os.getenv("GOAFFPRO_LIMIT", str(core.DEFAULT_OFFERS_PER_PAGE)) or str(core.DEFAULT_OFFERS_PER_PAGE)
    )
    max_pages_cap = core.goaffpro_max_pages_cap()
    start_page = int(filters.get("start_page") or 1)
    if start_page < 1:
        start_page = 1
    end_raw = filters.get("end_page")
    if end_raw is None:
        end_page = 1
    elif str(end_raw).strip() == "":
        end_page = None
    else:
        end_page = int(end_raw)
    if end_page is not None and end_page < start_page:
        end_page = start_page
    if end_page is not None and max_pages_cap is not None:
        end_page = min(end_page, max_pages_cap)
    delay_ms = int(
        os.getenv("GOAFFPRO_PAGE_DELAY_MS", str(core.DEFAULT_GOAFFPRO_PAGE_DELAY_MS))
        or str(core.DEFAULT_GOAFFPRO_PAGE_DELAY_MS)
    )

    raw_stores = []
    page = start_page
    while True:
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return []
        offset = (page - 1) * limit
        STATE.add_log(f"Goaffpro trang {page} (offset={offset}): đang tải...")
        body = core.fetch_goaffpro_page(base_url, offset, limit)
        stores = body.get("stores") or []
        if not isinstance(stores, list):
            stores = []
        if not stores:
            STATE.add_log(f"Goaffpro trang {page}: hết dữ liệu, dừng phân trang.")
            break
        raw_stores.extend(stores)
        STATE.add_log(f"Goaffpro trang {page}: +{len(stores)} cửa hàng (tổng {len(raw_stores)})")
        try:
            total_n = int(body.get("count")) if body.get("count") is not None else None
        except Exception:
            total_n = None
        if total_n is not None and offset + len(stores) >= total_n:
            STATE.add_log("Goaffpro: đã lấy hết theo tổng từ API.")
            break
        if end_page is not None and page >= end_page:
            STATE.add_log(f"Goaffpro: đã tới trang kết thúc đã chọn: {end_page}")
            break
        if max_pages_cap is not None and page >= max_pages_cap:
            STATE.add_log(f"Goaffpro: đã tới giới hạn trang trong cài đặt: {max_pages_cap}")
            break
        if len(stores) < limit:
            break
        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)

    return [core.map_goaffpro_store(s) for s in raw_stores]


def fetch_offers_refersion(filters: dict) -> list:
    base_url = (os.getenv("REFERSION_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu REFERSION_API_URL trong cài đặt")
    core.enforce_fixed_fetch_defaults()
    max_pages_cap = core.refersion_max_pages_cap()
    start_page = int(filters.get("start_page") or 1)
    if start_page < 1:
        start_page = 1
    end_raw = filters.get("end_page")
    if end_raw is None:
        end_page = 1
    elif str(end_raw).strip() == "":
        end_page = None
    else:
        end_page = int(end_raw)
    if end_page is not None and end_page < start_page:
        end_page = start_page
    if end_page is not None and max_pages_cap is not None:
        end_page = min(end_page, max_pages_cap)
    delay_ms = int(
        os.getenv("REFERSION_PAGE_DELAY_MS", str(core.DEFAULT_REFERSION_PAGE_DELAY_MS))
        or str(core.DEFAULT_REFERSION_PAGE_DELAY_MS)
    )

    raw_offers = []
    page = start_page
    while True:
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return []
        STATE.add_log(f"Refersion trang {page}: đang tải...")
        body = core.fetch_refersion_page(base_url, page)
        payload = body.get("data") or {}
        offers = payload.get("offers") or []
        if not isinstance(offers, list):
            offers = []
        if not offers:
            STATE.add_log(f"Refersion trang {page}: hết dữ liệu, dừng phân trang.")
            break
        raw_offers.extend(offers)
        STATE.add_log(f"Refersion trang {page}: +{len(offers)} offer (tổng {len(raw_offers)})")
        try:
            total_n = int(payload.get("total_results")) if payload.get("total_results") is not None else None
        except Exception:
            total_n = None
        if total_n is not None and len(raw_offers) >= total_n:
            STATE.add_log("Refersion: đã lấy hết theo tổng từ API.")
            break
        if end_page is not None and page >= end_page:
            STATE.add_log(f"Refersion: đã tới trang kết thúc đã chọn: {end_page}")
            break
        if max_pages_cap is not None and page >= max_pages_cap:
            STATE.add_log(f"Refersion: đã tới giới hạn trang trong cài đặt: {max_pages_cap}")
            break
        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)

    return [core.map_refersion_offer(o) for o in raw_offers]


def fetch_offers_collabs(filters: dict) -> list:
    discovery_mode = str((filters or {}).get("discovery_mode") or "in_discovery").strip().lower()
    if discovery_mode == "outside_discovery":
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            return []
        dedup_enabled = str(os.getenv("COLLABS_OUTSIDE_DEDUP_PERSIST", "")).strip().lower() in {
            "1",
            "true",
            "yes",
            "y",
            "on",
            "enable",
            "enabled",
        }
        cse_k = (os.getenv("GOOGLE_CUSTOM_SEARCH_API_KEY") or os.getenv("GOOGLE_CSE_API_KEY") or "").strip()
        cse_cx = (os.getenv("GOOGLE_CUSTOM_SEARCH_ENGINE_ID") or os.getenv("GOOGLE_CSE_CX") or "").strip()
        if (os.getenv("COLLABS_OUTSIDE_GOOGLE_ACTOR_ID") or "").strip():
            prov = "Apify Google actor"
        elif cse_k and cse_cx:
            prov = "Google Custom Search JSON API"
        else:
            prov = "Bing RSS"
        raw_target = (filters or {}).get("outside_target_results")
        if raw_target is None or str(raw_target).strip() == "":
            raw_target = os.getenv("COLLABS_OUTSIDE_MAX_RESULTS", "80") or "80"
        try:
            target_n = max(10, min(30, int(raw_target)))
        except Exception:
            target_n = 30
        dedup_txt = "dedupe=ON" if dedup_enabled else "dedupe=OFF"
        fcollab = dict(filters or {})
        # Luôn sinh query mới mỗi lần lọc (không giữ _effective_outside_query từ request cũ).
        fcollab.pop(core.EFFECTIVE_OUTSIDE_QUERY_KEY, None)
        q_google = core.resolve_outside_discovery_query_string(fcollab)
        fcollab[core.EFFECTIVE_OUTSIDE_QUERY_KEY] = q_google
        STATE.add_log(
            f"Collabs ngoài Discovery: Google query: {q_google!r} — đang tìm brand ({prov}, {dedup_txt}, target={target_n})..."
        )
        offers = core.discover_collabs_outside_discovery_offers(
            fcollab, should_stop=STATE.control.should_stop
        )
        outside_stats = fcollab.pop(core.OUTSIDE_DISCOVERY_STATS_KEY, None)
        if STATE.control.should_stop():
            STATE.add_log("Collabs ngoài Discovery: đã dừng theo yêu cầu.")
        else:
            STATE.add_log(f"Collabs ngoài Discovery: đã tìm được {len(offers)} brand hợp lệ.")
            if isinstance(outside_stats, dict):
                tb = outside_stats.get("target_batch")
                cu = outside_stats.get("candidate_urls")
                STATE.add_log(
                    f"Collabs ngoài Discovery: giải thích số lượng — target={tb} là tối đa "
                    f"số URL lấy từ Google (theo cursor/query), không phải đảm bảo đủ brand; "
                    f"lần này có {cu} URL ứng viên từ tìm kiếm, sau khi kiểm tra trang collab + Apply now + link signup "
                    f"còn {len(offers)} brand."
                )
        return offers

    base_url = (os.getenv("COLLABS_API_URL") or "").strip()
    if not base_url:
        raise RuntimeError("Thiếu COLLABS_API_URL trong cài đặt")
    core.enforce_fixed_fetch_defaults()
    max_pages_cap = core.collabs_max_pages_cap()
    page_size = core.DEFAULT_COLLABS_LIMIT
    max_discovery_pages_per_run = 3
    max_discovery_records_per_run = page_size * max_discovery_pages_per_run
    start_page = int(filters.get("start_page") or 1)
    if start_page < 1:
        start_page = 1
    end_raw = filters.get("end_page")
    if end_raw is None:
        end_page = 1
    elif str(end_raw).strip() == "":
        end_page = None
    else:
        end_page = int(end_raw)
    if end_page is not None and end_page < start_page:
        end_page = start_page
    if end_page is not None and max_pages_cap is not None:
        end_page = min(end_page, max_pages_cap)
    delay_ms = int(
        os.getenv("COLLABS_PAGE_DELAY_MS", str(core.DEFAULT_COLLABS_PAGE_DELAY_MS))
        or str(core.DEFAULT_COLLABS_PAGE_DELAY_MS)
    )
    detail_delay_ms = int(os.getenv("COLLABS_DETAIL_DELAY_MS", "100") or "100")
    progress_floor = 0.0

    def _set_progress_floor(v: float) -> None:
        nonlocal progress_floor
        try:
            progress_floor = max(progress_floor, float(v))
        except Exception:
            return
        with STATE.lock:
            if progress_floor > STATE.progress:
                STATE.progress = progress_floor

    raw_nodes = []
    after = None
    page = 1
    if start_page > 1:
        STATE.add_log(f"Collabs: bắt đầu từ trang {start_page}")
    forced_end_page = start_page + max_discovery_pages_per_run - 1
    while True:
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return []
        if page >= start_page:
            STATE.add_log(f"Collabs trang {page}: đang tải...")
        body = core.fetch_collabs_page(base_url, page_size, after=after)
        data = body.get("data") or {}
        search = data.get("brandsNetworkSearch") or {}
        nodes = search.get("nodes") or []
        if not isinstance(nodes, list):
            nodes = []
        if not nodes:
            if page >= start_page:
                STATE.add_log(f"Collabs trang {page}: hết dữ liệu, dừng phân trang.")
            break
        if page >= start_page:
            raw_nodes.extend(nodes)
            STATE.add_log(f"Collabs trang {page}: +{len(nodes)} brand (tổng {len(raw_nodes)})")
            if len(raw_nodes) >= max_discovery_records_per_run:
                raw_nodes = raw_nodes[:max_discovery_records_per_run]
                STATE.add_log(
                    f"Collabs: đạt giới hạn mỗi lần lọc {max_discovery_records_per_run} brand "
                    f"(~{max_discovery_pages_per_run} trang), dừng phân trang."
                )
                break
            # Hiển thị tiến trình ngay trong pha crawl trang Collabs (0-25%).
            seen_pages = max(1, page - start_page + 1)
            _set_progress_floor(min(25.0, 5.0 + (seen_pages * 2.0)))
        info = search.get("pageInfo") or {}
        has_next = bool(info.get("hasNextPage"))
        after = info.get("endCursor")
        if end_page is not None and page >= end_page:
            STATE.add_log(f"Collabs: đã tới trang kết thúc đã chọn: {end_page}")
            break
        if page >= forced_end_page:
            STATE.add_log(
                f"Collabs: chỉ lấy tối đa {max_discovery_pages_per_run} trang liên tiếp mỗi lần "
                f"(trang {start_page} → {forced_end_page})."
            )
            break
        if not has_next:
            break
        if max_pages_cap is not None and page >= max_pages_cap:
            STATE.add_log(f"Collabs: đã tới giới hạn trang trong cài đặt: {max_pages_cap}")
            break
        page += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000)
    offers = []
    total_raw = len(raw_nodes)
    redirect_timeout_sec = int(os.getenv("COLLABS_REDIRECT_TIMEOUT_SEC", "20") or "20")
    redirect_delay_ms = int(os.getenv("COLLABS_REDIRECT_DELAY_MS", "50") or "50")
    signup_timeout_sec = int(os.getenv("COLLABS_SIGNUP_TIMEOUT_SEC", "20") or "20")
    signup_delay_ms = int(os.getenv("COLLABS_SIGNUP_DELAY_MS", "50") or "50")
    signup_by_host = {}
    for idx, node in enumerate(raw_nodes, start=1):
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return []
        detail_brand = {}
        gid = core.collabs_shopify_store_gid(node)
        if gid:
            try:
                detail_data = core.fetch_collabs_brand_profile(base_url, gid)
                detail_brand = detail_data.get("brand") if isinstance(detail_data.get("brand"), dict) else {}
            except Exception as exc:
                STATE.add_log(f"Lỗi detail collabs ({gid}): {exc}")
        mapped = core.map_collabs_brand(node, detail_brand)
        before_url = str(mapped.get("url") or "").strip()
        if before_url:
            final_url = core.resolve_redirected_url(before_url, timeout_sec=redirect_timeout_sec)
            if final_url and final_url != before_url:
                mapped["url"] = final_url
                STATE.add_log(f"Collabs redirect: {before_url} -> {final_url}")
        effective_url = str(mapped.get("url") or "").strip()
        hk = core.host_key(effective_url)
        signup_url = ""
        if hk:
            signup_url = signup_by_host.get(hk, "")
            if not signup_url:
                signup_url = core.discover_collabs_signup_url(effective_url, timeout_sec=signup_timeout_sec)
                signup_by_host[hk] = signup_url
        if signup_url:
            mapped["client_url"] = signup_url
            STATE.add_log(f"Collabs signup: {hk} -> {signup_url}")
        offers.append(mapped)
        if idx % 10 == 0 or idx == total_raw:
            STATE.add_log(f"Collabs detail: {idx}/{total_raw}")
        if total_raw > 0:
            # Pha detail Collabs: 25% -> 55%
            _set_progress_floor(25.0 + (idx / total_raw) * 30.0)
        if detail_delay_ms > 0:
            time.sleep(detail_delay_ms / 1000)
        if redirect_delay_ms > 0:
            time.sleep(redirect_delay_ms / 1000)
        if signup_delay_ms > 0:
            time.sleep(signup_delay_ms / 1000)
    return offers


def run_pipeline(settings: dict, min_traffic: int, filters: dict, source: str = "uppromote"):
    apply_settings_for_run(settings)
    core.enforce_fixed_fetch_defaults()
    os.environ["MIN_VISITS"] = str(min_traffic)
    STATE.add_log("Đang kiểm tra kết nối Apify...")
    apify_user = core.check_apify_connection()
    STATE.add_log(f"Kết nối Apify OK ({apify_user}). Bắt đầu tiến trình lọc.")

    src = (source or "uppromote").lower()
    if src == "goaffpro":
        STATE.add_log("Đang tải offer từ Goaffpro...")
        offers = fetch_offers_goaffpro(filters)
        snapshot_path = BASE_DIR / "goaffpro-offers-last.json"
    elif src == "refersion":
        STATE.add_log("Đang tải offer từ Refersion...")
        offers = fetch_offers_refersion(filters)
        snapshot_path = BASE_DIR / "refersion-offers-last.json"
    elif src == "collabs":
        STATE.add_log("Đang tải offer từ Shopify Collabs...")
        offers = fetch_offers_collabs(filters)
        snapshot_path = BASE_DIR / "collabs-offers-last.json"
    else:
        STATE.add_log("Đang tải offer từ Uppromote...")
        offers = fetch_offers_uppromote(filters)
        snapshot_path = BASE_DIR / "uppromote-offers-last.json"

    if STATE.control.should_stop():
        STATE.add_log("Đã dừng.")
        return

    if not offers:
        STATE.add_log("Không có offer trong phạm vi trang đã chọn.")
        return

    cap = license_guard.export_offer_cap(len(offers), src)
    if cap == 0:
        STATE.add_log(license_guard.zero_export_cap_log_message(src))
        return
    if cap < len(offers):
        partial = license_guard.export_cap_partial_log(len(offers), cap, src)
        if partial:
            STATE.add_log(partial)
        offers = offers[:cap]

    STATE.add_log(f"Đã tải {len(offers)} offer.")
    snapshot_path.write_text(
        core.json.dumps(
            {"fetchedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()), "count": len(offers), "offers": offers},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    from_offers = core.unique_hosts([o.get("url", "") for o in offers])
    domain_file = Path(os.getenv("DOMAIN_FILE", str(BASE_DIR / "domain.txt")))
    from_file = core.read_domains_from_txt(domain_file)
    domains = sorted(set(from_offers + from_file))
    if not domains:
        raise RuntimeError("Không có tên miền: thêm URL từ offer hoặc file domain.txt.")

    collabs_discovery_mode = str((filters or {}).get("discovery_mode") or "in_discovery").strip().lower()
    outside_similarweb_fallback = src == "collabs" and collabs_discovery_mode == "outside_discovery"

    STATE.add_log(f"Chạy Apify cho {len(domains)} tên miền...")
    items = []
    chunk_size = int(
        os.getenv("APIFY_MAX_DOMAINS_PER_RUN", str(core.DEFAULT_APIFY_MAX_DOMAINS_PER_RUN))
        or str(core.DEFAULT_APIFY_MAX_DOMAINS_PER_RUN)
    )
    for idx, part in enumerate(core.chunked(domains, chunk_size), start=1):
        STATE.control.wait_if_paused()
        if STATE.control.should_stop():
            STATE.add_log("Đã dừng.")
            return
        STATE.add_log(f"Apify đợt {idx}: {len(part)} tên miền")
        dataset_id, apify_tok = core.apify_call_actor(part)
        batch = core.apify_list_items(dataset_id, token=apify_tok)
        if outside_similarweb_fallback:
            batch, n_fb = core.merge_outside_similarweb_fallback_batch(part, batch)
            if n_fb:
                STATE.add_log(
                    f"Collabs ngoài Discovery: Similarweb fallback Apify cho {n_fb} domain "
                    f"(thiếu traffic từ actor mặc định; xem COLLABS_OUTSIDE_SIMILARWEB_FALLBACK_ACTOR trong .env)."
                )
        items.extend(batch)

    by_host = {}
    for item in items:
        site = core.apify_site_field(item)
        key = core.host_key(site)
        if key:
            by_host[key] = item

    net_prefix = (
        "goaffpro"
        if src == "goaffpro"
        else ("refersion" if src == "refersion" else ("collabs" if src == "collabs" else "uppromote"))
    )
    export_start_page, export_end_page = _resolve_export_page_range(filters, src)
    now = datetime.now()
    date_part = f"{now.day}-{now.month}-{now.year}"
    time_part = f"{now.hour}-{now.minute:02d}"
    if src == "collabs" and collabs_discovery_mode == "outside_discovery":
        # File riêng cho luồng search Google ngoài Discovery.
        xlsx_name = f"collabs_google_{now.month}-{now.year}_{time_part}.xlsx"
    else:
        xlsx_name = f"{net_prefix}_page{export_start_page}-{export_end_page}_{date_part}_{time_part}.xlsx"
    xlsx_path = BASE_DIR / xlsx_name
    if src == "goaffpro":
        header = list(core.GOAFF_CSV_HEADER)
    elif src == "refersion":
        header = list(core.REFERSION_CSV_HEADER)
    elif src == "collabs":
        header = list(core.COLLABS_CSV_HEADER)
    else:
        header = list(core.UPPROMOTE_CSV_HEADER_VI)

    exported_rows = []
    total_offers = len(offers)
    export_interrupted = False

    def _flush_export_workbook(note: str = "") -> bool:
        if not exported_rows:
            return False
        try:
            core.write_xlsx_highlight_status(xlsx_path, header, exported_rows, status_col=0)
        except Exception as exc:
            STATE.add_log(f"Không ghi được Excel (cần openpyxl): {exc}")
            return False
        with STATE.lock:
            STATE.output_file = str(xlsx_path)
            STATE.output_files.insert(0, xlsx_path.name)
            STATE.output_files = STATE.output_files[:100]
        extra = f" — {note}" if note else ""
        abs_path = str(xlsx_path.resolve())
        STATE.add_log(
            f"Đã ghi Excel{extra}: {len(exported_rows)} dòng → {xlsx_path.name} "
            f"(cột trạng thái ĐẠT: nền xanh lá nhạt)\n"
            f"  Đường dẫn đầy đủ: {abs_path}"
        )
        return True

    try:
        STATE.add_log(
            f"─── Xuất dữ liệu: xử lý lần lượt {total_offers} offer; "
            f"file .xlsx được lưu khi hoàn tất hoặc khi dừng / lỗi (giữ các dòng đã xử lý) ───"
        )
        _w = STATE.wait_log_displayed(STATE.log_count())
        if _w == "stop":
            STATE.add_log("Đã dừng trước khi xuất từng offer.")
            return
        if _w == "timeout":
            STATE.add_log("[Cảnh báo] Chờ hiển thị log quá lâu — tiếp tục xử lý.")
        for idx, offer in enumerate(offers, start=1):
            STATE.control.wait_if_paused()
            if STATE.control.should_stop():
                STATE.add_log("Đã nhận lệnh dừng — lưu các dòng đã xử lý ra file…")
                export_interrupted = True
                break

            brand = offer.get("brand", "")
            url = offer.get("url", "")
            key = core.host_key(url)
            item = core.lookup_apify_item(url, by_host)
            eng = core.engagement_from_item(item)
            visits = core.parse_visits_from_engagement(eng)
            min_v = float(min_traffic)
            filters_ok = offer_passes_filters(offer, filters, src)
            traffic_ok = visits >= min_v
            status = core.STATUS_TRAFFIC_OK if row_is_dat(offer, filters, src, visits, min_v) else core.STATUS_TRAFFIC_FAIL
            if src == "goaffpro":
                row = core.build_goaff_csv_row(offer, item, status)
            elif src == "refersion":
                row = core.build_refersion_csv_row(offer, item, status)
            elif src == "collabs":
                row = core.build_collabs_csv_row(offer, item, status)
            else:
                row = core.build_uppromote_csv_row_vi(offer, item, status)
            exported_rows.append(row)
            license_guard.record_one_exported_row(src)
            visits_show = eng.get("VisitsFormatted") if eng else ""
            if not visits_show and visits:
                visits_show = int(visits) if visits == int(visits) else round(visits, 2)
            elif not visits_show:
                visits_show = 0
            block = (
                f"┌─ Record {idx}/{total_offers} ─────────────────────\n"
                f"│  Thương hiệu : {brand}\n"
                f"│  Domain/URL  : {key or url}\n"
                f"│  Trạng thái  : {status}\n"
                f"│  Lọc offer   : {'đạt' if filters_ok else 'chưa đạt'}\n"
                f"│  Traffic      : {'đạt' if traffic_ok else 'chưa đạt'} "
                f"(~{visits_show} so với ngưỡng {min_v})\n"
                f"└────────────────────────────────────────"
            )
            STATE.add_log(block)
            with STATE.lock:
                # Không cho tiến trình tụt lùi nếu trước đó đã cập nhật từ pha fetch/detail.
                STATE.progress = max(STATE.progress, (idx / total_offers) * 100)
            _w = STATE.wait_log_displayed(STATE.log_count())
            if _w == "stop":
                STATE.add_log("Đã nhận lệnh dừng — lưu các dòng đã xử lý ra file…")
                export_interrupted = True
                break
            if _w == "timeout":
                STATE.add_log("[Cảnh báo] Chờ hiển thị log quá lâu — tiếp record tiếp theo.")
    except Exception as exc:
        export_interrupted = True
        STATE.add_log(
            f"Lỗi trong vòng xuất: {exc} — đã xử lý xong {len(exported_rows)} dòng, sẽ ghi Excel phần đã có."
        )
        raise
    finally:
        if exported_rows:
            note = "dừng hoặc lỗi giữa chừng" if export_interrupted else "hoàn tất"
            _flush_export_workbook(note)


def _worker(settings: dict, min_traffic: int, filters: dict, source: str = "uppromote"):
    try:
        run_pipeline(settings, min_traffic, filters, source)
    except Exception as exc:
        STATE.add_log(f"LỖI: {exc}")
    finally:
        with STATE.lock:
            STATE.running = False
            STATE.paused = False
            STATE.status = "Sảnh"
            if STATE.progress < 100:
                STATE.progress = 100
        STATE.add_log("Hoàn tất.")


@app.get("/")
def index():
    return render_template("index.html", auto_apply_collabs_enabled=auto_apply_collabs_enabled())


@app.get("/api/settings")
def api_settings():
    return jsonify(load_env_defaults())


@app.post("/api/settings")
def api_save_settings():
    payload = filter_web_settings_payload(request.get_json(force=True))
    save_env(payload)
    return jsonify({"ok": True})


@app.get("/api/license")
def api_license():
    core.load_env_file(ENV_PATH)
    _refresh_license_env_from_file()
    license_guard.set_paths(BASE_DIR)
    return _no_cache_json(license_guard.license_status_payload())


@app.post("/api/license/activate")
def api_license_activate():
    core.load_env_file(ENV_PATH)
    _refresh_license_env_from_file()
    license_guard.set_paths(BASE_DIR)
    payload = request.get_json(force=True) or {}
    key = (payload.get("key") or "").strip()
    if not key:
        return jsonify({"ok": False, "error": "Nhập key kích hoạt."}), 400
    ok, msg = license_guard.activate_key(key)
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True, "message": msg, "license": license_guard.license_status_payload()})


@app.post("/api/license/deactivate")
def api_license_deactivate():
    core.load_env_file(ENV_PATH)
    _refresh_license_env_from_file()
    license_guard.set_paths(BASE_DIR)
    ok, msg = license_guard.deactivate_on_this_machine()
    if not ok:
        return jsonify({"ok": False, "error": msg}), 400
    return jsonify({"ok": True, "message": msg, "license": license_guard.license_status_payload()})


@app.post("/api/run")
def api_run():
    payload = request.get_json(force=True) or {}
    settings = payload.get("settings") or {}
    filters = payload.get("filters") or {}
    try:
        min_traffic = float(payload.get("min_traffic") if payload.get("min_traffic") not in (None, "") else 9000)
    except (TypeError, ValueError):
        min_traffic = 9000.0
    source = (payload.get("source") or "uppromote").strip().lower()
    if source not in ("uppromote", "goaffpro", "refersion", "collabs"):
        source = "uppromote"

    core.load_env_file(ENV_PATH)
    _refresh_license_env_from_file()
    license_guard.set_paths(BASE_DIR)
    ok_run, lic_err = license_guard.assert_can_start_pipeline(source)
    if not ok_run:
        return jsonify({"ok": False, "error": lic_err}), 400

    with STATE.lock:
        if STATE.running:
            return jsonify({"ok": False, "error": "Đang chạy sẵn, không thể bắt đầu thêm."}), 400
        STATE.running = True
        STATE.paused = False
        STATE.status = "Đang chạy"
        STATE.progress = 0
        STATE.logs = []
        with STATE.ack_lock:
            STATE.log_ack_seen = 0
        STATE.control = RunControl()
        STATE.worker = threading.Thread(target=_worker, args=(settings, min_traffic, filters, source), daemon=True)
        STATE.worker.start()
    return jsonify({"ok": True})


@app.post("/api/pause")
def api_pause():
    with STATE.lock:
        if not STATE.running or not STATE.control:
            return jsonify({"ok": False}), 400
        paused = STATE.control.toggle_pause()
        STATE.paused = paused
        STATE.status = "Tạm dừng" if paused else "Đang chạy"
    STATE.add_log("Đã tạm dừng." if paused else "Đã tiếp tục.")
    return jsonify({"ok": True, "paused": paused})


@app.post("/api/stop")
def api_stop():
    with STATE.lock:
        if not STATE.running or not STATE.control:
            return jsonify({"ok": False}), 400
        STATE.control.stop()
        STATE.status = "Đang dừng"
    STATE.add_log("Đang dừng...")
    return jsonify({"ok": True})


@app.post("/api/collabs-outside/reset")
def api_collabs_outside_reset():
    removed: list[str] = []
    errors: list[str] = []
    for p in (core._outside_cursor_path(), core._outside_dedup_path()):
        try:
            if p.exists():
                p.unlink()
                removed.append(p.name)
        except Exception as exc:
            errors.append(f"{p.name}: {exc}")
    if errors:
        return jsonify({"ok": False, "error": "; ".join(errors), "removed": removed}), 500
    return jsonify({"ok": True, "removed": removed})


def _open_collabs_tab_over_cdp(cdp_url: str, logs: list[str], label: str = "") -> None:
    """Kết nối CDP tạm thời, mở tab Collabs, đóng kết nối (không tắt Edge của user)."""
    prefix = f"{label} " if label else ""
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(cdp_url)
            try:
                context = browser.contexts[0] if browser.contexts else browser.new_context()
                page = context.new_page()
                try:
                    page.goto("https://collabs.shopify.com/", wait_until="domcontentloaded", timeout=45000)
                    logs.append(f"{prefix}Đã mở tab Collabs: {cdp_url}")
                except Exception as exc:
                    logs.append(f"{prefix}[Cảnh báo] Không mở được tab Collabs: {exc}")
                try:
                    page.close()
                except Exception:
                    pass
            finally:
                try:
                    browser.close()
                except Exception:
                    pass
    except Exception as exc:
        logs.append(f"{prefix}[Cảnh báo] Không kết nối CDP để mở tab: {exc}")


@app.post("/api/edge-cdp/start")
def api_edge_cdp_start():
    """
    Mở Edge CDP cho một hoặc nhiều tài khoản (tối đa 10).
    Body JSON (tuỳ chọn):
      account_indices: [1,2,3] — Tài khoản 1 → cổng 9222, TK2 → 9223, … (mặc định [1]).
    Mỗi tài khoản = profile Edge riêng, khớp Auto Apply đa tài khoản.
    """
    payload = request.get_json(force=True, silent=True) or {}
    raw = payload.get("account_indices")
    indices: list[int] = []
    if isinstance(raw, list) and len(raw) > 0:
        for x in raw:
            try:
                i = int(x)
            except (TypeError, ValueError):
                continue
            if 1 <= i <= 10:
                indices.append(i)
        indices = sorted(set(indices))
    if not indices:
        indices = [1]

    logs: list[str] = []
    edge_exe = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    opened: list[dict] = []
    host = "127.0.0.1"
    all_ok = True
    for acc in indices:
        try:
            port = port_for_collabs_account(acc)
        except ValueError as exc:
            logs.append(str(exc))
            all_ok = False
            opened.append({"account": acc, "port": None, "cdp_url": "", "ok": False, "error": str(exc)})
            continue
        udir = default_user_data_dir_for_port(port)
        cdp_url = cdp_url_for_collabs_account(acc, host=host)
        logs.append(f"Tài khoản {acc}: CDP {cdp_url} — profile {udir}")
        ok = ensure_edge_cdp_running(
            port=port,
            user_data_dir=udir,
            edge_exe=edge_exe,
            log=lambda m, a=acc: logs.append(f"[TK{a}] {m}"),
            wait_sec=12.0,
            host=host,
        )
        opened.append({"account": acc, "port": port, "cdp_url": cdp_url, "ok": ok})
        if not ok:
            all_ok = False
            continue
        _open_collabs_tab_over_cdp(cdp_url, logs, label=f"[TK{acc}]")

    if not all_ok:
        return _no_cache_json(
            {
                "ok": False,
                "error": "Một hoặc nhiều Edge không mở được. Xem logs.",
                "logs": logs,
                "opened": opened,
            }
        ), 500
    primary = cdp_url_for_collabs_account(indices[0], host=host)
    return _no_cache_json({"ok": True, "cdp_url": primary, "opened": opened, "logs": logs})


def _no_cache_json(data):
    resp = jsonify(data)
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.get("/api/status")
def api_status():
    with STATE.lock:
        return _no_cache_json(
            {
                "running": STATE.running,
                "paused": STATE.paused,
                "status": STATE.status,
                "progress": STATE.progress,
                "output_file": STATE.output_file,
            }
        )


@app.get("/api/logs")
def api_logs():
    since = int(request.args.get("since", "0"))
    logs, total = STATE.get_logs(since)
    return _no_cache_json({"logs": logs, "total": total})


@app.post("/api/logs/ack")
def api_logs_ack():
    """Client xác nhận đã hiển thị log tới mốc total (đồng bộ với /api/logs)."""
    payload = request.get_json(force=True) or {}
    try:
        seen = int(payload.get("seen_total", 0))
    except (TypeError, ValueError):
        seen = 0
    STATE.notify_log_displayed(seen)
    return jsonify({"ok": True})


@app.get("/api/results")
def api_results():
    files = []
    seen = set()
    globs = (
        list(BASE_DIR.glob("result-*.xlsx"))
        + list(BASE_DIR.glob("uppromote_*.xlsx"))
        + list(BASE_DIR.glob("goaffpro_*.xlsx"))
        + list(BASE_DIR.glob("refersion_*.xlsx"))
        + list(BASE_DIR.glob("collabs_*.xlsx"))
    )
    for p in sorted(globs, key=lambda x: x.stat().st_mtime, reverse=True):
        if p.name in seen:
            continue
        seen.add(p.name)
        files.append(
            {
                "name": p.name,
                "size": p.stat().st_size,
                "modified": int(p.stat().st_mtime),
            }
        )
    return jsonify({"files": files})


@app.get("/api/auto-collabs/files")
def api_auto_collabs_files():
    files = []
    for p in sorted(BASE_DIR.glob("collabs_import_*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
        files.append(
            {
                "name": p.name,
                "size": p.stat().st_size,
                "modified": int(p.stat().st_mtime),
            }
        )
    return jsonify({"files": files})


@app.post("/api/auto-collabs/import")
def api_auto_collabs_import():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Thiếu file upload."}), 400
    upl = request.files.get("file")
    if not upl:
        return jsonify({"ok": False, "error": "Thiếu file upload."}), 400
    original_name = str(getattr(upl, "filename", "") or "").strip()
    ext = Path(original_name).suffix.lower()
    if ext != ".xlsx":
        return jsonify({"ok": False, "error": "Chỉ hỗ trợ file .xlsx."}), 400

    now = datetime.now()
    safe_name = f"collabs_import_{now.day}-{now.month}-{now.year}_{now.hour}-{now.minute:02d}-{now.second:02d}.xlsx"
    full = BASE_DIR / safe_name
    try:
        upl.save(str(full))
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Không lưu được file: {exc}"}), 500

    try:
        links = auto_apply_core.extract_apply_links_from_xlsx(full, apply_mode="all")
    except Exception as exc:
        try:
            if full.exists():
                full.unlink()
        except OSError:
            pass
        return jsonify({"ok": False, "error": f"Không đọc được file Excel: {exc}"}), 400
    if not links:
        try:
            if full.exists():
                full.unlink()
        except OSError:
            pass
        return jsonify({"ok": False, "error": "Không tìm thấy link hợp lệ trong file Excel."}), 400

    return jsonify({"ok": True, "name": safe_name, "total_links": len(links)})


@app.get("/api/auto-collabs/template")
def api_auto_collabs_template():
    """
    Trả file Excel mẫu để user điền link đăng ký Collabs rồi import.
    Không lưu file lên đĩa: tạo workbook trong memory.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Thiếu thư viện openpyxl để tạo file mẫu: {exc}"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Auto Collabs"

    header = ["Trạng thái", "Link đăng ký"]
    ws.append(header)

    # Gợi ý: nếu user muốn chỉ apply hàng ĐẠT thì điền "ĐẠT" vào cột Trạng thái.
    ws.append(["ĐẠT", "https://collabs.shopify.com/brands/<...>/signup"])
    ws.append(["", "https://collabs.shopify.com/brands/<...>/signup"])

    # Style đơn giản cho header + highlight "ĐẠT"
    fill_head = PatternFill("solid", fgColor="1F2937")  # slate-800
    fill_head_font = "FFFFFF"
    for cell in ws[1]:
        try:
            cell.fill = fill_head
            cell.font = cell.font.copy(color=fill_head_font, bold=True)
        except Exception:
            pass
    fill_dat = PatternFill("solid", fgColor="DCFCE7")  # green-100
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        c = row[0]
        if str(c.value or "").strip().lower() == "đạt":
            try:
                c.fill = fill_dat
            except Exception:
                pass

    # set width
    try:
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 70
    except Exception:
        pass

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_name = "auto-collabs-template.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=safe_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
        conditional=False,
    )


def _allowed_export_basename(name: str) -> bool:
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if ext != "xlsx":
        return False
    return (
        name.startswith("result-")
        or name.startswith("uppromote_")
        or name.startswith("goaffpro_")
        or name.startswith("refersion_")
        or name.startswith("collabs_")
    )


def _safe_result_file_path(safe_name: str) -> Path | None:
    if not _allowed_export_basename(safe_name):
        return None
    base = BASE_DIR.resolve()
    full = (base / Path(safe_name).name).resolve()
    try:
        full.relative_to(base)
    except ValueError:
        return None
    return full


@app.get("/api/download/<path:filename>")
def api_download(filename: str):
    safe_name = Path(filename).name
    full = _safe_result_file_path(safe_name)
    if full is None:
        return jsonify({"ok": False, "error": "Invalid file"}), 400
    if not full.is_file():
        return jsonify({"ok": False, "error": "Not found"}), 404
    # str(path): PyInstaller/Windows ổn định hơn với đường dẫn UNC/unicode
    return send_file(
        str(full),
        as_attachment=True,
        download_name=safe_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
        conditional=False,
    )


@app.post("/api/results/delete")
def api_delete_result():
    payload = request.get_json(force=True) or {}
    name = (payload.get("name") or "").strip()
    safe_name = Path(name).name
    full = _safe_result_file_path(safe_name)
    if full is None:
        return jsonify({"ok": False, "error": "Invalid file"}), 400
    if not full.is_file():
        return jsonify({"ok": False, "error": "Not found"}), 404
    try:
        full.unlink()
    except OSError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": True})


@app.post("/api/auto-apply")
def api_auto_apply():
    if not auto_apply_collabs_enabled():
        return jsonify({"ok": False, "error": "Auto Apply Collabs đang tắt trên server."}), 403
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name") or "").strip()
    safe_name = Path(name).name
    full = _safe_result_file_path(safe_name)
    if full is None:
        return jsonify({"ok": False, "error": "Tên file không hợp lệ."}), 400
    if not full.is_file():
        return jsonify({"ok": False, "error": "Không tìm thấy file."}), 404

    profile_in = payload.get("profile") or {}
    profile = {
        "full_name": str(profile_in.get("full_name") or "").strip(),
        "first_name": str(profile_in.get("first_name") or "").strip(),
        "last_name": str(profile_in.get("last_name") or "").strip(),
        "email": str(profile_in.get("email") or "").strip(),
        "phone": str(profile_in.get("phone") or "").strip(),
        "website": str(profile_in.get("website") or "").strip(),
        "instagram": str(profile_in.get("instagram") or "").strip(),
        "tiktok": str(profile_in.get("tiktok") or "").strip(),
        "youtube": str(profile_in.get("youtube") or "").strip(),
        "message": str(profile_in.get("message") or "").strip(),
        "business_type": str(profile_in.get("business_type") or "").strip(),
        "dob": str(profile_in.get("dob") or "").strip(),
        "shipping_location": str(profile_in.get("shipping_location") or "").strip() or "United States",
        "purchase_before_choice": str(profile_in.get("purchase_before_choice") or "").strip() or "Yes",
        "identify": profile_in.get("identify") if isinstance(profile_in.get("identify"), list) else [],
        "brands_worked": str(profile_in.get("brands_worked") or "").strip(),
        "successful_partnership": str(profile_in.get("successful_partnership") or "").strip(),
        "content_inspires": str(profile_in.get("content_inspires") or "").strip(),
        "hope_gain": str(profile_in.get("hope_gain") or "").strip(),
        "how_found": str(profile_in.get("how_found") or "").strip(),
        "city_country": str(profile_in.get("city_country") or "").strip(),
        "demographic": str(profile_in.get("demographic") or "").strip(),
        "growth_strategy": str(profile_in.get("growth_strategy") or "").strip(),
        "children_age": str(profile_in.get("children_age") or "").strip(),
        "ugc_content": str(profile_in.get("ugc_content") or "").strip(),
        "content_ideas": str(profile_in.get("content_ideas") or "").strip(),
        "why_fit": str(profile_in.get("why_fit") or "").strip(),
        "purchase_love": str(profile_in.get("purchase_love") or "").strip(),
        "why_join": str(profile_in.get("why_join") or "").strip(),
        "generic_short": str(profile_in.get("generic_short") or "").strip(),
        "generic_long": str(profile_in.get("generic_long") or "").strip(),
    }
    # chuẩn hóa identify
    if not isinstance(profile["identify"], list):
        profile["identify"] = []
    profile["identify"] = [str(x or "").strip() for x in profile["identify"] if str(x or "").strip()]
    if not profile["identify"]:
        profile["identify"] = ["Prefer not to say"]
    if not profile["full_name"] and not profile["first_name"] and not profile["email"]:
        return (
            jsonify({"ok": False, "error": "Cần ít nhất 1 trong các trường: Họ tên, First name hoặc Email."}),
            400,
        )

    auto_submit = bool(payload.get("auto_submit"))
    use_cdp = bool(payload.get("use_cdp"))
    cdp_url = str(payload.get("cdp_url") or "").strip() or "http://127.0.0.1:9222"
    if not use_cdp:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": (
                        "Auto Apply hiện chỉ chạy trên trình duyệt thật đã login (CDP). "
                        "Vui lòng bật tùy chọn dùng trình duyệt đang mở."
                    ),
                }
            ),
            400,
        )
    # Cưỡng chế: luôn bắt buộc login Shopify Collabs trước khi chạy Auto Apply.
    login_first = True
    apply_mode = str(payload.get("apply_mode") or "only_dat").strip() or "only_dat"
    try:
        row_start = int(payload.get("row_start")) if str(payload.get("row_start") or "").strip() else None
    except (TypeError, ValueError):
        row_start = None
    try:
        row_end = int(payload.get("row_end")) if str(payload.get("row_end") or "").strip() else None
    except (TypeError, ValueError):
        row_end = None

    links = auto_apply_core.extract_apply_links_from_xlsx(
        full,
        apply_mode=apply_mode,
        row_start=row_start,
        row_end=row_end,
    )
    if not links:
        return jsonify({"ok": False, "error": "Không tìm thấy cột/link apply trong file."}), 400

    # Giữ endpoint cũ để tương thích: chạy dạng blocking (không có nút Hủy).
    logs: list[str] = []
    try:
        host, port = cdp_url_host_port(cdp_url)
        udir = default_user_data_dir_for_port(port)
        ensure_edge_cdp_running(
            port=port,
            user_data_dir=udir,
            edge_exe=r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            log=lambda m: logs.append(str(m)),
            host=host,
        )
        result = auto_apply_core.run_auto_apply(
            links=links,
            profile=profile,
            auto_submit=auto_submit,
            cdp_url=cdp_url,
            login_first=login_first,
            log=lambda m: logs.append(str(m)),
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc), "logs": logs}), 500
    return jsonify({"ok": True, "result": result, "logs": logs})


_EDGE_EXE_AUTO = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"


def _normalize_parallel_slots(payload: dict) -> list[dict]:
    default_cdp = str(payload.get("cdp_url") or "").strip() or "http://127.0.0.1:9222"
    raw = payload.get("parallel_slots")
    if isinstance(raw, list) and len(raw) > 0:
        out: list[dict] = []
        for it in raw:
            if not isinstance(it, dict):
                continue
            out.append(
                {
                    "cdp_url": str(it.get("cdp_url") or "").strip() or default_cdp,
                    "edge_user_data_dir": str(it.get("edge_user_data_dir") or "").strip(),
                    "email_override": str(it.get("email_override") or "").strip(),
                }
            )
        if not out:
            return [{"cdp_url": default_cdp, "edge_user_data_dir": "", "email_override": ""}]
        return out
    top_ud = str(payload.get("edge_user_data_dir") or "").strip()
    return [{"cdp_url": default_cdp, "edge_user_data_dir": top_ud, "email_override": ""}]


def _normalize_account_mode(raw) -> str:
    s = str(raw or "").strip().lower()
    if s in ("multi", "da_tai_khoan", "many"):
        return "multi"
    return "single"


def _normalize_collabs_account_indices_list(raw_list) -> tuple[list[int], str]:
    """Giữ thứ tự, bỏ trùng; mỗi phần tử 1..EDGE_CDP_ACCOUNT_MAX."""
    if not isinstance(raw_list, list):
        return [], "collabs_account_indices phải là mảng số."
    out: list[int] = []
    seen: set[int] = set()
    for x in raw_list:
        try:
            i = int(x)
        except (TypeError, ValueError):
            return [], "Mỗi phần tử collabs_account_indices phải là số nguyên."
        if i < 1 or i > EDGE_CDP_ACCOUNT_MAX:
            return [], f"Tài khoản phải từ 1 đến {EDGE_CDP_ACCOUNT_MAX}."
        if i not in seen:
            seen.add(i)
            out.append(i)
    return out, ""


def _parse_multi_collabs_account_spec(spec: str) -> list[int]:
    """
    - Có dấu phẩy: danh sách, vd 1,4,7 (thứ tự giữ nguyên, bỏ trùng).
    - Một dải a-b: vd 1-5 (hai đầu 1..MAX).
    - Chỉ số nguyên n (không phẩy, không gạch): legacy — tài khoản 1..n, 2 <= n <= MAX.
    """
    s = str(spec or "").strip()
    if not s:
        return []
    if "," in s:
        parts = [p.strip() for p in s.split(",") if str(p).strip()]
        out: list[int] = []
        seen: set[int] = set()
        for p in parts:
            try:
                i = int(p, 10)
            except ValueError:
                return []
            if i < 1 or i > EDGE_CDP_ACCOUNT_MAX:
                return []
            if i not in seen:
                seen.add(i)
                out.append(i)
        return out
    m = re.match(r"^(\d+)\s*-\s*(\d+)$", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if a > b:
            a, b = b, a
        if a < 1 or b > EDGE_CDP_ACCOUNT_MAX:
            return []
        return list(range(a, b + 1))
    if s.isdigit():
        n = int(s, 10)
        if n < 2 or n > EDGE_CDP_ACCOUNT_MAX:
            return []
        return list(range(1, n + 1))
    return []


def _collabs_account_indices_from_payload(payload: dict, legacy_count: int) -> tuple[list[int], str]:
    """Chọn danh sách tài khoản Collabs cho chế độ multi (lần lượt)."""
    raw_list = payload.get("collabs_account_indices")
    if isinstance(raw_list, list) and len(raw_list) > 0:
        out, err = _normalize_collabs_account_indices_list(raw_list)
        if err:
            return [], err
        if len(out) < 2:
            return [], "Đa tài khoản cần ít nhất 2 tài khoản trong collabs_account_indices."
        return out, ""
    spec = str(
        payload.get("multi_account_spec") or payload.get("multi_collabs_account_spec") or ""
    ).strip()
    if spec:
        out = _parse_multi_collabs_account_spec(spec)
        if not out:
            return (
                [],
                "Tài khoản không hợp lệ. Dùng dải (vd: 1-5), danh sách (vd: 1,4,7), hoặc số 2-10 (tài khoản 1→n).",
            )
        if len(out) < 2:
            return [], "Đa tài khoản cần ít nhất 2 tài khoản (vd: 1-2 hoặc 1,2)."
        return out, ""
    n = int(legacy_count)
    n = max(2, min(EDGE_CDP_ACCOUNT_MAX, n))
    return list(range(1, n + 1)), ""


def _normalize_collabs_account_profile_map(payload: dict, indices: list[int]) -> tuple[dict[int, dict], str]:
    raw = payload.get("collabs_account_profile_map")
    if raw is None:
        # Backward-compatible: map email cũ { "1": "a@x.com" }
        raw = payload.get("collabs_account_email_map")
    if raw is None:
        return {}, ""
    if not isinstance(raw, dict):
        return {}, "collabs_account_profile_map phải là object dạng {\"1\":{\"email\":\"...\"}}."
    allowed = set(int(x) for x in (indices or []))
    out: dict[int, dict] = {}

    def _norm(v) -> str:
        t = str(v or "").strip()
        if not t:
            return ""
        if t.lower() == "no":
            return ""
        return t

    for k, v in raw.items():
        try:
            acc = int(k)
        except (TypeError, ValueError):
            return {}, f"Key tài khoản không hợp lệ trong collabs_account_profile_map: {k}"
        if acc < 1 or acc > EDGE_CDP_ACCOUNT_MAX:
            return {}, f"Tài khoản trong collabs_account_profile_map phải từ 1 đến {EDGE_CDP_ACCOUNT_MAX}."
        if allowed and acc not in allowed:
            continue
        if isinstance(v, dict):
            em = _norm(v.get("email"))
            full_name = _norm(v.get("full_name"))
            phone = _norm(v.get("phone"))
            website = _norm(v.get("website"))
            instagram = _norm(v.get("instagram"))
        else:
            # Legacy string value = email
            em = _norm(v)
            full_name = ""
            phone = ""
            website = ""
            instagram = ""
        if em and "@" not in em:
            return {}, f"Email không hợp lệ cho TK{acc}: {em}"
        prof = {
            "email": em,
            "full_name": full_name,
            "phone": phone,
            "website": website,
            "instagram": instagram,
        }
        if any(str(x).strip() for x in prof.values()):
            out[acc] = prof
    return out, ""


def _slots_from_collabs_account_indices(indices: list[int], profile_map: dict[int, dict] | None = None) -> list[dict]:
    host = "127.0.0.1"
    pmap = profile_map or {}
    return [
        {
            "cdp_url": cdp_url_for_collabs_account(acc, host=host),
            "edge_user_data_dir": "",
            "email_override": str((pmap.get(acc) or {}).get("email") or ""),
            "profile_override": dict(pmap.get(acc) or {}),
        }
        for acc in indices
    ]


def _profile_for_slot(base: dict, email_override: str, profile_override: Optional[dict] = None) -> dict:
    p = dict(base) if isinstance(base, dict) else {}
    em = str(email_override or "").strip()
    if em:
        p["email"] = em
    ov = dict(profile_override) if isinstance(profile_override, dict) else {}
    for k in ("email", "full_name", "phone", "website", "instagram"):
        v = str(ov.get(k) or "").strip()
        if v and v.lower() != "no":
            p[k] = v
    return p


def _append_auto_apply_run_to_history(file_name: str, started_local: datetime, profile: dict, result: dict) -> None:
    submitted_items = result.get("submitted_items") if isinstance(result, dict) else []
    if not isinstance(submitted_items, list):
        submitted_items = []
    attempted_items = result.get("attempted_items") if isinstance(result, dict) else []
    if not isinstance(attempted_items, list):
        attempted_items = []
    entry: dict = {
        "file": str(file_name or ""),
        "started_at": started_local.isoformat(timespec="seconds"),
        "started_at_display": started_local.strftime("%d/%m/%Y %H:%M:%S"),
        "email": str(profile.get("email") or "").strip(),
        "submitted_count": int(result.get("submitted") or 0) if isinstance(result, dict) else 0,
        "submitted_items": [
            {
                "brand": str((it or {}).get("brand") or "").strip(),
                "email": str((it or {}).get("email") or "").strip() or str(profile.get("email") or "").strip(),
                "link": str((it or {}).get("link") or "").strip(),
            }
            for it in submitted_items
            if isinstance(it, dict)
        ],
        "attempted_items": [
            {
                "brand": str((it or {}).get("brand") or "").strip(),
                "domain": str((it or {}).get("domain") or "").strip(),
                "email": str((it or {}).get("email") or "").strip() or str(profile.get("email") or "").strip(),
                "link": str((it or {}).get("link") or "").strip(),
                "submitted": bool((it or {}).get("submitted")),
                "note": str((it or {}).get("note") or "").strip(),
            }
            for it in attempted_items
            if isinstance(it, dict)
        ],
    }
    if isinstance(result, dict) and result.get("parallel"):
        entry["parallel"] = True
        entry["batch_id"] = str(result.get("batch_id") or "")
        pj = result.get("parallel_jobs")
        if isinstance(pj, dict):
            entry["parallel_jobs"] = pj
    if isinstance(result, dict) and result.get("sequential_multi"):
        entry["sequential_multi"] = True
        entry["accounts"] = int(result.get("accounts") or 0)
    _append_auto_apply_history(entry)


def _merge_parallel_auto_apply_results(
    *,
    links: list[str],
    results: dict[str, dict],
    errors: dict[str, str],
    batch_id: str,
    job_order: list[str],
    parallel_link_split: str = "round_robin",
) -> dict:
    filled = 0
    submitted = 0
    submitted_items: list[dict] = []
    attempted_items: list[dict] = []
    parallel_jobs: dict[str, dict] = {}
    for jid in job_order:
        r = results.get(jid)
        err = errors.get(jid, "")
        if isinstance(r, dict):
            filled += int(r.get("filled") or 0)
            submitted += int(r.get("submitted") or 0)
            for it in r.get("submitted_items") or []:
                if isinstance(it, dict):
                    submitted_items.append(dict(it))
            for it in r.get("attempted_items") or []:
                if isinstance(it, dict):
                    attempted_items.append(dict(it))
            parallel_jobs[jid] = {
                "ok": True,
                "error": "",
                "filled": int(r.get("filled") or 0),
                "submitted": int(r.get("submitted") or 0),
            }
        else:
            parallel_jobs[jid] = {"ok": False, "error": err or "Lỗi", "filled": 0, "submitted": 0}
    return {
        "total": len(links),
        "filled": filled,
        "submitted": submitted,
        "submitted_items": submitted_items,
        "attempted_items": attempted_items,
        "parallel": True,
        "batch_id": batch_id,
        "parallel_jobs": parallel_jobs,
        "parallel_link_split": str(parallel_link_split or "round_robin"),
    }


def _normalize_parallel_link_split(raw) -> str:
    s = str(raw or "").strip().lower()
    if s in ("sequential", "chunks", "contiguous", "theo_thu_tu", "order", "lan_luot"):
        return "sequential"
    return "round_robin"


def _split_links_for_parallel_jobs(links: list[str], n_slots: int, mode: str) -> list[list[str]]:
    """
    Chia links thành đúng n_slots phần.
    - round_robin: slot i nhận links[i::n] (xen kẽ).
    - sequential: slot i nhận một khối liên tiếp theo thứ tự file (slot 0 = đầu list, …).
    """
    if n_slots <= 0:
        return []
    mode_n = _normalize_parallel_link_split(mode)
    L = len(links)
    if mode_n == "sequential":
        out: list[list[str]] = []
        for i in range(n_slots):
            a = (i * L) // n_slots
            b = ((i + 1) * L) // n_slots
            out.append(links[a:b])
        return out
    return [links[i::n_slots] for i in range(n_slots)]


def _parallel_auto_apply_supervisor(
    *,
    parallel_slots: list[dict],
    links: list[str],
    base_profile: dict,
    auto_submit: bool,
    login_first: bool,
    file_name: str,
    started_local: datetime,
    parallel_link_split: str = "round_robin",
) -> None:
    st = AUTO_APPLY_STATE
    n = len(parallel_slots)
    split_mode = _normalize_parallel_link_split(parallel_link_split)
    batch_id = (st.batch_id or "").strip() or uuid.uuid4().hex[:10]
    ctx = multiprocessing.get_context("spawn")
    log_q: multiprocessing.Queue = ctx.Queue()
    res_q: multiprocessing.Queue = ctx.Queue()
    children: list[tuple[str, multiprocessing.Process, AutoApplyJob]] = []
    job_order: list[str] = []
    link_slices = _split_links_for_parallel_jobs(links, n, split_mode)
    try:
        split_label = "theo thứ tự (khối liên tiếp)" if split_mode == "sequential" else "round-robin (xen kẽ)"
        st.add_log(f"Khởi động {n} tiến trình Auto Apply song song (batch {batch_id}), chia link: {split_label}.")
        for i, slot in enumerate(parallel_slots):
            jid = f"j{i + 1}-{uuid.uuid4().hex[:6]}"
            sub_links = link_slices[i] if i < len(link_slices) else []
            if not sub_links:
                st.add_log(f"[slot {i + 1}] Không có link sau khi chia — bỏ slot.")
                continue
            prof = _profile_for_slot(
                base_profile,
                str(slot.get("email_override") or ""),
                profile_override=slot.get("profile_override"),
            )
            cdp = str(slot.get("cdp_url") or "").strip()
            _, port = cdp_url_host_port(cdp)
            udir = str(slot.get("edge_user_data_dir") or "").strip() or default_user_data_dir_for_port(port)
            stop_ev = ctx.Event()
            proc = ctx.Process(
                target=run_auto_apply_in_subprocess,
                kwargs={
                    "job_id": jid,
                    "links": sub_links,
                    "profile": prof,
                    "auto_submit": auto_submit,
                    "cdp_url": cdp,
                    "login_first": login_first,
                    "file_name": file_name,
                    "edge_user_data_dir": udir,
                    "log_queue": log_q,
                    "stop_event": stop_ev,
                    "result_queue": res_q,
                },
            )
            job = AutoApplyJob(
                jid,
                cdp_url=cdp,
                edge_user_data_dir=udir,
                profile_email=str(prof.get("email") or ""),
                links=sub_links,
                process=proc,
                stop_event=stop_ev,
            )
            children.append((jid, proc, job))
            job_order.append(jid)
        with st.lock:
            st.mp_ctx = ctx
            st.shared_log_queue = log_q
            st.shared_result_queue = res_q
            st.jobs = {jid: job for jid, _, job in children}
        for jid, proc, _job in children:
            proc.start()
        results: dict[str, dict] = {}
        errors: dict[str, str] = {}
        expected = {jid for jid, _, _ in children}
        if not expected:
            with st.lock:
                st.status = "Error"
                st.error = "Không khởi chạy được job nào (kiểm tra parallel_slots / link)."
            return
        while True:
            if st.stop_event.is_set():
                for _jid, _proc, job in children:
                    job.stop_event.set()
            while True:
                try:
                    jid, msg = log_q.get_nowait()
                except pyqueue.Empty:
                    break
                with st.lock:
                    j = st.jobs.get(jid)
                    if j:
                        j.logs.append(str(msg))
                        if len(j.logs) > 500:
                            j.logs = j.logs[-500:]
                st.add_log(f"[{jid}] {msg}")
            while True:
                try:
                    item = res_q.get_nowait()
                except pyqueue.Empty:
                    break
                if not isinstance(item, dict):
                    continue
                jid2 = str(item.get("job_id") or "").strip()
                if item.get("ok"):
                    r = item.get("result")
                    results[jid2] = r if isinstance(r, dict) else {}
                    with st.lock:
                        if jid2 in st.jobs:
                            st.jobs[jid2].result = results[jid2]
                            st.jobs[jid2].status = "Done"
                else:
                    err = str(item.get("error") or "").strip() or "Lỗi không xác định."
                    errors[jid2] = err
                    with st.lock:
                        if jid2 in st.jobs:
                            st.jobs[jid2].error = err
                            st.jobs[jid2].status = "Error"
            all_dead = all(not p.is_alive() for _jid, p, _ in children)
            for jid, p, job in children:
                if not p.is_alive() and job.status == "Running" and jid not in results and jid not in errors:
                    errors[jid] = "Tiến trình kết thúc không trả kết quả."
                    job.status = "Error"
                    job.error = errors[jid]
                    st.add_log(f"[{jid}] {errors[jid]}")
            got = set(results) | set(errors)
            if all_dead and expected.issubset(got):
                break
            time.sleep(0.12)
        merged = _merge_parallel_auto_apply_results(
            links=links,
            results=results,
            errors=errors,
            batch_id=batch_id,
            job_order=job_order,
            parallel_link_split=split_mode,
        )
        with st.lock:
            st.result = merged
            if errors and not results:
                st.status = "Error"
                st.error = "; ".join(f"{k}: {v}" for k, v in sorted(errors.items()))
            else:
                st.status = "Done"
                if errors and results:
                    st.error = "Một số slot lỗi: " + "; ".join(f"{k}: {v}" for k, v in sorted(errors.items()))
        _append_auto_apply_run_to_history(file_name, started_local, base_profile, merged)
    except Exception as exc:
        with st.lock:
            st.error = str(exc)
            st.status = "Error"
        st.add_log(f"LỖI (parallel): {exc}")
    finally:
        for _jid, p, _job in children:
            if p.is_alive():
                p.join(timeout=4)
            if p.is_alive():
                try:
                    p.terminate()
                except Exception:
                    pass
        with st.lock:
            st.running = False


def _sequential_multi_apply_worker(
    *,
    slots: list[dict],
    links: list[str],
    base_profile: dict,
    auto_submit: bool,
    login_first: bool,
    file_name: str,
    started_local: datetime,
) -> None:
    """
    Đa tài khoản lần lượt: mỗi slot chạy lại toàn bộ danh sách link (cùng file brand),
    sau khi chờ login Collabs (login_first) trên Edge của slot đó.
    """
    st = AUTO_APPLY_STATE
    n = len(slots)
    total_filled = 0
    total_submitted = 0
    submitted_items: list[dict] = []
    attempted_items: list[dict] = []
    per_slot: list[dict] = []
    L = len(links)
    try:
        st.add_log(
            f"Đa tài khoản (lần lượt): {n} phiên Edge — mỗi tài khoản apply lại cả {L} link trong file."
        )
        for i, slot in enumerate(slots):
            if st.stop_event.is_set():
                st.add_log("Đã nhận lệnh hủy — dừng trước khi sang tài khoản kế.")
                break
            if L == 0:
                break
            sub_links = list(links)
            cdp = str(slot.get("cdp_url") or "").strip()
            _, port = cdp_url_host_port(cdp)
            udir = str(slot.get("edge_user_data_dir") or "").strip() or default_user_data_dir_for_port(port)
            prof = _profile_for_slot(
                base_profile,
                str(slot.get("email_override") or ""),
                profile_override=slot.get("profile_override"),
            )
            st.add_log(
                f"=== Tài khoản {i + 1}/{n} — CDP {cdp} — {L} link (lặp lại cả danh sách) — "
                "nếu chưa login Collabs, hãy đăng nhập trong cửa sổ Edge vừa mở ==="
            )
            st.add_log("Đang kiểm tra/mở Edge CDP…")
            host, p = cdp_url_host_port(cdp)
            ensure_edge_cdp_running(
                port=p,
                user_data_dir=udir,
                edge_exe=_EDGE_EXE_AUTO,
                log=st.add_log,
                host=host,
            )
            st.add_log(f"Bắt đầu Auto Apply cho tài khoản {i + 1}/{n}…")
            result = auto_apply_core.run_auto_apply(
                links=sub_links,
                profile=prof,
                auto_submit=auto_submit,
                cdp_url=cdp,
                login_first=login_first,
                should_stop=lambda: st.stop_event.is_set(),
                log=st.add_log,
            )
            if not isinstance(result, dict):
                result = {}
            fi = int(result.get("filled") or 0)
            su = int(result.get("submitted") or 0)
            total_filled += fi
            total_submitted += su
            for it in result.get("submitted_items") or []:
                if isinstance(it, dict):
                    submitted_items.append(dict(it))
            for it in result.get("attempted_items") or []:
                if isinstance(it, dict):
                    attempted_items.append(dict(it))
            per_slot.append(
                {
                    "index": i + 1,
                    "cdp_url": cdp,
                    "edge_user_data_dir": udir,
                    "email": str(prof.get("email") or ""),
                    "links": L,
                    "filled": fi,
                    "submitted": su,
                    "skipped": False,
                }
            )
            st.add_log(f"--- Xong tài khoản {i + 1}/{n}: đã điền {fi}, đã submit {su} ---")
        merged = {
            "total": L,
            "total_brand_runs": L * n,
            "filled": total_filled,
            "submitted": total_submitted,
            "submitted_items": submitted_items,
            "attempted_items": attempted_items,
            "sequential_multi": True,
            "accounts": n,
            "per_slot": per_slot,
            "batch_id": str(st.batch_id or ""),
        }
        with st.lock:
            st.result = merged
            st.status = "Done"
        _append_auto_apply_run_to_history(file_name, started_local, base_profile, merged)
    except Exception as exc:
        with st.lock:
            st.error = str(exc)
            st.status = "Error"
        st.add_log(f"LỖI (đa tài khoản lần lượt): {exc}")
    finally:
        with st.lock:
            st.running = False


def _auto_apply_worker(
    *,
    links: list[str],
    profile: dict,
    auto_submit: bool,
    cdp_url: str,
    login_first: bool,
    file_name: str,
    edge_user_data_dir: str = "",
):
    st = AUTO_APPLY_STATE
    started_local = datetime.now()
    try:
        st.add_log("Đang kiểm tra/mở Edge CDP…")
        host, port = cdp_url_host_port(cdp_url)
        udir = str(edge_user_data_dir or "").strip() or default_user_data_dir_for_port(port)
        ensure_edge_cdp_running(
            port=port,
            user_data_dir=udir,
            edge_exe=_EDGE_EXE_AUTO,
            log=st.add_log,
            host=host,
        )
        st.add_log("Bắt đầu Auto Apply…")
        result = auto_apply_core.run_auto_apply(
            links=links,
            profile=profile,
            auto_submit=auto_submit,
            cdp_url=cdp_url,
            login_first=login_first,
            should_stop=lambda: st.stop_event.is_set(),
            log=st.add_log,
        )
        with st.lock:
            st.result = result
            st.status = "Done"
        _append_auto_apply_run_to_history(file_name, started_local, profile, result)
    except Exception as exc:
        with st.lock:
            st.error = str(exc)
            st.status = "Error"
        st.add_log(f"LỖI: {exc}")
    finally:
        with st.lock:
            st.running = False


def _auto_apply_jobs_snapshot_unlocked(st: AutoApplyState) -> dict[str, dict]:
    snap: dict[str, dict] = {}
    for jid, j in st.jobs.items():
        snap[jid] = {
            "cdp_url": j.cdp_url,
            "edge_user_data_dir": j.edge_user_data_dir,
            "email": j.profile_email,
            "links": len(j.links),
            "status": j.status,
            "error": j.error,
            "logs_tail": list(j.logs[-80:]),
        }
    return snap


@app.post("/api/auto-apply/start")
def api_auto_apply_start():
    if not auto_apply_collabs_enabled():
        return jsonify({"ok": False, "error": "Auto Apply Collabs đang tắt trên server."}), 403
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name") or "").strip()
    safe_name = Path(name).name
    full = _safe_result_file_path(safe_name)
    if full is None:
        return jsonify({"ok": False, "error": "Tên file không hợp lệ."}), 400
    if not full.is_file():
        return jsonify({"ok": False, "error": "Không tìm thấy file."}), 404

    profile_in = payload.get("profile") or {}
    profile = profile_in if isinstance(profile_in, dict) else {}
    auto_submit = bool(payload.get("auto_submit"))
    use_cdp = bool(payload.get("use_cdp"))
    cdp_url = str(payload.get("cdp_url") or "").strip() or "http://127.0.0.1:9222"
    if not use_cdp:
        return jsonify({"ok": False, "error": "Auto Apply hiện chỉ chạy qua CDP (trình duyệt đang mở)."}), 400
    login_first = True
    apply_mode = str(payload.get("apply_mode") or "only_dat").strip() or "only_dat"
    try:
        row_start = int(payload.get("row_start")) if str(payload.get("row_start") or "").strip() else None
    except (TypeError, ValueError):
        row_start = None
    try:
        row_end = int(payload.get("row_end")) if str(payload.get("row_end") or "").strip() else None
    except (TypeError, ValueError):
        row_end = None

    links = auto_apply_core.extract_apply_links_from_xlsx(
        full,
        apply_mode=apply_mode,
        row_start=row_start,
        row_end=row_end,
    )
    if not links:
        return jsonify({"ok": False, "error": "Không tìm thấy cột/link apply trong file."}), 400

    DEFAULT_CDP = "http://127.0.0.1:9222"
    payload_cdp = str(payload.get("cdp_url") or "").strip() or DEFAULT_CDP

    parallel_link_split = _normalize_parallel_link_split(payload.get("parallel_link_split"))
    explicit_account_mode = "account_mode" in payload
    account_mode = _normalize_account_mode(payload.get("account_mode"))
    multi_parallel = bool(payload.get("multi_parallel"))
    try:
        multi_browser_count = int(payload.get("multi_browser_count"))
    except (TypeError, ValueError):
        multi_browser_count = 2
    multi_browser_count = max(2, min(EDGE_CDP_ACCOUNT_MAX, multi_browser_count))

    raw_slots = payload.get("parallel_slots")
    has_custom_slots = isinstance(raw_slots, list) and len(raw_slots) >= 2

    slots: list[dict] = []
    use_parallel_supervisor = False
    use_sequential_multi = False

    if explicit_account_mode and account_mode == "multi":
        indices, spec_err = _collabs_account_indices_from_payload(payload, multi_browser_count)
        if spec_err:
            return jsonify({"ok": False, "error": spec_err}), 400
        profile_map, profile_map_err = _normalize_collabs_account_profile_map(payload, indices)
        if profile_map_err:
            return jsonify({"ok": False, "error": profile_map_err}), 400
        slots = _slots_from_collabs_account_indices(indices, profile_map=profile_map)
        use_sequential_multi = True
    elif explicit_account_mode and account_mode == "single":
        slots = [{"cdp_url": DEFAULT_CDP, "edge_user_data_dir": "", "email_override": ""}]
    elif not explicit_account_mode and has_custom_slots and multi_parallel:
        account_mode = "multi"
        slots = _normalize_parallel_slots(payload)
        use_parallel_supervisor = True
    elif not explicit_account_mode and has_custom_slots:
        account_mode = "multi"
        slots = _normalize_parallel_slots(payload)
        use_sequential_multi = True
    else:
        slots = [
            {
                "cdp_url": payload_cdp,
                "edge_user_data_dir": str(payload.get("edge_user_data_dir") or "").strip(),
                "email_override": "",
            }
        ]

    if use_parallel_supervisor and len(slots) < 2:
        return jsonify({"ok": False, "error": "Chế độ song song cần ít nhất 2 slot CDP trong parallel_slots."}), 400
    if use_sequential_multi and len(slots) < 2:
        return jsonify({"ok": False, "error": "Đa tài khoản cần ít nhất 2 phiên Edge."}), 400

    st = AUTO_APPLY_STATE
    with st.lock:
        if st.running:
            return jsonify({"ok": False, "error": "Auto Apply đang chạy sẵn."}), 400
        st.reset()
        st.running = True
        st.status = "Running"
        st.current_file = safe_name
        st.batch_id = uuid.uuid4().hex[:12]
        st.account_mode = account_mode
        st.sequential_multi = use_sequential_multi

        if use_parallel_supervisor:
            st.parallel_link_split = parallel_link_split
            st.worker = threading.Thread(
                target=_parallel_auto_apply_supervisor,
                kwargs={
                    "parallel_slots": slots,
                    "links": links,
                    "base_profile": profile,
                    "auto_submit": auto_submit,
                    "login_first": login_first,
                    "file_name": safe_name,
                    "started_local": datetime.now(),
                    "parallel_link_split": parallel_link_split,
                },
                daemon=True,
            )
            st.worker.start()
            return jsonify(
                {
                    "ok": True,
                    "total_links": len(links),
                    "parallel": True,
                    "parallel_jobs": len(slots),
                    "batch_id": st.batch_id,
                    "parallel_link_split": parallel_link_split,
                    "account_mode": "multi",
                    "multi_parallel": True,
                }
            )

        if use_sequential_multi:
            st.parallel_link_split = "sequential"
            st.worker = threading.Thread(
                target=_sequential_multi_apply_worker,
                kwargs={
                    "slots": slots,
                    "links": links,
                    "base_profile": profile,
                    "auto_submit": auto_submit,
                    "login_first": login_first,
                    "file_name": safe_name,
                    "started_local": datetime.now(),
                },
                daemon=True,
            )
            st.worker.start()
            return jsonify(
                {
                    "ok": True,
                    "total_links": len(links),
                    "parallel": False,
                    "sequential_multi": True,
                    "accounts": len(slots),
                    "batch_id": st.batch_id,
                    "account_mode": "multi",
                    "multi_parallel": False,
                }
            )

        slot0 = slots[0]
        one_profile = _profile_for_slot(
            profile,
            str(slot0.get("email_override") or ""),
            profile_override=slot0.get("profile_override"),
        )
        one_cdp = str(slot0.get("cdp_url") or "").strip() or cdp_url
        one_udir = str(slot0.get("edge_user_data_dir") or "").strip()
        st.parallel_link_split = ""
        st.worker = threading.Thread(
            target=_auto_apply_worker,
            kwargs={
                "links": links,
                "profile": one_profile,
                "auto_submit": auto_submit,
                "cdp_url": one_cdp,
                "login_first": login_first,
                "file_name": safe_name,
                "edge_user_data_dir": one_udir,
            },
            daemon=True,
        )
        st.worker.start()
        return jsonify(
            {
                "ok": True,
                "total_links": len(links),
                "parallel": False,
                "sequential_multi": False,
                "batch_id": st.batch_id,
                "account_mode": account_mode,
            }
        )


@app.post("/api/auto-apply/stop")
def api_auto_apply_stop():
    if not auto_apply_collabs_enabled():
        return jsonify({"ok": False, "error": "Auto Apply Collabs đang tắt trên server."}), 403
    st = AUTO_APPLY_STATE
    with st.lock:
        if not st.running:
            return jsonify({"ok": False, "error": "Auto Apply không đang chạy."}), 400
        st.stop_event.set()
        st.status = "Stopping"
        jobs_copy = list(st.jobs.values())
    for job in jobs_copy:
        try:
            job.stop_event.set()
        except Exception:
            pass
    st.add_log("Đang hủy Auto Apply…")
    return jsonify({"ok": True})


@app.get("/api/auto-apply/status")
def api_auto_apply_status():
    if not auto_apply_collabs_enabled():
        return _no_cache_json(
            {
                "running": False,
                "status": "Disabled",
                "file": "",
                "logs": [],
                "result": None,
                "error": "",
                "parallel": False,
                "jobs": {},
                "batch_id": "",
                "parallel_link_split": "",
                "account_mode": "",
                "sequential_multi": False,
            }
        )
    st = AUTO_APPLY_STATE
    with st.lock:
        jobs_snap = _auto_apply_jobs_snapshot_unlocked(st)
        parallel = len(jobs_snap) > 0
        return _no_cache_json(
            {
                "running": st.running,
                "status": st.status,
                "file": st.current_file,
                "logs": list(st.logs[-200:]),
                "result": st.result,
                "error": st.error,
                "parallel": parallel,
                "jobs": jobs_snap,
                "batch_id": st.batch_id,
                "parallel_link_split": str(st.parallel_link_split or ""),
                "account_mode": str(st.account_mode or ""),
                "sequential_multi": bool(st.sequential_multi),
            }
        )


@app.get("/api/auto-apply/history")
def api_auto_apply_history():
    if not auto_apply_collabs_enabled():
        return _no_cache_json({"items": []})
    file_name = str(request.args.get("name") or "").strip()
    safe_name = Path(file_name).name if file_name else ""
    with AUTO_APPLY_HISTORY_LOCK:
        items = _load_auto_apply_history()
    if safe_name:
        items = [it for it in items if str((it or {}).get("file") or "") == safe_name]
    return _no_cache_json({"items": items[:100]})


if __name__ == "__main__":
    multiprocessing.freeze_support()
    # threaded=True: worker chạy pipeline không chặn request /api/logs (log theo thời gian thực)
    app.run(host="127.0.0.1", port=5050, debug=False, threaded=True)
